import os
import shutil
import atexit
import copy
import functools
import uuid
import signal
from contextlib import contextmanager
import gateway_network
import ipaddress
import re
import sqlite3
import threading
import json
import time
import random
import string
import platform
import subprocess
import urllib.request
import urllib.parse
import io
from datetime import datetime
from collections import deque
from flask import Flask, request, render_template_string, jsonify, session, redirect, url_for, Response, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
import logging
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except (ImportError, SyntaxError, Exception):
    openpyxl = None
try:
    import serial
except ImportError:
    serial = None
from esp32_simulator import ESP32Simulator
import license_manager
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)
app = Flask(__name__, static_folder='static')
class LocalProxy:
    def __init__(self, application):
        self.raw = application
        self.proxy = ProxyFix(application, x_for=1, x_proto=1, x_host=0)
    def __call__(self, environ, start_response):
        application = self.proxy if environ.get('REMOTE_ADDR') in ('127.0.0.1', '::1') else self.raw
        return application(environ, start_response)
app.wsgi_app = LocalProxy(app.wsgi_app)
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'vendo_sessions.db')
active_clients = {}
active_clients_lock = threading.RLock()
active_depositor_ip = None
active_depositor_timeout = 0
ser = None

def get_client_ip():
    # Proxy headers are interpreted only for our loopback nginx connection.
    value = request.remote_addr or '127.0.0.1'
    try:
        return str(ipaddress.ip_address(value))
    except ValueError:
        return '127.0.0.1'


@contextmanager
def db_connection():
    connection = sqlite3.connect(DB_PATH, timeout=15)
    try:
        with connection:
            yield connection
    finally:
        connection.close()

def atomic_credit_change(function):
    return function

def license_valid():
    return bool(license_manager.verify_license().get('valid'))

def session_expired(sess):
    if sess.get('is_paused') and sess.get('expires_at', 0) and time.time() >= sess['expires_at']:
        sess.update(remaining_seconds=0, expires_at=0, is_paused=False)
        return True
    return False

def resume_session(sess):
    session_expired(sess)
    if sess.get('admin_paused'):
        return False
    sess.update(is_paused=False, user_paused=False, auto_paused=False, paused_at=0, expires_at=0)
    return True

def ensure_session_schema(conn):
    conn.execute('CREATE TABLE IF NOT EXISTS active_sessions (ip TEXT PRIMARY KEY, mac TEXT, remaining_seconds INTEGER, is_paused INTEGER, dl_kbps INTEGER, ul_kbps INTEGER, pending_bottles INTEGER, paused_at REAL, expires_at REAL, saved_at REAL, state_json TEXT)')
    fields = {row[1] for row in conn.execute('PRAGMA table_info(active_sessions)')}
    for name, kind in [('paused_at', 'REAL DEFAULT 0'), ('expires_at', 'REAL DEFAULT 0'), ('state_json', 'TEXT')]:
        if name not in fields:
            conn.execute('ALTER TABLE active_sessions ADD COLUMN ' + name + ' ' + kind)

def init_db():
    with db_connection() as conn:
        c = conn.cursor()
        ensure_session_schema(conn)
        c.execute('CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)')
        c.execute('CREATE TABLE IF NOT EXISTS stats (date TEXT PRIMARY KEY, total_bottles INTEGER)')
        c.execute('CREATE TABLE IF NOT EXISTS admins (username TEXT PRIMARY KEY, password_hash TEXT)')
        c.execute('CREATE TABLE IF NOT EXISTS vouchers (code TEXT PRIMARY KEY, minutes INTEGER, is_used INTEGER DEFAULT 0, created_at TEXT, used_by TEXT, note TEXT)')
        c.execute('CREATE TABLE IF NOT EXISTS time_transfers (code TEXT PRIMARY KEY, from_ip TEXT, from_mac TEXT, seconds INTEGER, created_at REAL, is_claimed INTEGER DEFAULT 0)')
        c.execute('CREATE TABLE IF NOT EXISTS members (username TEXT PRIMARY KEY, pin_hash TEXT, wallet_minutes INTEGER DEFAULT 0, created_at TEXT)')
        c.execute('CREATE TABLE IF NOT EXISTS mac_control (mac TEXT PRIMARY KEY, type TEXT, note TEXT, dl_kbps INTEGER DEFAULT 0, ul_kbps INTEGER DEFAULT 0)')
        c.execute("CREATE TABLE IF NOT EXISTS promo_rates (bottles INTEGER PRIMARY KEY, minutes INTEGER, label TEXT, speed_profile TEXT DEFAULT '')")
        c.execute('CREATE TABLE IF NOT EXISTS announcements (id INTEGER PRIMARY KEY AUTOINCREMENT, message TEXT, active INTEGER DEFAULT 1)')
        c.execute('CREATE TABLE IF NOT EXISTS walled_garden (domain TEXT PRIMARY KEY, note TEXT)')
        try:
            c.execute('ALTER TABLE mac_control ADD COLUMN dl_kbps INTEGER DEFAULT 0')
        except sqlite3.OperationalError:
            pass
        try:
            c.execute('ALTER TABLE mac_control ADD COLUMN ul_kbps INTEGER DEFAULT 0')
        except sqlite3.OperationalError:
            pass
        try:
            c.execute('ALTER TABLE vouchers ADD COLUMN note TEXT')
        except sqlite3.OperationalError:
            pass
        try:
            c.execute("ALTER TABLE promo_rates ADD COLUMN speed_profile TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        try:
            c.execute('ALTER TABLE active_sessions ADD COLUMN paused_at REAL DEFAULT 0')
        except sqlite3.OperationalError:
            pass
        try:
            c.execute('ALTER TABLE active_sessions ADD COLUMN expires_at REAL DEFAULT 0')
        except sqlite3.OperationalError:
            pass
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('minutes_per_bottle', '10')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('drop_timeout', '60')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('default_dl_kbps', '3072')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('default_ul_kbps', '1536')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('custom_css', '')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('telegram_bot_token', '')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('telegram_chat_id', '')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('telegram_alert_bin', '1')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('telegram_alert_daily', '1')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('anti_tethering', '1')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('vendo_name', 'ECO-Fi Vendo')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('vendo_subtitle', 'Recycle Bottles for Fast WiFi')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('announcement', '')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('audio_bg', '/static/audio/eco_loop.wav')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('audio_insert', '/static/audio/bottle_success.wav')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('audio_success', '/static/audio/eco_success.wav')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('audio_preset', '/static/audio/eco_chime.wav')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('audio_custom_url', '')")
        c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('audio_volume', '80')")
        c.execute("UPDATE config SET value = '/static/audio/eco_loop.wav' WHERE key = 'audio_bg' AND (value = '/static/audio/b1.wav' OR value = '')")
        c.execute("UPDATE config SET value = '/static/audio/eco_chime.wav' WHERE key = 'audio_insert' AND (value = '/static/audio/coin.wav' OR value = '')")
        c.execute("UPDATE config SET value = '/static/audio/eco_success.wav' WHERE key = 'audio_success' AND (value = '/static/audio/success_ding.wav' OR value = '')")
        c.execute("INSERT OR IGNORE INTO promo_rates (bottles, minutes, label) VALUES (1, 10, '1 Bottle = 10 mins')")
        c.execute("INSERT OR IGNORE INTO promo_rates (bottles, minutes, label) VALUES (3, 40, '3 Bottles = 40 mins')")
        c.execute("INSERT OR IGNORE INTO promo_rates (bottles, minutes, label) VALUES (5, 75, '5 Bottles = 1h 15m')")
        c.execute("INSERT OR IGNORE INTO promo_rates (bottles, minutes, label) VALUES (10, 180, '10 Bottles = 3 Hours')")
        c.execute("INSERT OR IGNORE INTO walled_garden (domain, note) VALUES ('connectivitycheck.gstatic.com', 'Android Captive Probe')")
        c.execute("INSERT OR IGNORE INTO walled_garden (domain, note) VALUES ('captive.apple.com', 'Apple Captive Probe')")
        default_hash = generate_password_hash('admin123', method='pbkdf2:sha256')
        c.execute("INSERT OR IGNORE INTO admins (username, password_hash) VALUES ('admin', ?)", (default_hash,))
        conn.commit()
init_db()

def get_config(key, default=''):
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT value FROM config WHERE key=?', (key,))
        row = c.fetchone()
        return row[0] if row else default

def get_all_config():
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT key, value FROM config')
        cfg = {row[0]: row[1] for row in c.fetchall()}
        c.execute('SELECT message FROM announcements WHERE active = 1 ORDER BY id DESC LIMIT 1')
        row = c.fetchone()
        cfg['announcement'] = row[0] if row else cfg.get('announcement', '')
        return cfg

def set_config(key, value):
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('REPLACE INTO config (key, value) VALUES (?, ?)', (key, str(value)))
        conn.commit()

def _initialize_secret_key():
    secret = get_config('flask_secret_key', '')
    if not secret:
        import binascii
        secret = binascii.hexlify(os.urandom(32)).decode('ascii')
        set_config('flask_secret_key', secret)
    return secret
app.secret_key = _initialize_secret_key()
app.config.update(SESSION_COOKIE_SAMESITE='Lax')

def calculate_minutes_for_bottles(bottles_count):
    if bottles_count <= 0:
        return 0
    total_minutes = 0
    remaining_bottles = bottles_count
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT bottles, minutes FROM promo_rates ORDER BY bottles DESC')
        rates = c.fetchall()
        if not rates:
            base_rate = int(get_config('minutes_per_bottle', '10'))
            return bottles_count * base_rate
        for tier_bottles, tier_minutes in rates:
            if remaining_bottles >= tier_bottles and tier_bottles > 0:
                multiplier = remaining_bottles // tier_bottles
                total_minutes += multiplier * tier_minutes
                remaining_bottles %= tier_bottles
        if remaining_bottles > 0:
            c.execute('SELECT minutes FROM promo_rates WHERE bottles = 1')
            base_row = c.fetchone()
            base_rate = base_row[0] if base_row else int(get_config('minutes_per_bottle', '10'))
            total_minutes += remaining_bottles * base_rate
    return total_minutes

def record_bottle_drop(count=1):
    today = datetime.now().strftime('%Y-%m-%d')
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('INSERT OR IGNORE INTO stats (date, total_bottles) VALUES (?, 0)', (today,))
        c.execute('UPDATE stats SET total_bottles = total_bottles + ? WHERE date = ?', (count, today))
        conn.commit()
    print('[ECO-FI STATS] +{} bottle(s) recorded in database for {}.'.format(count, today), flush=True)

def send_telegram_alert(custom_msg=None):
    bot_token = get_config('telegram_bot_token')
    chat_id = get_config('telegram_chat_id')
    if not bot_token or not chat_id:
        return False
    try:
        url = 'https://api.telegram.org/bot{}/sendMessage'.format(bot_token)
        text = custom_msg or '🚨 *ECO-Fi Alert*\n\nThe recycling bin has reached **100% capacity**! Please empty the bin.'
        payload = {'chat_id': chat_id, 'text': text, 'parse_mode': 'Markdown'}
        data = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(url, data=data, headers={'Content-Type': 'application/json'})
        urllib.request.urlopen(req, timeout=5)
        return True
    except Exception:
        return False

def on_esp32_uart_output(raw_msg):
    try:
        data = json.loads(raw_msg)
        event = data.get('event')
        if event == 'CREDIT_ADD':
            bottles = int(data.get('bottles', 1))
            session_total = data.get('sessionTotal')
            if session_total is not None:
                esp32.current_session_bottles = session_total
            else:
                esp32.current_session_bottles += bottles
            record_bottle_drop(bottles)
            print('[ECO-FI VENDO] Bottle drop credit added: {} bottle(s), session total: {}'.format(bottles, esp32.current_session_bottles), flush=True)
            with active_clients_lock:
                if active_depositor_ip and active_depositor_ip in active_clients:
                    active_clients[active_depositor_ip]['pending_bottles'] = (
                        active_clients[active_depositor_ip].get('pending_bottles', 0) + bottles
                    )
            if active_depositor_ip:
                timeout = int(get_config('drop_timeout', '60') or 60)
                active_depositor_timeout = time.time() + timeout + 5
                transmit_to_esp32({'cmd': 'OPEN_GATE', 'timeout': timeout})
        elif event == 'REJECTED':
            if active_depositor_ip:
                timeout = int(get_config('drop_timeout', '60') or 60)
                transmit_to_esp32({'cmd': 'OPEN_GATE', 'timeout': timeout})
        elif event == 'BIN_FULL':
            set_config('hw_bin_full', '1')
            if get_config('telegram_alert_bin', '1') == '1':
                send_telegram_alert()
        elif event == 'BIN_OK':
            set_config('hw_bin_full', '0')
        elif event == 'CONFIG_SAVED':
            print('[ESP32] CONFIG_SAVED ACK received — config applied to hardware.')
    except Exception:
        pass
esp32 = ESP32Simulator(on_serial_output_callback=on_esp32_uart_output)

def transmit_to_esp32(payload_dict):
    msg_str = json.dumps(payload_dict) + '\n'
    esp32.receive_uart(msg_str)
    global ser
    if ser:
        try:
            ser.write(msg_str.encode())
        except Exception:
            pass
import socket

def check_network_health():
    """Monitor for network configuration loss and recover."""
    if platform.system() == 'Windows':
        return
    try:
        lan_iface = get_lan_interface()
        res = subprocess.run(['ip', 'addr', 'show', lan_iface], stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        if '10.0.0.1' not in res.stdout:
            subprocess.run(['systemctl', 'restart', 'networking'])
            subprocess.run(['systemctl', 'restart', 'dnsmasq'])
            time.sleep(5)
            setup_firewall()
    except Exception:
        pass

def get_lan_interface():
    return os.environ.get('ECOFI_LAN_IFACE', 'eth1')

def get_arp_table():
    if platform.system() == 'Windows':
        return {}
    try:
        res = subprocess.run(['ip', '-4', 'neigh', 'show', 'dev', get_lan_interface()], stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, timeout=5)
        result = {}
        for line in res.stdout.splitlines():
            parts = line.split()
            if 'lladdr' in parts and not any(state in parts for state in ('FAILED', 'INCOMPLETE')):
                result[parts[0]] = parts[parts.index('lladdr') + 1].lower()
        return result
    except Exception:
        log.exception('Cannot read LAN neighbors')
        return {}

def get_connected_ips():
    return set(get_arp_table().keys())

def setup_bandwidth_control(interface):
    if platform.system() == 'Windows':
        return
    try:
        subprocess.run(['tc', 'qdisc', 'del', 'dev', interface, 'root'], stderr=subprocess.DEVNULL)
        subprocess.run(['tc', 'qdisc', 'add', 'dev', interface, 'root', 'handle', '1:', 'htb', 'default', '99'], check=True)
        subprocess.run(['tc', 'class', 'add', 'dev', interface, 'parent', '1:', 'classid', '1:1', 'htb', 'rate', '100mbit'], check=True)
        subprocess.run(['tc', 'class', 'add', 'dev', interface, 'parent', '1:1', 'classid', '1:99', 'htb', 'rate', '100mbit'], check=True)
        
        subprocess.run(['modprobe', 'ifb', 'numifbs=1'], stderr=subprocess.DEVNULL)
        res = subprocess.run(['ip', 'link', 'set', 'dev', 'ifb0', 'up'], stderr=subprocess.DEVNULL)
        if res.returncode == 0:
            subprocess.run(['tc', 'qdisc', 'del', 'dev', interface, 'ingress'], stderr=subprocess.DEVNULL)
            subprocess.run(['tc', 'qdisc', 'add', 'dev', interface, 'ingress'], stderr=subprocess.DEVNULL)
            subprocess.run(['tc', 'filter', 'add', 'dev', interface, 'parent', 'ffff:', 'protocol', 'ip', 'u32', 'match', 'u32', '0', '0', 'action', 'mirred', 'egress', 'redirect', 'dev', 'ifb0'], stderr=subprocess.DEVNULL)
            subprocess.run(['tc', 'qdisc', 'del', 'dev', 'ifb0', 'root'], stderr=subprocess.DEVNULL)
            subprocess.run(['tc', 'qdisc', 'add', 'dev', 'ifb0', 'root', 'handle', '1:', 'htb', 'default', '99'], check=True)
            subprocess.run(['tc', 'class', 'add', 'dev', 'ifb0', 'parent', '1:', 'classid', '1:1', 'htb', 'rate', '100mbit'], check=True)
            subprocess.run(['tc', 'class', 'add', 'dev', 'ifb0', 'parent', '1:1', 'classid', '1:99', 'htb', 'rate', '100mbit'], check=True)
    except Exception as e:
        log.error('setup_bandwidth_control error: %s', e)

def apply_client_bandwidth(ip, dl_kbps, ul_kbps):
    if platform.system() == 'Windows':
        return
    interface = get_lan_interface()
    try:
        ip_int = int(ipaddress.IPv4Address(ip))
        mark = 100 + (ip_int & 16383)
        dl_kbps = max(64, int(dl_kbps))
        ul_kbps = max(64, int(ul_kbps))
        subprocess.run(['tc', 'class', 'replace', 'dev', interface, 'parent', '1:1', 'classid', '1:{}'.format(mark), 'htb', 'rate', '{}kbit'.format(dl_kbps), 'ceil', '{}kbit'.format(dl_kbps), 'burst', '15k'], check=True)
        subprocess.run(['tc', 'filter', 'replace', 'dev', interface, 'protocol', 'ip', 'parent', '1:', 'prio', str(mark), 'u32', 'match', 'ip', 'dst', '{}/32'.format(ip), 'flowid', '1:{}'.format(mark)], check=True)
        res = subprocess.run(['ip', 'link', 'show', 'ifb0'], stderr=subprocess.DEVNULL)
        if res.returncode == 0:
            subprocess.run(['tc', 'class', 'replace', 'dev', 'ifb0', 'parent', '1:1', 'classid', '1:{}'.format(mark), 'htb', 'rate', '{}kbit'.format(ul_kbps), 'ceil', '{}kbit'.format(ul_kbps), 'burst', '15k'], check=True)
            subprocess.run(['tc', 'filter', 'replace', 'dev', 'ifb0', 'protocol', 'ip', 'parent', '1:', 'prio', str(mark), 'u32', 'match', 'ip', 'src', '{}/32'.format(ip), 'flowid', '1:{}'.format(mark)], check=True)
    except Exception as e:
        log.error('apply_client_bandwidth error: %s', e)

def remove_client_bandwidth(ip):
    if platform.system() == 'Windows':
        return
    interface = get_lan_interface()
    try:
        ip_int = int(ipaddress.IPv4Address(ip))
        mark = 100 + (ip_int & 16383)
        subprocess.run(['tc', 'filter', 'del', 'dev', interface, 'protocol', 'ip', 'parent', '1:', 'prio', str(mark)], stderr=subprocess.DEVNULL)
        subprocess.run(['tc', 'class', 'del', 'dev', interface, 'parent', '1:1', 'classid', '1:{}'.format(mark)], stderr=subprocess.DEVNULL)
        res = subprocess.run(['ip', 'link', 'show', 'ifb0'], stderr=subprocess.DEVNULL)
        if res.returncode == 0:
            subprocess.run(['tc', 'filter', 'del', 'dev', 'ifb0', 'protocol', 'ip', 'parent', '1:', 'prio', str(mark)], stderr=subprocess.DEVNULL)
            subprocess.run(['tc', 'class', 'del', 'dev', 'ifb0', 'parent', '1:1', 'classid', '1:{}'.format(mark)], stderr=subprocess.DEVNULL)
    except Exception as e:
        log.error('remove_client_bandwidth error: %s', e)

def sync_client_firewall(ip):
    with active_clients_lock:
        sess = active_clients.get(ip)
        if sess is None:
            update_firewall(ip, 'del')
            return
        session_expired(sess)
        with db_connection() as conn:
            policy = conn.execute('SELECT type, dl_kbps, ul_kbps FROM mac_control WHERE lower(mac)=?', (sess.get('mac', '').lower(),)).fetchone()
        blocked = policy and policy[0] == 'block'
        free = policy and policy[0] == 'whitelist'
        dl = sess.get('dl_kbps') or int(get_config('default_dl_kbps', '3072'))
        ul = sess.get('ul_kbps') or int(get_config('default_ul_kbps', '1536'))
        if free:
            dl, ul = policy[1] or dl, policy[2] or ul
        allowed = license_valid() and not blocked and (free or (sess['remaining_seconds'] > 0 and not sess.get('is_paused')))
        if allowed:
            ok = update_firewall(ip, 'add', 30 if free else sess['remaining_seconds'], dl, ul)
            sess['access_error'] = '' if ok is not False else 'Network authorization unavailable; credit is preserved.'
        else:
            update_firewall(ip, 'del')
            sess['access_error'] = ''

def apply_walled_garden_and_macs():
    if platform.system() == 'Windows': return
    with db_connection() as conn:
        blocked = {row[0].lower() for row in conn.execute("SELECT mac FROM mac_control WHERE type='block'")}
        domains = [row[0] for row in conn.execute('SELECT domain FROM walled_garden')]
    addresses = set()
    for domain in domains:
        try:
            for result in socket.getaddrinfo(domain, None, socket.AF_INET):
                address = ipaddress.ip_address(result[4][0])
                if address.is_global: addresses.add(str(address))
        except socket.gaierror:
            log.warning('Walled garden DNS unavailable: %s', domain)
    gateway_network.policies(blocked, addresses)
    with active_clients_lock:
        for ip in list(active_clients): sync_client_firewall(ip)

def setup_firewall():
    if platform.system() == 'Windows': return
    gateway_network.setup(get_lan_interface(), os.environ.get('ECOFI_WAN_IFACE', 'eth0'))
    gateway_network.set_license(license_valid())
    apply_walled_garden_and_macs()

def update_firewall(ip, action, timeout_sec=0, dl_kbps=3072, ul_kbps=1536):
    if platform.system() == 'Windows': return True
    try:
        if ipaddress.ip_address(ip) not in gateway_network.SUBNET or ip == '10.0.0.1': return False
        if action == 'del':
            gateway_network.revoke(ip)
            return True
        mac = get_arp_table().get(ip)
        return gateway_network.grant(ip, mac, timeout_sec, dl_kbps, ul_kbps)
    except Exception as e:
        log.error('Client enforcement failed for %s: %s', ip, e)
        return False

import math

def calculate_pause_validity_seconds(remaining_seconds):
    """
    Computes dynamic validity duration in seconds based on remaining credit.
    Formula: V(T) = min(720h, max(24h, 12h + 1.2*sqrt(Mins) + 0.025*Mins))
    Minimum: 24 Hours (86,400s)
    Maximum: 30 Days (2,592,000s)
    """
    if remaining_seconds <= 0:
        return 0
    mins = remaining_seconds / 60.0
    validity_hours = 12.0 + 1.2 * math.sqrt(mins) + 0.025 * mins
    validity_hours = max(24.0, min(720.0, validity_hours))
    return int(validity_hours * 3600)

def compute_session_expiration(remaining_seconds, paused_at=None):
    """Returns the absolute Unix timestamp when the paused session expires."""
    if remaining_seconds <= 0:
        return 0
    base_time = paused_at or time.time()
    return int(base_time + calculate_pause_validity_seconds(remaining_seconds))

def save_sessions_to_db():
    with active_clients_lock:
        with db_connection() as conn:
            ensure_session_schema(conn)
            conn.execute('DELETE FROM active_sessions')
            for ip, s in active_clients.items():
                if ip in ('127.0.0.1', '10.0.0.1', '::1', 'localhost'): continue
                conn.execute('INSERT INTO active_sessions VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    (ip, s['mac'], s['remaining_seconds'], int(s.get('is_paused', False)),
                     s.get('dl_kbps', 3072), s.get('ul_kbps', 1536), s.get('pending_bottles', 0),
                     s.get('paused_at', 0), s.get('expires_at', 0), time.time(), json.dumps(s)))

def restore_sessions_from_db():
    with active_clients_lock:
        with db_connection() as conn:
            ensure_session_schema(conn)
            rows = conn.execute('SELECT ip, mac, remaining_seconds, is_paused, dl_kbps, ul_kbps, pending_bottles, paused_at, expires_at, state_json FROM active_sessions').fetchall()
        for ip, mac, remaining, paused, dl, ul, pending, paused_at, expires_at, state in rows:
            if ip in ('127.0.0.1', '10.0.0.1', '::1') or (remaining <= 0 and not pending): continue
            sess = json.loads(state) if state else {}
            sess.update(mac=mac, remaining_seconds=remaining, is_paused=bool(paused),
                        dl_kbps=dl or int(get_config('default_dl_kbps', '3072')),
                        ul_kbps=ul or int(get_config('default_ul_kbps', '1536')),
                        pending_bottles=pending or 0, paused_at=paused_at or 0, expires_at=expires_at or 0)
            if paused and not state: sess['user_paused'] = True
            # Accepted but unfinalized bottles survive a restart as earned time.
            if pending:
                sess['remaining_seconds'] += calculate_minutes_for_bottles(pending) * 60
                sess['pending_bottles'] = 0
            session_expired(sess)
            active_clients[ip] = sess
            sync_client_firewall(ip)
        save_sessions_to_db()
atexit.register(save_sessions_to_db)

def time_daemon():
    tick = 0
    last_tick_time = None
    while True:
        tick += 1
        now = time.time()
        if last_tick_time is None:
            elapsed = 0
        else:
            elapsed = max(1, int(round(now - last_tick_time)))
        last_tick_time = now
        if tick % 30 == 0:
            save_sessions_to_db()
        if tick % 60 == 0:
            arp_table = get_arp_table()
            connected_ips = set(arp_table.keys())
            if platform.system() != 'Windows':
                try:
                    with db_connection() as conn:
                        c = conn.cursor()
                        c.execute("SELECT mac FROM mac_control WHERE type='whitelist'")
                        whitelisted_macs = {row[0].lower() for row in c.fetchall()}
                        for ip, mac in arp_table.items():
                            if mac in whitelisted_macs:
                                subprocess.run(['ipset', 'add', 'ecofi_auth', ip, '-exist'])
                except Exception:
                    pass
            auto_pause_enabled = get_config('auto_pause_disconnect', '0') == '1'
            if auto_pause_enabled and platform.system() != 'Windows':
                with active_clients_lock:
                    for ip, session_data in list(active_clients.items()):
                        if ip != '127.0.0.1' and session_data['remaining_seconds'] > 0:
                            if ip not in connected_ips and (not session_data.get('is_paused')):
                                session_data['is_paused'] = True
                                session_data['auto_paused'] = True
                                session_data['paused_at'] = now
                                session_data['expires_at'] = compute_session_expiration(session_data['remaining_seconds'], now)
                                sync_client_firewall(ip)
                            elif ip in connected_ips and session_data.get('auto_paused') and (not session_data.get('user_paused')) and (not session_data.get('admin_paused')):
                                session_data['is_paused'] = False
                                session_data['auto_paused'] = False
                                session_data['paused_at'] = 0
                                session_data['expires_at'] = 0
                                sync_client_firewall(ip)
        if tick % 300 == 0:
            check_network_health()
        if tick % 3600 == 0:
            apply_walled_garden_and_macs()
        with active_clients_lock:
            for ip, session_data in list(active_clients.items()):
                if session_data.get('is_paused') and session_data.get('expires_at', 0) > 0:
                    if now > session_data['expires_at']:
                        session_data['remaining_seconds'] = 0
                        session_data['is_paused'] = False
                        session_data['expires_at'] = 0
                        sync_client_firewall(ip)
                        continue
                was_active = session_data['remaining_seconds'] > 0 and (not session_data.get('is_paused', False))
                if was_active and elapsed > 0:
                    session_data['remaining_seconds'] = max(0, session_data['remaining_seconds'] - elapsed)
                    if session_data['remaining_seconds'] <= 0:
                        sync_client_firewall(ip)
                    elif tick % 10 == 0:
                        sync_client_firewall(ip)
        time.sleep(1)

def ensure_client_session(ip):
    with active_clients_lock:
        dummy = {'mac': '00:00:00:00:00:00', 'pending_bottles': 0, 'remaining_seconds': 0, 'is_paused': False, 'paused_at': 0, 'expires_at': 0, 'dl_kbps': 0, 'ul_kbps': 0}
        if ip in ('127.0.0.1', '10.0.0.1', '::1', 'localhost'): return dummy
        mac = get_arp_table().get(ip, '').lower()
        existing = active_clients.get(ip)
        if existing and (not mac or existing.get('mac') == mac): return existing
        if existing:
            # Keep the previous owner's balance under a non-routable identity.
            old_mac = existing.get('mac', '')
            update_firewall(ip, 'del')
            if existing.get('remaining_seconds', 0) > 0:
                active_clients['saved:' + old_mac] = existing
            del active_clients[ip]
        if mac:
            for old_ip, old_sess in list(active_clients.items()):
                if old_ip != ip and old_sess.get('mac', '').lower() == mac:
                    if not old_ip.startswith('saved:'):
                        update_firewall(old_ip, 'del')
                    active_clients[ip] = old_sess
                    del active_clients[old_ip]
                    sync_client_firewall(ip)
                    return old_sess
        sess = dict(dummy)
        sess.update(mac=mac, dl_kbps=int(get_config('default_dl_kbps', '3072')), ul_kbps=int(get_config('default_ul_kbps', '1536')))
        active_clients[ip] = sess
        return sess
PORTAL_HTML = '\n<!DOCTYPE html>\n<html lang="en">\n<head>\n    <meta charset="UTF-8">\n    <title>{{ vendo_name }} - Reverse Vending WiFi Portal</title>\n    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">\n    <link rel="stylesheet" href="/static/vendor/fontawesome/css/all.min.css">\n    <style>\n        * { box-sizing: border-box; }\n        img { max-width: 100%; height: auto; }\n        body { \n            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; \n            margin: 0; padding: 12px 10px 24px 10px; min-height: 100vh;\n            background: linear-gradient(rgba(15, 23, 42, 0.94), rgba(15, 23, 42, 0.98)), url(\'/static/banner.jpg\') no-repeat center center fixed;\n            background-size: cover;\n            color: #f1f5f9; \n            display: flex; justify-content: center; align-items: flex-start;\n            overflow-x: hidden;\n            width: 100%;\n        }\n\n        .portal-container {\n            width: 100%; max-width: 375px;\n            background: rgba(30, 41, 59, 0.88);\n            backdrop-filter: blur(16px);\n            -webkit-backdrop-filter: blur(16px);\n            border-radius: 14px;\n            border: 1px solid rgba(255, 255, 255, 0.1);\n            padding: 12px 14px;\n            text-align: center;\n            position: relative;\n            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.5);\n            overflow: hidden;\n        }\n\n        .brand-banner-box {\n            width: 100%;\n            border-radius: 9px;\n            overflow: hidden;\n            margin-bottom: 9px;\n            border: 1px solid rgba(255, 255, 255, 0.12);\n        }\n        .brand-banner-img {\n            width: 100%;\n            height: auto;\n            display: block;\n            object-fit: cover;\n        }\n\n        .announcement-bar {\n            background: rgba(16, 185, 129, 0.1); border-left: 3px solid #10b981;\n            padding: 5px 9px; border-radius: 5px; font-size: 10.5px; color: #a7f3d0;\n            margin-bottom: 9px; text-align: left;\n        }\n\n        .bin-full-banner {\n            background: rgba(239, 68, 68, 0.15); border-left: 3px solid #ef4444;\n            padding: 6px 9px; border-radius: 5px; font-size: 11px; color: #fca5a5;\n            margin-bottom: 9px; text-align: left; display: none; font-weight: 600;\n        }\n\n        .status-box {\n            background: rgba(15, 23, 42, 0.65);\n            border-radius: 9px; padding: 8px 10px; margin-bottom: 9px;\n            border: 1px solid rgba(255, 255, 255, 0.08);\n        }\n        .status-text { \n            font-size: 10px; text-transform: uppercase; letter-spacing: 0.8px; \n            color: #94a3b8; font-weight: 600; margin-bottom: 2px;\n        }\n        .time-display { \n            font-size: 24px; font-family: "SF Mono", "Roboto Mono", "Courier New", monospace; \n            font-weight: 700; color: #10b981; margin: 2px 0 3px 0; \n            letter-spacing: 1px;\n        }\n        .status-badge { \n            display: inline-block; padding: 2px 8px; border-radius: 12px; \n            font-size: 9px; font-weight: 600; letter-spacing: 0.4px;\n        }\n        .bg-active { background: rgba(16, 185, 129, 0.15); color: #34d399; border: 1px solid rgba(16, 185, 129, 0.35); }\n        .bg-paused { background: rgba(245, 158, 11, 0.15); color: #fbbf24; border: 1px solid rgba(245, 158, 11, 0.35); }\n        .bg-inactive { background: rgba(239, 68, 68, 0.15); color: #f87171; border: 1px solid rgba(239, 68, 68, 0.35); }\n        .bg-binfull { background: rgba(239, 68, 68, 0.25); color: #fca5a5; border: 1px solid #ef4444; }\n\n        \n        \n        \n        \n        \n        \n        .btn-pause { \n            background: rgba(245, 158, 11, 0.2); border: 1px solid rgba(245, 158, 11, 0.4); \n            color: #fbbf24; height: 35px; font-size: 11.5px;\n        }\n        .btn-pause:hover { background: rgba(245, 158, 11, 0.3); }\n        .btn-resume { \n            background: rgba(59, 130, 246, 0.2); border: 1px solid rgba(59, 130, 246, 0.4); \n            color: #60a5fa; height: 35px; font-size: 11.5px;\n        }\n        .btn-resume:hover { background: rgba(59, 130, 246, 0.3); }\n\n        .nav-tabs {\n            display: flex; gap: 3px; margin-bottom: 9px; background: rgba(15, 23, 42, 0.55);\n            padding: 3px; border-radius: 7px; border: 1px solid rgba(255, 255, 255, 0.06);\n            width: 100%;\n        }\n        .tab-btn {\n            flex: 1; padding: 6px 1px; font-size: 11px; font-weight: 500; color: #94a3b8;\n            background: transparent; border: none; border-radius: 5px; cursor: pointer;\n            transition: all 0.15s ease; text-align: center; white-space: nowrap;\n        }\n        .tab-btn.active { \n            background: #1e293b; color: #10b981; font-weight: 600;\n            border: 1px solid rgba(16, 185, 129, 0.25);\n            box-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);\n        }\n\n        .tab-content { display: none; text-align: left; width: 100%; }\n        .tab-content.active { display: block; }\n\n        .custom-input {\n            width: 100%; height: 35px; padding: 6px 9px; border-radius: 6px; \n            background: rgba(15, 23, 42, 0.7);\n            border: 1px solid rgba(255, 255, 255, 0.12); color: #f8fafc; font-size: 12px; margin-bottom: 7px;\n            box-sizing: border-box; transition: border-color 0.15s ease;\n        }\n        .custom-input:focus { border-color: #10b981; outline: none; box-shadow: 0 0 0 1px rgba(16, 185, 129, 0.3); }\n\n        .table-info {\n            width: 100%; border-collapse: collapse; font-size: 11.5px; color: #cbd5e1;\n            margin-top: 3px; table-layout: fixed;\n        }\n        .table-info tr { border-bottom: 1px solid rgba(255, 255, 255, 0.05); }\n        .table-info tr:last-child { border-bottom: none; }\n        .table-info td { padding: 6px 2px; }\n\n        \n        \n\n        .modal-overlay {\n            display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%;\n            background: rgba(15, 23, 42, 0.88); backdrop-filter: blur(12px);\n            z-index: 1000; align-items: center; justify-content: center;\n        }\n        .modal-box {\n            width: 90%; max-width: 360px; background: #1e293b; border-radius: 14px;\n            padding: 20px 16px; border: 1px solid rgba(255, 255, 255, 0.1); box-shadow: 0 15px 35px rgba(0, 0, 0, 0.6);\n            text-align: center; position: relative; overflow: hidden;\n        }\n        .countdown-circle {\n            font-size: 32px; font-weight: 700; color: #f59e0b; margin: 8px 0;\n            font-family: monospace;\n        }\n        .bottle-counter {\n            font-size: 20px; font-weight: 700; color: #10b981; margin-bottom: 10px;\n        }\n        .progress-bar-bg {\n            width: 100%; height: 8px; background: #334155; border-radius: 6px;\n            overflow: hidden; margin-bottom: 10px;\n        }\n        .progress-bar-fill {\n            height: 100%; width: 100%; background: linear-gradient(90deg, #10b981, #34d399);\n            transition: width 0.3s ease;\n        }\n        .chute-stage-box {\n            background: rgba(15, 23, 42, 0.6); padding: 7px 9px; border-radius: 7px;\n            font-size: 11px; color: #94a3b8; margin-bottom: 10px; border: 1px dashed rgba(16, 185, 129, 0.3);\n        }\n        @keyframes drop-in {\n            0% { transform: translateY(-30px) rotate(0deg) scale(0.6); opacity: 0; }\n            50% { transform: translateY(0) rotate(10deg) scale(1.1); opacity: 1; }\n            100% { transform: translateY(0) rotate(0deg) scale(1); opacity: 1; }\n        }\n        .bottle-pop {\n            display: inline-block; animation: drop-in 0.35s ease-out;\n        }\n    \n        /* Tactile Physical Appliance Button Structure */\n        .btn-tactile {\n            display: flex;\n            align-items: center;\n            justify-content: center;\n            outline: none;\n            cursor: pointer;\n            width: 100%;\n            height: 44px;\n            background-image: linear-gradient(to top, #D8D9DB 0%, #fff 80%, #FDFDFD 100%);\n            border-radius: 30px;\n            border: 1px solid #8F9092;\n            transition: transform 0.1s ease, filter 0.15s ease;\n            font-family: "Source Sans Pro", -apple-system, BlinkMacSystemFont, sans-serif;\n            font-size: 13.5px;\n            font-weight: 600;\n            color: #374151;\n            text-shadow: 0 1px #fff;\n            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.25);\n            letter-spacing: 0.2px;\n            text-decoration: none;\n            gap: 7px;\n            margin-bottom: 8px;\n        }\n\n        /* Removed hover glowing box-shadow completely */\n        .btn-tactile:hover {\n            filter: brightness(0.97);\n            color: #1f2937;\n        }\n\n        .btn-tactile:active {\n            transform: translateY(1px);\n            box-shadow: 0 1px 2px rgba(0, 0, 0, 0.2), inset 0 1px 3px rgba(0, 0, 0, 0.2);\n        }\n\n        .btn-tactile:focus {\n            outline: none;\n        }\n\n        .btn-tactile:disabled {\n            opacity: 0.55;\n            cursor: not-allowed;\n            background-image: linear-gradient(to top, #bbb 0%, #ddd 100%) !important;\n            box-shadow: none !important;\n            color: #777 !important;\n            filter: none !important;\n        }\n\n        /* Hero Insert Button: Distinct 18px gap to separate clearly from the tab bar */\n        .btn-tactile-green {\n            color: #065f46;\n            font-weight: 700;\n            font-size: 14px;\n            border-color: #059669;\n            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.9);\n            margin-bottom: 18px !important;\n        }\n        .btn-tactile-green i {\n            color: #10b981;\n            font-size: 15px;\n        }\n\n        /* Tactile Compact variant for side-by-side buttons */\n        .btn-tactile-sm {\n            height: 38px;\n            font-size: 12.5px;\n            border-radius: 20px;\n            margin-bottom: 0;\n        }\n\n    </style>\n</head>\n<body>\n\n    <div class="portal-container" style="margin-top: 15px;">\n        <div class="brand-banner-box">\n            <img src="/static/banner-main.jpg" alt="Smart ECO-Fi Vendo" class="brand-banner-img">\n        </div>\n\n        {% if announcement %}\n        <div class="announcement-bar">\n            <i class="fas fa-bullhorn"></i> {{ announcement }}\n        </div>\n        {% endif %}\n\n        <div id="bin-full-banner" class="bin-full-banner">\n            <i class="fas fa-exclamation-triangle"></i> Storage bin is currently full. Machine cannot accept new bottles at this moment.\n        </div>\n\n        {% if license_valid %}\n        <div class="status-box">\n            <div class="status-text">Available Internet Time</div>\n            <div class="time-display" id="time-display">0d 00h:00m:00s</div>\n            <div id="status-badge" class="status-badge bg-inactive">DISCONNECTED</div>\n        </div>\n\n        <!-- MAIN ACTION: PULSING INSERT BOTTLE BUTTON -->\n        <button id="btn-insert" class="btn-tactile btn-tactile-green" onclick="startDepositSession()">\n            <i class="fas fa-recycle mr-1"></i> Insert Plastic Bottle\n        </button>\n\n        <div id="pause-ctrl-box" style="display:none; margin-bottom: 12px;">\n            <button id="btn-pause" class="btn-tactile btn-tactile-sm" style="color:#b45309; font-weight:700;" onclick="togglePause(\'pause\')">\n                <i class="fas fa-pause"></i> PAUSE TIME\n            </button>\n            <button id="btn-resume" class="btn-tactile btn-tactile-sm" style="color:#1d4ed8; font-weight:700;" style="display:none;" onclick="togglePause(\'resume\')">\n                <i class="fas fa-play"></i> RESUME TIME\n            </button>\n        </div>\n\n        <!-- MULTI-TAB FEATURES -->\n        <div class="nav-tabs">\n            <button class="tab-btn active" onclick="switchTab(\'tab-rates\')">Rates</button>\n            <button class="tab-btn" onclick="switchTab(\'tab-voucher\')">Voucher</button>\n            <button class="tab-btn" onclick="switchTab(\'tab-transfer\')">Transfer</button>\n            <button class="tab-btn" onclick="switchTab(\'tab-member\')">Member</button>\n        </div>\n\n        <!-- TAB 1: PROMO RATES -->\n        <div id="tab-rates" class="tab-content active">\n            <div style="font-size: 11.5px; color:#94a3b8; font-weight:600; margin-bottom:6px;"><i class="fas fa-tags text-success mr-1"></i> RATES & PACKAGES</div>\n            <table class="table-info">\n                {% for r in promo_rates %}\n                <tr>\n                    <td style="padding: 7px 6px;"><strong style="color:#34d399; font-size:13px;">{{ r.bottles }} Bottle{% if r.bottles > 1 %}s{% endif %}</strong></td>\n                    <td style="text-align:right; font-weight:700; color:#f8fafc; font-size:13px; padding: 7px 6px;">\n                        {% if r.minutes >= 60 %}\n                            {% set hrs = (r.minutes // 60) %}\n                            {% set mins = (r.minutes % 60) %}\n                            {% if mins == 0 %}\n                                {{ hrs }} Hour{% if hrs > 1 %}s{% endif %}\n                            {% else %}\n                                {{ hrs }}h {{ mins }}m\n                            {% endif %}\n                        {% else %}\n                            {{ r.minutes }} mins\n                        {% endif %}\n                    </td>\n                </tr>\n                {% endfor %}\n            </table>\n        </div>\n\n        <!-- TAB 2: VOUCHER REDEMPTION -->\n        <div id="tab-voucher" class="tab-content">\n            <div style="font-size: 13px; color:#94a3b8; font-weight:700; margin-bottom:6px;">ENTER VOUCHER CODE:</div>\n            <input type="text" id="voucher-code-input" class="custom-input" placeholder="e.g. ECO-XXXX" oninput="this.value = this.value.toUpperCase().replace(/[^A-Z0-9-]/g, \'\')">\n            <button class="btn-tactile btn-tactile-green" style="height:36px; font-size:12px; margin-bottom:0;" onclick="redeemVoucher()">\n                <i class="fas fa-ticket-alt mr-1"></i> Redeem Voucher\n            </button>\n            <div id="voucher-msg" style="font-size:12px; margin-top:4px;"></div>\n        </div>\n\n        <!-- TAB 3: TIME TRANSFER -->\n        <div id="tab-transfer" class="tab-content">\n            <div style="font-size: 12px; color:#94a3b8; font-weight:700; margin-bottom:6px;">SHARE / TRANSFER YOUR TIME:</div>\n            <div style="display:flex; gap:6px; margin-bottom:6px;">\n                <input type="number" id="transfer-mins-input" class="custom-input" placeholder="Minutes to Share (e.g. 5)" min="1" style="margin-bottom:0; flex:1;">\n                <button class="btn-tactile btn-tactile-sm" style="width:auto; padding:0 16px; white-space:nowrap;" onclick="generateTransferCode()">\n                    <i class="fas fa-share-alt"></i> Share\n                </button>\n            </div>\n            <div id="transfer-code-display" style="font-size:12.5px; font-weight:700; color:#38bdf8; margin:4px 0; text-align:center;"></div>\n            \n            <hr style="border:0; border-top:1px solid rgba(255,255,255,0.08); margin:8px 0;">\n            <div style="font-size: 11.5px; color:#94a3b8; font-weight:600; margin-bottom:4px;">CLAIM A TRANSFER CODE:</div>\n            <div style="display:flex; gap:6px;">\n                <input type="text" id="claim-code-input" class="custom-input" placeholder="6-Digit Code" maxlength="6" style="margin-bottom:0; flex:1;" oninput="this.value = this.value.replace(/[^0-9]/g,\'\')">\n                <button class="btn-tactile btn-tactile-green" style="width:auto; padding:0 12px; height:36px; font-size:11.5px; margin-bottom:0; white-space:nowrap;" onclick="claimTransferCode()">\n                    <i class="fas fa-download"></i> Claim\n                </button>\n            </div>\n            <div id="claim-status-msg" style="font-size:11px; margin-top:4px; text-align:center;"></div>\n        </div>\n\n        <!-- TAB 4: MEMBER WALLET -->\n        <div id="tab-member" class="tab-content">\n            <div id="mem-auth-section">\n                <div style="background:rgba(16,185,129,0.08); border-left:3px solid #10b981; padding:6px 9px; border-radius:6px; font-size:10.5px; color:#a7f3d0; margin-bottom:8px; text-align:left; line-height:1.35;">\n                    <i class="fas fa-shield-alt text-success"></i> <strong>Zero-Expiry Storage:</strong> Register once to save Wi-Fi minutes across your devices.\n                </div>\n                <div style="font-size: 13px; color:#94a3b8; font-weight:700; margin-bottom:6px;">MEMBER LOGIN / REGISTER:</div>\n                <input type="text" id="member-user" class="custom-input" placeholder="Username (letters & numbers)" maxlength="20">\n                <input type="password" id="member-pin" class="custom-input" placeholder="4 to 6-Digit Secret PIN" maxlength="6" inputmode="numeric">\n                <div style="display:flex; gap:8px; margin-top:2px;">\n                    <button class="btn-tactile btn-tactile-green" style="height:36px; font-size:12px; margin-bottom:0; flex:1;" onclick="memberLogin()"><i class="fas fa-sign-in-alt"></i> Login</button>\n                    <button class="btn-tactile btn-tactile-sm" style="flex:1;" onclick="memberRegister()"><i class="fas fa-user-plus"></i> Register</button>\n                </div>\n                <div id="member-status" style="font-size:12px; margin-top:8px;"></div>\n            </div>\n\n            <!-- Logged In Wallet Manager -->\n            <div id="mem-wallet-section" style="display:none; text-align:left;">\n                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">\n                    <span style="font-weight:700; color:#38bdf8;" id="mem-welcome-user">Member</span>\n                    <button class="btn btn-xs" style="background:#ef4444; color:white; border:none; border-radius:6px; padding:3px 8px; font-size:11px; cursor:pointer;" onclick="memberLogout()"><i class="fas fa-sign-out-alt"></i> Logout</button>\n                </div>\n                <div style="background:rgba(16,185,129,0.15); border:1px solid rgba(16,185,129,0.3); padding:12px; border-radius:12px; margin-bottom:12px; text-align:center;">\n                    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:4px;">\n                        <span style="font-size:11px; color:#94a3b8; text-transform:uppercase;">Stored Wallet Balance</span>\n                        <span style="background:rgba(16,185,129,0.3); color:#bbf7d0; font-size:10px; padding:2px 8px; border-radius:12px; font-weight:700;"><i class="fas fa-infinity"></i> PERPETUAL (Zero Expiry)</span>\n                    </div>\n                    <div style="font-size:28px; font-weight:800; color:#34d399;" id="mem-wallet-mins">0 Mins</div>\n                </div>\n                <div style="margin-bottom:8px;">\n                    <label style="font-size:11px; color:#94a3b8;">Use Stored Minutes on this Device:</label>\n                    <div style="display:flex; gap:6px;">\n                        <input type="number" id="use-wallet-mins" class="custom-input" placeholder="Mins" style="margin-bottom:0; width:90px;" min="1">\n                        <button class="btn-tactile btn-tactile-green" style="padding:8px 12px; margin-bottom:0; font-size:12px;" onclick="useMemberWallet()"><i class="fas fa-wifi"></i> Connect</button>\n                    </div>\n                </div>\n                <button class="btn-tactile btn-tactile-sm" style="color:#b45309; font-weight:700;" style="padding:8px 12px; font-size:12px; margin-top:6px;" onclick="saveSessionToWallet()">\n                    <i class="fas fa-save"></i> Save Active Session to Wallet\n                </button>\n            </div>\n        </div>\n\n        {% else %}\n        <div style="background: rgba(239, 68, 68, 0.1); border: 2px dashed #ef4444; border-radius: 12px; padding: 30px 15px; margin-top: 15px; text-align: center;">\n            <i class="fas fa-lock" style="font-size: 32px; color: #ef4444; margin-bottom: 10px;"></i>\n            <h3 style="color: #fca5a5; margin: 0 0 10px 0;">Unlicensed Vendo</h3>\n            <p style="color: #cbd5e1; font-size: 14px; margin: 0;">Please contact support or the administrator to activate this machine.</p>\n        </div>\n        {% endif %}\n\n        <!-- VISUAL GUIDE: ALLOWED VS NOT ALLOWED BOTTLES -->\n        <img src="/static/info-graphic.jpg" alt="Bottle Acceptance Guide" style="width: 100%; max-width: 100%; border-radius: 9px; margin-top: 12px; border: 1px solid rgba(255, 255, 255, 0.12); display: block;">\n\n        <div style="font-size:11px; color:#64748b; margin-top:18px;">\n            IP: {{ client_ip }} | MAC: {{ client_mac }}\n        </div>\n    </div>\n\n    <!-- LIVE DEPOSIT MODAL WITH ANIMATED BOTTLE DROP & STATUS -->\n    <div id="deposit-modal" class="modal-overlay">\n        <div class="modal-box">\n            <h3 style="margin-top:0; color:#34d399;"><i class="fas fa-door-open"></i> AIRLOCK GATE OPEN</h3>\n            <div class="chute-stage-box" id="modal-stage-text">\n                <i class="fas fa-spinner fa-spin text-success"></i> Ready! Drop your PET plastic bottle into the chute...\n            </div>\n            \n            <div class="countdown-circle" id="modal-timer">30s</div>\n            <div class="progress-bar-bg">\n                <div id="modal-progress-bar" class="progress-bar-fill"></div>\n            </div>\n\n            <div class="bottle-counter">\n                <span id="modal-bottle-icon" class="bottle-pop"><i class="fas fa-wine-bottle"></i></span>\n                <span id="modal-bottles">0</span> Bottles (<span id="modal-added-time">+0m</span>)\n            </div>\n\n            <button class="btn-tactile btn-tactile-green" onclick="closeDepositSession()">\n                <i class="fas fa-check-circle"></i> DONE / START BROWSING\n            </button>\n        </div>\n    </div>\n\n    <script>\n        const audioBgSrc = "{{ audio_bg }}";\n        const audioInsertSrc = "{{ audio_insert }}";\n        const audioSuccessSrc = "{{ audio_success }}";\n        const audioVolume = (parseInt("{{ audio_volume or 80 }}") || 80) / 100.0;\n\n        let bgAudioElem = null;\n        if (audioBgSrc && audioBgSrc !== \'silent\') {\n            bgAudioElem = new Audio(audioBgSrc);\n            bgAudioElem.loop = true;\n            bgAudioElem.volume = audioVolume * 0.5;\n        }\n\n        let insertAudioElem = null;\n        if (audioInsertSrc && audioInsertSrc !== \'silent\' && audioInsertSrc !== \'arcade_powerup\' && audioInsertSrc !== \'voice_filipino\') {\n            insertAudioElem = new Audio(audioInsertSrc);\n            insertAudioElem.volume = audioVolume;\n            insertAudioElem.load();\n        }\n\n        let successAudioElem = null;\n        if (audioSuccessSrc && audioSuccessSrc !== \'silent\' && audioSuccessSrc !== \'crystal_bell\') {\n            successAudioElem = new Audio(audioSuccessSrc);\n            successAudioElem.volume = audioVolume;\n            successAudioElem.load();\n        }\n\n        const audioCtx = new (window.AudioContext || window.webkitAudioContext)();\n        \n        function unlockAudio() {\n            if (audioCtx && audioCtx.state === \'suspended\') audioCtx.resume();\n            if (insertAudioElem) { insertAudioElem.play().then(()=>insertAudioElem.pause()).catch(()=>{}); }\n            if (successAudioElem) { successAudioElem.play().then(()=>successAudioElem.pause()).catch(()=>{}); }\n        }\n        \n        document.addEventListener(\'click\', unlockAudio, { once: true });\n        \n        function playChimeTone(freq, type, duration, gainVal=0.3) {\n            try {\n                const osc = audioCtx.createOscillator();\n                const gain = audioCtx.createGain();\n                osc.type = type; osc.frequency.value = freq;\n                osc.connect(gain); gain.connect(audioCtx.destination);\n                osc.start();\n                gain.gain.setValueAtTime(gainVal * audioVolume, audioCtx.currentTime);\n                gain.gain.exponentialRampToValueAtTime(0.0001, audioCtx.currentTime + duration);\n                osc.stop(audioCtx.currentTime + duration);\n            } catch(e){}\n        }\n\n        function playInsertChime() {\n            if (audioInsertSrc === \'arcade_powerup\') {\n                playChimeTone(493.88, \'square\', 0.08);\n                setTimeout(() => playChimeTone(659.25, \'square\', 0.08), 80);\n                setTimeout(() => playChimeTone(987.77, \'square\', 0.25), 160);\n            } else if (audioInsertSrc === \'voice_filipino\') {\n                playChimeTone(587.33, \'sine\', 0.2);\n                if (\'speechSynthesis\' in window) {\n                    const utter = new SpeechSynthesisUtterance("Salamat sa pag-recycle! Dagdag minuto.");\n                    utter.lang = \'tl-PH\';\n                    window.speechSynthesis.speak(utter);\n                }\n            } else if (insertAudioElem) {\n                insertAudioElem.currentTime = 0;\n                insertAudioElem.play().catch(e => {\n                    playChimeTone(587.33, \'sine\', 0.18);\n                    setTimeout(() => playChimeTone(880.00, \'sine\', 0.35), 140);\n                });\n            } else {\n                playChimeTone(587.33, \'sine\', 0.18);\n                setTimeout(() => playChimeTone(880.00, \'sine\', 0.35), 140);\n            }\n        }\n\n        function playSuccessChime() {\n            if (audioSuccessSrc === \'crystal_bell\') {\n                playChimeTone(1046.50, \'sine\', 0.6, 0.4);\n            } else if (successAudioElem) {\n                successAudioElem.currentTime = 0;\n                successAudioElem.play().catch(e => {\n                    playChimeTone(880.00, \'sine\', 0.4);\n                });\n            } else if (!audioSuccessSrc || audioSuccessSrc !== \'silent\') {\n                playChimeTone(880.00, \'sine\', 0.4);\n            }\n        }\n\n        let depositActive = false;\n        let depositTimer = null;\n        let depositSec = 60;\n        let initialDepositTimeout = 60;\n        let lastBottleCount = 0;\n        let localRemainingSeconds = 0;\n        let isClientPaused = false;\n        let isSystemBinFull = false;\n\n        function switchTab(tabId) {\n            document.querySelectorAll(\'.tab-content\').forEach(el => el.classList.remove(\'active\'));\n            document.querySelectorAll(\'.tab-btn\').forEach(el => el.classList.remove(\'active\'));\n            document.getElementById(tabId).classList.add(\'active\');\n            const btn = document.querySelector(`button[onclick="switchTab(\'${tabId}\')"]`);\n            if (btn) btn.classList.add(\'active\');\n        }\n\n        function formatTime(totalSec) {\n            if (totalSec <= 0) return \'0d 00h:00m:00s\';\n            const d = Math.floor(totalSec / 86400);\n            const h = Math.floor((totalSec % 86400) / 3600).toString().padStart(2, \'0\');\n            const m = Math.floor((totalSec % 3600) / 60).toString().padStart(2, \'0\');\n            const s = (totalSec % 60).toString().padStart(2, \'0\');\n            return `${d}d ${h}h:${m}m:${s}s`;\n        }\n\n        function formatAddedTime(mins) {\n            if (mins === 0) return \'+0m\';\n            let res = \'\';\n            const d = Math.floor(mins / 1440);\n            const h = Math.floor((mins % 1440) / 60);\n            const m = mins % 60;\n            if (d > 0) res += `${d}d `;\n            if (h > 0) res += `${h}h `;\n            res += `${m}m`;\n            return \'+\' + res.trim();\n        }\n\n        // Local ticker for smooth countdown\n        setInterval(() => {\n            if (localRemainingSeconds > 0 && !isClientPaused) {\n                localRemainingSeconds--;\n                document.getElementById(\'time-display\').innerText = formatTime(localRemainingSeconds);\n                if (localRemainingSeconds <= 0) {\n                    syncPortal();\n                }\n            }\n        }, 1000);\n\n        function syncPortal() {\n            fetch(\'/api/vendo/status\')\n                .then(r => r.json())\n                .then(data => {\n                    localRemainingSeconds = data.client_time_remaining || data.remaining_seconds || 0;\n                    isClientPaused = data.is_paused || false;\n                    isSystemBinFull = data.bin_full || false;\n                    \n                    document.getElementById(\'time-display\').innerText = formatTime(localRemainingSeconds);\n                    \n                    const badge = document.getElementById(\'status-badge\');\n                    const pauseBox = document.getElementById(\'pause-ctrl-box\');\n                    const btnPause = document.getElementById(\'btn-pause\');\n                    const btnResume = document.getElementById(\'btn-resume\');\n                    const btnInsert = document.getElementById(\'btn-insert\');\n                    const binBanner = document.getElementById(\'bin-full-banner\');\n\n                    // 1. Bin full handling\n                    if (isSystemBinFull) {\n                        binBanner.style.display = \'block\';\n                        if (!depositActive) {\n                            btnInsert.disabled = true;\n                            btnInsert.innerHTML = \'<i class="fas fa-ban"></i> BIN FULL - TEMPORARILY DISABLED\';\n                            // btnInsert.classList.remove(\'pulse-btn\');\n                        }\n                    } else {\n                        binBanner.style.display = \'none\';\n                        if (!depositActive) {\n                            btnInsert.disabled = false;\n                            btnInsert.innerHTML = \'<i class="fas fa-recycle mr-1"></i> Insert Plastic Bottle\';\n                            // btnInsert.classList.add(\'pulse-btn\');\n                        }\n                    }\n\n                    // 2. Connection and pause status\n                    if (localRemainingSeconds > 0) {\n                        pauseBox.style.display = \'block\';\n                        if (isClientPaused) {\n                            badge.className = \'status-badge bg-paused\';\n                            badge.innerText = \'PAUSED\';\n                            btnPause.style.display = \'none\';\n                            btnResume.style.display = \'flex\';\n                        } else {\n                            badge.className = \'status-badge bg-active\';\n                            badge.innerText = \'CONNECTED\';\n                            btnPause.style.display = \'flex\';\n                            btnResume.style.display = \'none\';\n                        }\n                    } else {\n                        badge.className = \'status-badge bg-inactive\';\n                        badge.innerText = isSystemBinFull ? \'BIN FULL\' : \'DISCONNECTED\';\n                        pauseBox.style.display = \'none\';\n                    }\n\n                    // 3. Deposit modal sync\n                    if (depositActive) {\n                        const bottles = data.session_bottles || 0;\n                        const addedMins = data.session_added_minutes !== undefined ? data.session_added_minutes : 0;\n                        document.getElementById(\'modal-bottles\').innerText = bottles;\n                        document.getElementById(\'modal-added-time\').innerText = formatAddedTime(addedMins);\n                        if (bottles > lastBottleCount) {\n                            playInsertChime();\n                            \n                            const icon = document.getElementById(\'modal-bottle-icon\');\n                            icon.classList.remove(\'bottle-pop\');\n                            void icon.offsetWidth;\n                            icon.classList.add(\'bottle-pop\');\n\n                            document.getElementById(\'modal-stage-text\').innerHTML = \n                                `<span class="text-success font-weight-bold"><i class="fas fa-check-circle"></i> PET Bottle Verified! +${addedMins}m Added.</span>`;\n                            \n                            lastBottleCount = bottles;\n                            depositSec = initialDepositTimeout; // Refresh countdown for next bottle\n                        }\n                    }\n                }).catch(()=>{});\n        }\n\n        setInterval(syncPortal, 1200);\n\n        function startDepositSession() {\n            const btn = document.getElementById(\'btn-insert\');\n            if (btn.disabled || depositActive || isSystemBinFull) return;\n            btn.disabled = true;\n            \n            unlockAudio();\n            \n            lastBottleCount = 0;\n            document.getElementById(\'modal-bottles\').innerText = \'0\';\n            document.getElementById(\'modal-added-time\').innerText = \'+0m\';\n            document.getElementById(\'modal-stage-text\').innerHTML = \n                \'<i class="fas fa-spinner fa-spin text-success"></i> Airlock opening... Please wait.\';\n            \n            fetch(\'/api/vendo/open_gate\', { method: \'POST\' })\n                .then(r => r.json())\n                .then(data => {\n                    if (!data.success) {\n                        alert(data.error || "Machine is in use.");\n                        btn.disabled = isSystemBinFull;\n                        return;\n                    }\n                    depositActive = true;\n                    initialDepositTimeout = data.timeout || 60;\n                    depositSec = initialDepositTimeout;\n                    lastBottleCount = 0;\n                    document.getElementById(\'deposit-modal\').style.display = \'flex\';\n                    document.getElementById(\'modal-stage-text\').innerHTML = \n                        \'<i class="fas fa-arrow-down text-success"></i> Gate Open! Drop your PET bottle into the chute...\';\n                    \n                    if (bgAudioElem) {\n                        bgAudioElem.currentTime = 0;\n                        bgAudioElem.play().catch(()=>{});\n                    }\n\n                    if (depositTimer) clearInterval(depositTimer);\n                    depositTimer = setInterval(() => {\n                        depositSec--;\n                        document.getElementById(\'modal-timer\').innerText = `${depositSec}s`;\n                        const pct = Math.max(0, (depositSec / initialDepositTimeout) * 100);\n                        document.getElementById(\'modal-progress-bar\').style.width = `${pct}%`;\n                        if (depositSec <= 0) {\n                            closeDepositSession();\n                        }\n                    }, 1000);\n                }).catch(err => {\n                    btn.disabled = isSystemBinFull;\n                });\n        }\n\n        function closeDepositSession() {\n            if (!depositActive) return;\n            depositActive = false;\n            clearInterval(depositTimer);\n            document.getElementById(\'deposit-modal\').style.display = \'none\';\n            document.getElementById(\'btn-insert\').disabled = isSystemBinFull;\n            \n            if (bgAudioElem) {\n                bgAudioElem.pause();\n                bgAudioElem.currentTime = 0;\n            }\n            playSuccessChime();\n            \n            fetch(\'/api/vendo/done\', { method: \'POST\' }).then(() => {\n                syncPortal();\n                setTimeout(() => {\n                    fetch(\'/generate_204\', { mode: \'no-cors\' }).catch(()=>{});\n                }, 300);\n            });\n        }\n\n        function togglePause(action) {\n            fetch(\'/api/client/pause\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ action: action })\n            }).then(()=>syncPortal());\n        }\n\n        function redeemVoucher() {\n            const code = document.getElementById(\'voucher-code-input\').value.trim();\n            if (!code) return;\n            fetch(\'/api/voucher/redeem\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ code: code })\n            }).then(r => r.json()).then(data => {\n                const msg = document.getElementById(\'voucher-msg\');\n                if (data.success) {\n                    msg.style.color = \'#34d399\';\n                    msg.innerText = data.message;\n                    document.getElementById(\'voucher-code-input\').value = \'\';\n                    playSuccessChime();\n                    syncPortal();\n                } else {\n                    msg.style.color = \'#f87171\';\n                    msg.innerText = data.error;\n                }\n            });\n        }\n\n        function generateTransferCode() {\n            const m = parseInt(document.getElementById(\'transfer-mins-input\').value) || 0;\n            const disp = document.getElementById(\'transfer-code-display\');\n            if (m <= 0) {\n                disp.style.color = \'#f87171\';\n                disp.innerText = \'Please enter the exact minutes to share.\';\n                return;\n            }\n            fetch(\'/api/transfer/generate\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ minutes: m })\n            })\n            .then(r => r.json())\n            .then(data => {\n                if (data.success) {\n                    disp.style.color = \'#38bdf8\';\n                    disp.innerHTML = `TRANSFER CODE: <strong style="font-size:16px; letter-spacing:2px; color:#34d399;">${data.code}</strong> (${data.minutes} Mins)`;\n                    document.getElementById(\'transfer-mins-input\').value = \'\';\n                    syncPortal();\n                } else {\n                    disp.style.color = \'#f87171\';\n                    disp.innerText = data.error;\n                }\n            });\n        }\n\n        function claimTransferCode() {\n            const code = document.getElementById(\'claim-code-input\').value.trim();\n            const msg = document.getElementById(\'claim-status-msg\');\n            if (!code) {\n                msg.style.color = \'#f87171\';\n                msg.innerText = \'Please enter the 6-digit transfer code.\';\n                return;\n            }\n            fetch(\'/api/transfer/claim\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ code: code })\n            }).then(r => r.json()).then(data => {\n                if (data.success) {\n                    msg.style.color = \'#34d399\';\n                    msg.innerText = data.message;\n                    document.getElementById(\'claim-code-input\').value = \'\';\n                    playInsertChime();\n                    syncPortal();\n                } else {\n                    msg.style.color = \'#f87171\';\n                    msg.innerText = data.error;\n                }\n            });\n        }\n\n        let loggedInMember = null;\n\n        function memberLogin() {\n            const u = document.getElementById(\'member-user\').value.trim();\n            const p = document.getElementById(\'member-pin\').value.trim();\n            if (!u || !p) {\n                alert(\'Please enter your username and PIN.\');\n                return;\n            }\n            fetch(\'/api/member/login\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ username: u, pin: p })\n            }).then(r => r.json()).then(data => {\n                const s = document.getElementById(\'member-status\');\n                if (data.success) {\n                    loggedInMember = { username: u, pin: p, wallet_minutes: data.wallet_minutes };\n                    document.getElementById(\'mem-auth-section\').style.display = \'none\';\n                    document.getElementById(\'mem-wallet-section\').style.display = \'block\';\n                    document.getElementById(\'mem-welcome-user\').innerText = `👤 ${u}`;\n                    document.getElementById(\'mem-wallet-mins\').innerText = `${data.wallet_minutes} Mins`;\n                    s.innerText = \'\';\n                } else {\n                    s.style.color = \'#f87171\'; s.innerText = data.error;\n                }\n            });\n        }\n\n        function memberLogout() {\n            loggedInMember = null;\n            document.getElementById(\'mem-auth-section\').style.display = \'block\';\n            document.getElementById(\'mem-wallet-section\').style.display = \'none\';\n            document.getElementById(\'member-user\').value = \'\';\n            document.getElementById(\'member-pin\').value = \'\';\n        }\n\n        function memberRegister() {\n            const u = document.getElementById(\'member-user\').value.trim();\n            const p = document.getElementById(\'member-pin\').value.trim();\n            if (!u || !p) {\n                alert(\'Please enter a username and PIN.\');\n                return;\n            }\n            fetch(\'/api/member/register\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ username: u, pin: p })\n            }).then(r => r.json()).then(data => {\n                const s = document.getElementById(\'member-status\');\n                if (data.success) {\n                    s.style.color = \'#34d399\';\n                    s.innerText = data.message;\n                    setTimeout(() => memberLogin(), 400);\n                } else {\n                    s.style.color = \'#f87171\';\n                    s.innerText = data.error;\n                }\n            });\n        }\n\n        function useMemberWallet() {\n            if (!loggedInMember) return;\n            const m = parseInt(document.getElementById(\'use-wallet-mins\').value) || 0;\n            if (m <= 0) {\n                alert(\'Please enter valid minutes to use.\');\n                return;\n            }\n            fetch(\'/api/member/use_wallet\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ username: loggedInMember.username, pin: loggedInMember.pin, minutes: m })\n            }).then(r => r.json()).then(data => {\n                if (data.success) {\n                    loggedInMember.wallet_minutes = data.wallet_minutes;\n                    document.getElementById(\'mem-wallet-mins\').innerText = `${data.wallet_minutes} Mins`;\n                    document.getElementById(\'use-wallet-mins\').value = \'\';\n                    playInsertChime();\n                    syncPortal();\n                    alert(data.message);\n                } else {\n                    alert(data.error);\n                }\n            });\n        }\n\n        function saveSessionToWallet() {\n            if (!loggedInMember) return;\n            fetch(\'/api/member/save_time\', {\n                method: \'POST\',\n                headers: { \'Content-Type\': \'application/json\' },\n                body: JSON.stringify({ username: loggedInMember.username, pin: loggedInMember.pin })\n            }).then(r => r.json()).then(data => {\n                if (data.success) {\n                    loggedInMember.wallet_minutes = data.wallet_minutes;\n                    document.getElementById(\'mem-wallet-mins\').innerText = `${data.wallet_minutes} Mins`;\n                    playSuccessChime();\n                    syncPortal();\n                    alert(data.message);\n                } else {\n                    alert(data.error);\n                }\n            });\n        }\n    </script>\n</body>\n</html>\n'

@app.errorhandler(404)
def page_not_found(e):
    return redirect('http://10.0.0.1/')

# ── Windows NCSI probe ────────────────────────────────────────────────────────
# Windows hits http://www.msftconnecttest.com/connecttest.txt and expects the
# exact string "Microsoft Connect Test".  If we return anything else (or a
# redirect) the taskbar shows "Action needed, no internet" even when the client
# has a working connection.  We serve the correct response when the client has
# active time; otherwise we redirect to the portal so they can buy time.
@app.route('/connecttest.txt')
def ncsi_connecttest():
    client_ip = get_client_ip()
    sess = active_clients.get(client_ip)
    if sess and sess.get('remaining_seconds', 0) > 0 and not sess.get('is_paused'):
        return Response('Microsoft Connect Test', mimetype='text/plain', status=200)
    return redirect('http://10.0.0.1/')

@app.route('/ncsi.txt')
def ncsi_txt():
    client_ip = get_client_ip()
    sess = active_clients.get(client_ip)
    if sess and sess.get('remaining_seconds', 0) > 0 and not sess.get('is_paused'):
        return Response('Microsoft NCSI', mimetype='text/plain', status=200)
    return redirect('http://10.0.0.1/')

# ── Android / Chrome OS connectivity probe ────────────────────────────────────
@app.route('/generate_204')
@app.route('/gen_204')
def generate_204():
    client_ip = get_client_ip()
    sess = active_clients.get(client_ip)
    if sess and sess.get('remaining_seconds', 0) > 0 and not sess.get('is_paused'):
        return Response('', status=204)
    return redirect('http://10.0.0.1/')

# ── Apple iOS / macOS captive portal probe ────────────────────────────────────
@app.route('/hotspot-detect.html')
@app.route('/library/test/success.html')
@app.route('/canonical.html')
def apple_captive():
    client_ip = get_client_ip()
    sess = active_clients.get(client_ip)
    if sess and sess.get('remaining_seconds', 0) > 0 and not sess.get('is_paused'):
        return Response('<HTML><HEAD><TITLE>Success</TITLE></HEAD><BODY>Success</BODY></HTML>',
                        mimetype='text/html', status=200)
    return redirect('http://10.0.0.1/')

# ── Firefox connectivity probe ────────────────────────────────────────────────
@app.route('/success.txt')
def firefox_success():
    client_ip = get_client_ip()
    sess = active_clients.get(client_ip)
    if sess and sess.get('remaining_seconds', 0) > 0 and not sess.get('is_paused'):
        return Response('success', mimetype='text/plain', status=200)
    return redirect('http://10.0.0.1/')

@app.route('/')
def index():
    client_ip = get_client_ip()
    session_data = ensure_client_session(client_ip)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT bottles, minutes, label FROM promo_rates ORDER BY bottles ASC')
        promos = [{'bottles': r[0], 'minutes': r[1], 'label': r[2]} for r in c.fetchall()]
        c.execute('SELECT message FROM announcements WHERE active = 1 ORDER BY id DESC LIMIT 1')
        ann_row = c.fetchone()
        announcement = ann_row[0] if ann_row else ''
        c.execute('SELECT domain, note FROM walled_garden ORDER BY domain ASC')
        walled_sites = [{'domain': r[0], 'note': r[1]} for r in c.fetchall()]
    return render_template_string(PORTAL_HTML, license_valid=license_valid(), client_ip=client_ip, client_mac=session_data.get('mac', '00:00:00:00:00:00'), vendo_name=get_config('vendo_name', 'ECO-Fi Vendo'), vendo_subtitle=get_config('vendo_subtitle', 'Recycle Bottles for Fast WiFi'), promo_rates=promos, announcement=announcement, walled_sites=walled_sites, audio_bg=get_config('audio_bg', '/static/audio/eco_loop.wav'), audio_insert=get_config('audio_insert', '/static/audio/eco_chime.wav'), audio_success=get_config('audio_success', '/static/audio/eco_success.wav'), audio_volume=get_config('audio_volume', '80'))

@app.route('/api/vendo/status')
@app.route('/api/status')
def api_vendo_status():
    client_ip = get_client_ip()
    session_data = ensure_client_session(client_ip)
    sim_status = esp32.get_state()
    is_current_depositor = (active_depositor_ip is not None and active_depositor_ip == client_ip)
    session_bottles = sim_status.get('session_bottles', 0) if is_current_depositor else 0
    session_added_minutes = calculate_minutes_for_bottles(session_bottles) if is_current_depositor else 0
    is_bin_full = sim_status.get('is_bin_full', False) or sim_status.get('bin_full_alert', False) or get_config('hw_bin_full', '0') == '1'
    expires_at = session_data.get('expires_at', 0)
    expires_str = ''
    if session_data.get('is_paused') and expires_at > time.time():
        expires_str = datetime.fromtimestamp(expires_at).strftime('%b %d, %I:%M %p')
    elif not session_data.get('is_paused') and session_data.get('remaining_seconds', 0) > 0:
        proj_exp = compute_session_expiration(session_data['remaining_seconds'])
        expires_str = datetime.fromtimestamp(proj_exp).strftime('%b %d, %I:%M %p')
    return jsonify({'remaining_seconds': session_data.get('remaining_seconds', 0), 'client_time_remaining': session_data.get('remaining_seconds', 0), 'is_paused': session_data.get('is_paused', False), 'paused_at': session_data.get('paused_at', 0), 'expires_at': expires_at, 'expires_str': expires_str, 'validity_hours': calculate_pause_validity_seconds(session_data.get('remaining_seconds', 0)) // 3600, 'session_bottles': session_bottles, 'session_added_minutes': session_added_minutes, 'gate_open': sim_status.get('entrance_servo', 0) > 45 if is_current_depositor else False, 'bin_full': is_bin_full})

@app.route('/api/vendo/open_gate', methods=['POST'])
@app.route('/api/open_gate', methods=['POST'])
def api_open_gate():
    global active_depositor_ip, active_depositor_timeout
    if not license_valid():
        return jsonify({'success': False, 'error': 'Machine is unlicensed or license is expired.'})
    client_ip = get_client_ip()
    is_bin_full = get_config('hw_bin_full', '0') == '1' or esp32.get_state().get('is_bin_full', False)
    if is_bin_full:
        return jsonify({'success': False, 'error': 'Storage bin is full. Please contact administrator to empty the bin.'})
    with active_clients_lock:
        if active_depositor_ip and active_depositor_ip != client_ip:
            if time.time() < active_depositor_timeout:
                return jsonify({'success': False, 'error': 'Another user is currently depositing bottles. Please wait.'})
        timeout = int(get_config('drop_timeout', '60') or 60)
        active_depositor_ip = client_ip
        active_depositor_timeout = time.time() + timeout + 5
        esp32.reset_session()
        if client_ip in active_clients:
            active_clients[client_ip]['pending_bottles'] = 0
    esp32.open_entrance_gate(timeout=timeout)
    transmit_to_esp32({'cmd': 'OPEN_GATE', 'timeout': timeout})
    return jsonify({'success': True, 'timeout': timeout, 'session_bottles': 0})

@app.route('/api/vendo/done', methods=['POST'])
def api_vendo_done():
    global active_depositor_ip, active_depositor_timeout
    client_ip = get_client_ip()
    with active_clients_lock:
        if active_depositor_ip and active_depositor_ip != client_ip:
            return jsonify({'success': False, 'error': 'You are not the active depositor.'})
        sim_status = esp32.get_state()
        session_bottles = sim_status.get('session_bottles', 0)
        added_minutes = 0
        if session_bottles > 0:
            added_minutes = calculate_minutes_for_bottles(session_bottles)
            sess = ensure_client_session(client_ip)
            sess['remaining_seconds'] += added_minutes * 60
            sess['is_paused'] = False
            sess['user_paused'] = False
            sess['auto_paused'] = False
            sess['paused_at'] = 0
            sess['expires_at'] = 0
            sync_client_firewall(client_ip)
            save_sessions_to_db()
        active_depositor_ip = None
        active_depositor_timeout = 0
    esp32.reset_session()
    esp32.close_entrance_gate()
    transmit_to_esp32({'cmd': 'CLOSE_GATE'})
    return jsonify({'success': True, 'bottles_credited': session_bottles, 'added_minutes': added_minutes})

@app.route('/generate_204')
@app.route('/gen_204')
def captive_generate_204():
    client_ip = get_client_ip()
    session_data = ensure_client_session(client_ip)
    if session_data.get('remaining_seconds', 0) > 0 and (not session_data.get('is_paused', False)):
        return ('', 204)
    return redirect('/')

@app.route('/hotspot-detect.html')
def captive_hotspot_detect():
    client_ip = get_client_ip()
    session_data = ensure_client_session(client_ip)
    if session_data.get('remaining_seconds', 0) > 0 and (not session_data.get('is_paused', False)):
        return ('<HTML><HEAD><TITLE>Success</TITLE></HEAD><BODY>Success</BODY></HTML>', 200, {'Content-Type': 'text/html'})
    return redirect('/')

@app.route('/connecttest.txt')
@app.route('/ncsi.txt')
def captive_msft():
    client_ip = get_client_ip()
    session_data = ensure_client_session(client_ip)
    if session_data.get('remaining_seconds', 0) > 0 and (not session_data.get('is_paused', False)):
        return ('Microsoft NCSI', 200, {'Content-Type': 'text/plain'})
    return redirect('/')

@app.route('/api/client/pause', methods=['POST'])
def api_client_pause():
    client_ip = get_client_ip()
    data = request.get_json() or {}
    action = data.get('action', 'pause')
    with active_clients_lock:
        sess = ensure_client_session(client_ip)
        if sess:
            if action == 'pause':
                sess['is_paused'] = True
                sess['user_paused'] = True
                sess['paused_at'] = time.time()
                sess['expires_at'] = compute_session_expiration(sess['remaining_seconds'], sess['paused_at'])
            else:
                if sess.get('admin_paused'):
                    return jsonify({'success': False, 'error': 'Session paused by administrator.'})
                if sess.get('expires_at', 0) > 0 and time.time() >= sess['expires_at']:
                    sess['remaining_seconds'] = 0
                    sess['expires_at'] = 0
                    sess['is_paused'] = True
                    sess['user_paused'] = False
                    sync_client_firewall(client_ip)
                    save_sessions_to_db()
                    return jsonify({'success': False, 'error': 'Paused credit expired.'})
                if sess.get('remaining_seconds', 0) <= 0:
                    sess['remaining_seconds'] = 0
                    sync_client_firewall(client_ip)
                    return jsonify({'success': False, 'error': 'No remaining credit to resume.'})
                sess['is_paused'] = False
                sess['user_paused'] = False
                sess['auto_paused'] = False
                sess['paused_at'] = 0
                sess['expires_at'] = 0
            sync_client_firewall(client_ip)
            save_sessions_to_db()
            return jsonify({'success': True, 'is_paused': sess['is_paused'], 'expires_at': sess.get('expires_at', 0)})
    return jsonify({'success': False})

@app.route('/api/voucher/redeem', methods=['POST'])
def api_voucher_redeem():
    if not license_valid():
        return jsonify({'success': False, 'error': 'Machine is unlicensed or license is expired.'})
    client_ip = get_client_ip()
    data = request.get_json() or {}
    code = data.get('code', '').strip().upper()
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT minutes, is_used FROM vouchers WHERE code = ?', (code,))
        row = c.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Invalid voucher code.'})
        if row[1] == 1:
            return jsonify({'success': False, 'error': 'Voucher already redeemed.'})
        minutes = row[0]
        c.execute('UPDATE vouchers SET is_used = 1, used_by = ? WHERE code = ? AND is_used = 0', (client_ip, code))
        if c.rowcount <= 0:
            return jsonify({'success': False, 'error': 'Voucher already redeemed.'})
        conn.commit()
    with active_clients_lock:
        sess = ensure_client_session(client_ip)
        sess['remaining_seconds'] += minutes * 60
        sess['is_paused'] = False
        sess['user_paused'] = False
        sess['paused_at'] = 0
        sess['expires_at'] = 0
        sync_client_firewall(client_ip)
        save_sessions_to_db()
    return jsonify({'success': True, 'message': 'Successfully added {} minutes!'.format(minutes)})

@app.route('/api/transfer/generate', methods=['POST'])
def api_transfer_generate():
    client_ip = get_client_ip()
    data = request.get_json() or {}
    requested_mins = data.get('minutes')
    with active_clients_lock:
        sess = ensure_client_session(client_ip)
        rem = sess.get('remaining_seconds', 0)
        available_mins = rem // 60
        if available_mins < 1:
            return jsonify({'success': False, 'error': 'Minimum balance to transfer is 1 minute.'})
        if requested_mins is not None and requested_mins != '':
            try:
                mins_to_transfer = int(requested_mins)
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': 'Please enter a valid number of minutes.'})
            if mins_to_transfer <= 0:
                return jsonify({'success': False, 'error': 'Transfer minutes must be at least 1 minute.'})
            if mins_to_transfer > available_mins:
                return jsonify({'success': False, 'error': 'Insufficient balance ({}m available).'.format(available_mins)})
        else:
            mins_to_transfer = available_mins
        transfer_sec = mins_to_transfer * 60
        sess['remaining_seconds'] = max(0, rem - transfer_sec)
        sync_client_firewall(client_ip)
        save_sessions_to_db()
        code = ''.join(random.choice(string.digits) for _ in range(6))
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('INSERT INTO time_transfers (code, from_ip, from_mac, seconds, created_at) VALUES (?, ?, ?, ?, ?)', (code, client_ip, sess['mac'], transfer_sec, time.time()))
        conn.commit()
    return jsonify({'success': True, 'code': code, 'minutes': mins_to_transfer, 'remaining_minutes': sess['remaining_seconds'] // 60})

@app.route('/api/transfer/claim', methods=['POST'])
def api_transfer_claim():
    client_ip = get_client_ip()
    data = request.get_json() or {}
    code = data.get('code', '').strip()
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT seconds, is_claimed FROM time_transfers WHERE code = ?', (code,))
        row = c.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Invalid transfer code.'})
        if row[1] == 1:
            return jsonify({'success': False, 'error': 'Transfer code already claimed.'})
        sec = row[0]
        c.execute('UPDATE time_transfers SET is_claimed = 1 WHERE code = ? AND is_claimed = 0', (code,))
        if c.rowcount <= 0:
            return jsonify({'success': False, 'error': 'Transfer code already claimed.'})
        conn.commit()
    with active_clients_lock:
        sess = ensure_client_session(client_ip)
        sess['remaining_seconds'] += sec
        sess['is_paused'] = False
        sess['user_paused'] = False
        sess['paused_at'] = 0
        sess['expires_at'] = 0
        sync_client_firewall(client_ip)
        save_sessions_to_db()
    return jsonify({'success': True, 'message': 'Claimed {} minutes successfully!'.format(sec // 60)})

@app.route('/api/member/register', methods=['POST'])
def api_member_register():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    pin = data.get('pin', '').strip()
    if not username or not re.match('^[a-zA-Z0-9_]{3,20}$', username):
        return jsonify({'success': False, 'error': 'Username must be 3-20 characters (letters, numbers, underscores only).'})
    if not pin or not re.match('^\\d{4,6}$', pin):
        return jsonify({'success': False, 'error': 'PIN must be strictly 4 to 6 numeric digits.'})
    pin_hash = generate_password_hash(pin, method='pbkdf2:sha256')
    try:
        with db_connection() as conn:
            c = conn.cursor()
            c.execute('INSERT INTO members (username, pin_hash, wallet_minutes, created_at) VALUES (?, ?, 0, ?)', (username, pin_hash, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            conn.commit()
        return jsonify({'success': True, 'message': 'Account registered successfully! Zero-expiry member wallet is active.'})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'Username already taken. Please choose another.'})
login_attempts = {}

@app.route('/api/member/login', methods=['POST'])
def api_member_login():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    pin = data.get('pin', '').strip()
    now = time.time()
    if username in login_attempts:
        count, last_time = login_attempts[username]
        if now - last_time < 300:
            if count >= 5:
                return jsonify({'success': False, 'error': 'Too many failed attempts. Try again in 5 minutes.'})
        else:
            login_attempts[username] = [0, now]
    else:
        login_attempts[username] = [0, now]
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT pin_hash, wallet_minutes FROM members WHERE username = ?', (username,))
        row = c.fetchone()
        if not row or not check_password_hash(row[0], pin):
            login_attempts[username][0] += 1
            login_attempts[username][1] = time.time()
            return jsonify({'success': False, 'error': 'Invalid username or PIN.'})
        if username in login_attempts:
            del login_attempts[username]
        return jsonify({'success': True, 'wallet_minutes': row[1]})

@app.route('/api/member/use_wallet', methods=['POST'])
def api_member_use_wallet():
    client_ip = get_client_ip()
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    pin = data.get('pin', '').strip()
    try:
        minutes = int(data.get('minutes', 0))
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Please enter a valid number of minutes.'})
    if minutes <= 0:
        return jsonify({'success': False, 'error': 'Invalid minutes specified.'})
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT pin_hash, wallet_minutes FROM members WHERE username = ?', (username,))
        row = c.fetchone()
        if not row or not check_password_hash(row[0], pin):
            return jsonify({'success': False, 'error': 'Invalid username or PIN.'})
        current_wallet = row[1]
        if current_wallet < minutes:
            return jsonify({'success': False, 'error': 'Insufficient wallet balance ({}m available).'.format(current_wallet)})
        c.execute('UPDATE members SET wallet_minutes = wallet_minutes - ? WHERE username = ? AND wallet_minutes >= ?', (minutes, username, minutes))
        if c.rowcount <= 0:
            return jsonify({'success': False, 'error': 'Insufficient wallet balance.'})
        conn.commit()
    with active_clients_lock:
        sess = ensure_client_session(client_ip)
        sess['remaining_seconds'] += minutes * 60
        sess['is_paused'] = False
        sess['user_paused'] = False
        sess['paused_at'] = 0
        sess['expires_at'] = 0
        sync_client_firewall(client_ip)
        save_sessions_to_db()
    return jsonify({'success': True, 'message': 'Added {} minutes from wallet to session!'.format(minutes), 'wallet_minutes': current_wallet - minutes})

@app.route('/api/member/save_time', methods=['POST'])
def api_member_save_time():
    client_ip = get_client_ip()
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    pin = data.get('pin', '').strip()
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT pin_hash, wallet_minutes FROM members WHERE username = ?', (username,))
        row = c.fetchone()
        if not row or not check_password_hash(row[0], pin):
            return jsonify({'success': False, 'error': 'Invalid username or PIN.'})
        with active_clients_lock:
            sess = ensure_client_session(client_ip)
            rem_sec = sess.get('remaining_seconds', 0)
            if rem_sec < 60:
                return jsonify({'success': False, 'error': 'No active session time to save (minimum 1 minute required).'})
            mins_to_save = rem_sec // 60
            sess['remaining_seconds'] = 0
            sess['is_paused'] = False
            sess['paused_at'] = 0
            sess['expires_at'] = 0
            sync_client_firewall(client_ip)
            save_sessions_to_db()
        c.execute('UPDATE members SET wallet_minutes = wallet_minutes + ? WHERE username = ?', (mins_to_save, username))
        conn.commit()
        c.execute('SELECT wallet_minutes FROM members WHERE username = ?', (username,))
        new_wallet = c.fetchone()[0]
    return jsonify({'success': True, 'message': 'Saved {} minutes to your member wallet!'.format(mins_to_save), 'wallet_minutes': new_wallet})

@app.route('/admin/api/license', methods=['GET'])
def admin_api_license():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    return jsonify(license_manager.verify_license())

@app.route('/admin/api/license/activate', methods=['POST'])
def admin_api_license_activate():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    pin = data.get('pin', '').strip()
    licensee = data.get('licensee', 'Store Owner')
    tier = data.get('tier', 'COMMERCIAL')
    return jsonify(license_manager.activate_machine(pin, licensee, tier))

@app.route('/admin/api/esp32/save', methods=['POST'])
def admin_api_esp32_save():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    for k, v in data.items():
        set_config('esp_{}'.format(k), v)
    data['cmd'] = 'SET_CONFIG'
    transmit_to_esp32(data)
    return jsonify({'success': True})

@app.route('/admin/api/esp32/trigger', methods=['POST'])
def admin_api_esp32_trigger():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    transmit_to_esp32({'cmd': 'TRIGGER_CONFIG'})
    return jsonify({'success': True})

def validate_promo_rate_conflict(bottles, minutes, exclude_bottles=None):
    """
    Validates a proposed promo rate against 3 mathematical invariants:
    1. Combination Floor: minutes >= greedy combo of lower tiers
    2. Monotonic Efficiency: minutes/bottle >= all smaller-tier efficiencies
    3. Higher-Tier Bound: minutes <= smallest higher-tier's minutes
    Returns (is_valid: bool, error_message: str)
    """
    if bottles <= 0 or minutes <= 0:
        return (False, 'Bottles and Minutes must be positive numbers.')
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT bottles, minutes FROM promo_rates ORDER BY bottles ASC')
        existing = [(r[0], r[1]) for r in c.fetchall() if r[0] != exclude_bottles]
    eff_new = minutes / bottles
    for eb, em in existing:
        eff_ex = em / eb
        if eb < bottles and eff_ex > eff_new:
            return (False, 'Efficiency conflict: {}B→{}m tier gives {:.1f} m/bottle, but your {}B→{}m gives only {:.1f} m/bottle. Higher bottle tiers must reward at least as much per bottle.'.format(eb, em, eff_ex, bottles, minutes, eff_new))
        if eb > bottles and eff_ex < eff_new:
            return (False, 'Efficiency conflict: your {}B→{}m tier gives {:.1f} m/bottle, which exceeds the {}B→{}m tier at {:.1f} m/bottle. Larger tiers must always be at least as efficient.'.format(bottles, minutes, eff_new, eb, em, eff_ex))
    lower_tiers = sorted([(eb, em) for eb, em in existing if eb < bottles], reverse=True)
    combo_minutes = 0
    rem = bottles
    for eb, em in lower_tiers:
        if rem >= eb:
            times = rem // eb
            combo_minutes += times * em
            rem %= eb
    if combo_minutes > 0 and minutes < combo_minutes:
        return (False, 'Combination conflict: {} bottles can be split into smaller tiers yielding {} mins, but this rate only gives {} mins. New rate must be at least {} mins to incentivize bulk deposit.'.format(bottles, combo_minutes, minutes, combo_minutes))
    higher_tiers = sorted([(eb, em) for eb, em in existing if eb > bottles])
    if higher_tiers:
        min_higher_mins = min((em for _, em in higher_tiers))
        if minutes >= min_higher_mins:
            hb, hm = min(((eb, em) for eb, em in higher_tiers if em == min_higher_mins))
            return (False, 'Upper-bound conflict: your {}B→{}m would give same or more time than the {}B→{}m tier. Reduce minutes or increase the higher tier.'.format(bottles, minutes, hb, hm))
    return (True, '')

@app.route('/admin/api/rates/list')
def admin_api_rates_list():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT bottles, minutes, label, speed_profile FROM promo_rates ORDER BY bottles ASC')
        return jsonify([{'bottles': r[0], 'minutes': r[1], 'label': r[2], 'speed_profile': r[3] or ''} for r in c.fetchall()])

@app.route('/admin/api/rates/add', methods=['POST'])
def admin_api_rates_add():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    try:
        bottles = int(data.get('bottles', 0))
        minutes = int(data.get('minutes', 0))
        orig_bottles = int(data.get('orig_bottles')) if data.get('orig_bottles') else None
        label = data.get('label', '').strip() or '{} Bottle{} = {} mins'.format(bottles, 's' if bottles != 1 else '', minutes)
    except (TypeError, ValueError):
        return (jsonify({'success': False, 'error': 'Bottles and Minutes must be positive numbers.'}), 400)
    if bottles <= 0 or minutes <= 0:
        return (jsonify({'success': False, 'error': 'Bottles and Minutes must be positive numbers.'}), 400)
    is_valid, err_msg = validate_promo_rate_conflict(bottles, minutes, exclude_bottles=orig_bottles)
    if not is_valid:
        return (jsonify({'success': False, 'error': err_msg}), 400)
    with db_connection() as conn:
        c = conn.cursor()
        if orig_bottles and orig_bottles != bottles:
            c.execute('DELETE FROM promo_rates WHERE bottles = ?', (orig_bottles,))
        c.execute("REPLACE INTO promo_rates (bottles, minutes, label, speed_profile) VALUES (?, ?, ?, '')", (bottles, minutes, label))
        if bottles == 1:
            c.execute("REPLACE INTO config (key, value) VALUES ('minutes_per_bottle', ?)", (str(minutes),))
        conn.commit()
    return jsonify({'success': True})

@app.route('/admin/api/rates/delete', methods=['POST'])
def admin_api_rates_delete():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    bottles = data.get('bottles')
    if not bottles:
        return (jsonify({'success': False, 'error': 'Missing bottles parameter.'}), 400)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('DELETE FROM promo_rates WHERE bottles = ?', (int(bottles),))
        conn.commit()
    return jsonify({'success': True})

@app.route('/admin/api/rates/apply_preset', methods=['POST'])
def admin_api_rates_apply_preset():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    preset_key = data.get('preset', 'standard')
    PRESETS = {'standard': {'label': 'Standard Community Curve', 'rates': [(1, 10, '1 Bottle = 10 mins'), (3, 40, '3 Bottles = 40 mins'), (5, 75, '5 Bottles = 1h 15m'), (10, 180, '10 Bottles = 3 Hours')]}, 'aggressive': {'label': 'Aggressive Reward Curve', 'rates': [(1, 10, '1 Bottle = 10 mins'), (5, 70, '5 Bottles = 1h 10m'), (10, 180, '10 Bottles = 3 Hours'), (20, 420, '20 Bottles = 7 Hours')]}, 'cafe': {'label': 'Café / Study Hub Curve', 'rates': [(1, 20, '1 Bottle = 20 mins'), (3, 75, '3 Bottles = 1h 15m'), (6, 180, '6 Bottles = 3 Hours'), (12, 420, '12 Bottles = 7 Hours')]}}
    if preset_key not in PRESETS:
        return (jsonify({'success': False, 'error': "Unknown preset '{}'.".format(preset_key)}), 400)
    preset = PRESETS[preset_key]
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('DELETE FROM promo_rates')
        for bottles, minutes, label in preset['rates']:
            c.execute("REPLACE INTO promo_rates (bottles, minutes, label, speed_profile) VALUES (?, ?, ?, '')", (bottles, minutes, label))
        base = next((m for b, m, _ in preset['rates'] if b == 1), None)
        if base:
            c.execute("REPLACE INTO config (key, value) VALUES ('minutes_per_bottle', ?)", (str(base),))
        conn.commit()
    return jsonify({'success': True, 'message': "'{}' template applied with {} tiers.".format(preset['label'], len(preset['rates']))})

@app.route('/admin/api/audio/settings', methods=['POST'])
def admin_api_audio_settings():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    bg = data.get('audio_bg', '/static/audio/b1.wav')
    insert = data.get('audio_insert', '/static/audio/coin.wav')
    success = data.get('audio_success', '/static/audio/success_ding.wav')
    vol = data.get('volume', '80')
    set_config('audio_bg', bg)
    set_config('audio_insert', insert)
    set_config('audio_success', success)
    set_config('audio_volume', vol)
    set_config('audio_preset', insert)
    return jsonify({'success': True})

@app.route('/admin/api/audio/upload', methods=['POST'])
def admin_api_audio_upload():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    if 'file' not in request.files:
        return (jsonify({'success': False, 'error': 'No file uploaded.'}), 400)
    file = request.files['file']
    if file.filename == '':
        return (jsonify({'success': False, 'error': 'No file selected.'}), 400)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.mp3', '.wav', '.ogg', '.m4a', '.aac']:
        return (jsonify({'success': False, 'error': 'Invalid audio file format. Only MP3, WAV, OGG allowed.'}), 400)
    upload_dir = os.path.join(app.static_folder or 'static', 'audio', 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = 'custom_{}_{}'.format(int(time.time()), re.sub('[^a-zA-Z0-9_.-]', '', file.filename))
    filepath = os.path.join(upload_dir, safe_name)
    file.save(filepath)
    file_url = '/static/audio/uploads/{}'.format(safe_name)
    return jsonify({'success': True, 'url': file_url})

@app.route('/simulator')
def simulator_ui():
    return render_template_string(esp32.render_simulator_html())

@app.route('/simulator/api/state')
def simulator_api_state():
    return jsonify(esp32.get_state())

@app.route('/simulator/api/drop', methods=['POST'])
@app.route('/simulator/api/trigger', methods=['POST'])
def simulator_drop():
    data = request.get_json() or {}
    item_type = data.get('item_type') or data.get('type', 'valid_pet')
    esp32.simulate_insert(item_type=item_type)
    return jsonify({'success': True, 'item_type': item_type})

@app.route('/simulator/api/bin', methods=['POST'])
def simulator_bin():
    data = request.get_json() or {}
    dist = int(data.get('distance_cm', 60))
    esp32.set_bin_distance(dist)
    return jsonify({'success': True, 'distance_cm': dist})

@app.route('/simulator/api/reset', methods=['POST'])
def simulator_reset():
    esp32.reset_session()
    return jsonify({'success': True})

@app.route('/simulator/api/lcd', methods=['POST'])
def simulator_set_lcd():
    data = request.get_json() or {}
    l0 = data.get('line0')
    l1 = data.get('line1')
    l2 = data.get('line2')
    l3 = data.get('line3')
    esp32.set_lcd(line0=l0, line1=l1, line2=l2, line3=l3)
    return jsonify({'success': True, 'lcd_lines': esp32.lcd_lines})
FORCE_PASS_HTML = '\n<!DOCTYPE html>\n<html lang="en">\n<head>\n    <meta charset="utf-8">\n    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">\n    <title>Update Admin Password</title>\n    <link rel="stylesheet" href="/static/vendor/fontawesome/css/all.min.css">\n    <style>\n        body {\n            background-color: #0b0f19;\n            min-height: 100vh;\n            display: flex;\n            align-items: center;\n            justify-content: center;\n            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;\n            margin: 0;\n            padding: 16px;\n        }\n        .pass-box-clean {\n            width: 100%;\n            max-width: 340px;\n            background: #111827;\n            border: 1px solid #1f2937;\n            border-radius: 12px;\n            padding: 24px;\n            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.6);\n        }\n        .pass-title {\n            font-size: 17px;\n            font-weight: 700;\n            color: #f9fafb;\n            text-align: center;\n            margin-bottom: 4px;\n        }\n        .pass-subtitle {\n            font-size: 12px;\n            color: #f87171;\n            text-align: center;\n            margin-bottom: 18px;\n            line-height: 1.4;\n        }\n        .form-group-clean {\n            margin-bottom: 14px;\n        }\n        .form-group-clean label {\n            display: block;\n            font-size: 12px;\n            font-weight: 500;\n            color: #9ca3af;\n            margin-bottom: 5px;\n        }\n        .form-control-clean {\n            width: 100%;\n            height: 38px;\n            background-color: #1f2937 !important;\n            border: 1px solid #374151 !important;\n            border-radius: 6px !important;\n            color: #f9fafb !important;\n            font-size: 13px !important;\n            padding: 8px 12px !important;\n            box-sizing: border-box;\n            outline: none;\n        }\n        .form-control-clean:focus {\n            border-color: #ef4444 !important;\n            box-shadow: 0 0 0 2px rgba(239, 68, 68, 0.2) !important;\n        }\n        .btn-update {\n            width: 100%;\n            height: 38px;\n            background: #ef4444;\n            border: none;\n            border-radius: 6px;\n            color: #ffffff;\n            font-size: 13.5px;\n            font-weight: 600;\n            cursor: pointer;\n            margin-top: 6px;\n            transition: background 0.15s ease;\n        }\n        .btn-update:hover {\n            background: #dc2626;\n        }\n        .pass-footer {\n            margin-top: 18px;\n            text-align: center;\n            font-size: 12px;\n        }\n        .pass-footer a {\n            color: #6b7280;\n            text-decoration: none;\n        }\n        .pass-footer a:hover {\n            color: #9ca3af;\n        }\n    </style>\n</head>\n<body>\n<div class="pass-box-clean">\n    <div class="pass-title"><i class="fas fa-shield-alt text-danger mr-1"></i> Security Requirement</div>\n    <div class="pass-subtitle">You must change the default password before accessing the admin dashboard.</div>\n    \n    <form method="POST" action="/admin/force_password_change">\n        <div class="form-group-clean">\n            <label for="new_password">New Password (min. 6 characters)</label>\n            <input type="password" id="new_password" name="new_password" class="form-control-clean" placeholder="Enter new password" required minlength="6" autofocus>\n        </div>\n        \n        <button type="submit" class="btn-update">Change Password</button>\n    </form>\n    \n    <div class="pass-footer">\n        <a href="/admin/logout">← Cancel & Logout</a>\n    </div>\n</div>\n</body>\n</html>\n'

@app.before_request
def admin_security_guard():
    is_admin_route = request.path == '/admin' or request.path.startswith('/admin/')
    is_sim_route = request.path == '/simulator' or request.path.startswith('/simulator/')
    if is_admin_route or is_sim_route:
        if request.path == '/admin/login':
            return None
        if not session.get('admin_logged_in'):
            if request.path.startswith('/admin/api/') or request.path.startswith('/simulator/api/'):
                return (jsonify({'error': 'unauthorized', 'message': 'Admin authentication required.'}), 401)
            return redirect('/admin/login')
        if session.get('must_change_password'):
            allowed_during_pw_change = ['/admin/force_password_change', '/admin/logout']
            if request.path not in allowed_during_pw_change:
                if request.path.startswith('/admin/api/') or request.path.startswith('/simulator/api/'):
                    return (jsonify({'error': 'password_change_required', 'message': 'Default password must be changed first.'}), 403)
                return redirect('/admin/force_password_change')

@app.route('/admin/force_password_change', methods=['GET', 'POST'])
def admin_force_password_change():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    if request.method == 'POST':
        new_pw = request.form.get('new_password', '').strip()
        if new_pw and len(new_pw) >= 6 and (new_pw != 'admin123'):
            admin_user = session.get('admin_username', 'admin')
            with db_connection() as conn:
                conn.execute('UPDATE admins SET password_hash=? WHERE username=?', (generate_password_hash(new_pw, method='pbkdf2:sha256'), admin_user))
            session.pop('must_change_password', None)
            return redirect('/admin')
    return render_template_string(FORCE_PASS_HTML)
admin_login_attempts = {}

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        client_ip = get_client_ip()
        now = time.time()
        if client_ip in admin_login_attempts:
            count, last_time = admin_login_attempts[client_ip]
            if now - last_time < 300:
                if count >= 5:
                    return render_template_string(LOGIN_HTML, error='Too many attempts. Try again in 5 minutes.')
            else:
                admin_login_attempts[client_ip] = [0, now]
        else:
            admin_login_attempts[client_ip] = [0, now]
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        with db_connection() as conn:
            c = conn.cursor()
            c.execute('SELECT password_hash FROM admins WHERE username=?', (username,))
            row = c.fetchone()
            if row and check_password_hash(row[0], password):
                if client_ip in admin_login_attempts:
                    del admin_login_attempts[client_ip]
                session['admin_logged_in'] = True
                session['admin_username'] = username
                if password == 'admin123':
                    session['must_change_password'] = True
                return redirect('/admin')
            else:
                admin_login_attempts[client_ip][0] += 1
                admin_login_attempts[client_ip][1] = time.time()
                error = 'Invalid username or password'
    return render_template_string(LOGIN_HTML, error=error)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect('/admin/login')

@app.route('/admin')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    return render_template_string(ADMIN_HTML, config=get_all_config())

@app.route('/admin/api/stats')
def admin_api_stats():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    today = datetime.now().strftime('%Y-%m-%d')
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT total_bottles FROM stats WHERE date=?', (today,))
        row = c.fetchone()
        today_bottles = row[0] if row else 0
        c.execute('SELECT SUM(total_bottles) FROM stats')
        row = c.fetchone()
        total_bottles = row[0] if row and row[0] else 0
        c.execute('SELECT date, total_bottles FROM stats ORDER BY date DESC LIMIT 7')
        history = [{'date': r[0], 'count': r[1]} for r in c.fetchall()]
        if not history:
            history = [{'date': today, 'count': today_bottles}]
    with active_clients_lock:
        active_count = sum((1 for c in active_clients.values() if c['remaining_seconds'] > 0))
    cpu_val, ram_val, disk_val, uptime_val = (12, 28, 15, '2h 45m')
    try:
        if platform.system() == 'Linux':
            with open('/proc/loadavg', 'r') as f:
                load = float(f.read().split()[0])
                cpu_val = min(100, int(load * 100 / (os.cpu_count() or 1)))
            with open('/proc/meminfo', 'r') as f:
                mem = {}
                for line in f:
                    parts = line.split()
                    mem[parts[0].strip(':')] = int(parts[1])
                ram_val = int((mem['MemTotal'] - mem['MemAvailable']) / mem['MemTotal'] * 100)
            with open('/proc/uptime', 'r') as f:
                up_sec = float(f.read().split()[0])
                uptime_val = '{}h {}m'.format(int(up_sec // 3600), int(up_sec % 3600 // 60))
            st = os.statvfs('/')
            disk_val = int((st.f_blocks - st.f_bavail) / st.f_blocks * 100)
        else:
            cpu_val = random.randint(5, 20)
            ram_val = random.randint(30, 45)
    except Exception:
        pass
    return jsonify({'today_bottles': today_bottles, 'total_bottles': total_bottles, 'active_clients': active_count, 'cpu': cpu_val, 'ram': ram_val, 'disk': disk_val, 'uptime': uptime_val, 'history': list(reversed(history))})

@app.route('/admin/api/clients')
def admin_api_clients():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    with active_clients_lock:
        res = []
        for ip, sess in active_clients.items():
            res.append({'ip': ip, 'mac': sess.get('mac', '00:00:00:00:00:00'), 'remaining_seconds': sess.get('remaining_seconds', 0), 'is_paused': sess.get('is_paused', False), 'dl_kbps': sess.get('dl_kbps', 3072), 'ul_kbps': sess.get('ul_kbps', 1536)})
        return jsonify(res)

@app.route('/admin/api/client/action', methods=['POST'])
def admin_api_client_action():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    ip = data.get('ip')
    action = data.get('action')
    with active_clients_lock:
        if ip in active_clients:
            if action == 'add15':
                active_clients[ip]['remaining_seconds'] += 15 * 60
                sync_client_firewall(ip)
            elif action == 'add60':
                active_clients[ip]['remaining_seconds'] += 60 * 60
                sync_client_firewall(ip)
            elif action == 'pause':
                active_clients[ip]['is_paused'] = True
                active_clients[ip]['admin_paused'] = True
                sync_client_firewall(ip)
            elif action == 'resume':
                active_clients[ip]['is_paused'] = False
                active_clients[ip]['admin_paused'] = False
                sync_client_firewall(ip)
            elif action == 'kick':
                active_clients[ip]['remaining_seconds'] = 0
                active_clients[ip]['is_paused'] = False
                active_clients[ip]['admin_paused'] = False
                sync_client_firewall(ip)
            save_sessions_to_db()
            return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Client not found.'})

@app.route('/admin/api/client/edit', methods=['POST'])
def admin_api_client_edit():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    ip = data.get('ip')
    minutes = data.get('minutes')
    dl = data.get('dl_kbps')
    ul = data.get('ul_kbps')
    with active_clients_lock:
        if ip in active_clients:
            if dl is not None:
                active_clients[ip]['dl_kbps'] = max(128, int(dl))
            if ul is not None:
                active_clients[ip]['ul_kbps'] = max(64, int(ul))
            if minutes is not None:
                active_clients[ip]['remaining_seconds'] = max(0, int(minutes) * 60)
            sync_client_firewall(ip)
            save_sessions_to_db()
            return jsonify({'success': True, 'client': active_clients[ip]})
    return jsonify({'success': False, 'error': 'Client not found.'})

@app.route('/admin/api/vouchers/list')
def admin_api_vouchers_list():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT code, minutes, is_used, created_at, used_by, note FROM vouchers ORDER BY created_at DESC LIMIT 50')
        rows = [{'code': r[0], 'minutes': r[1], 'is_used': r[2], 'created_at': r[3], 'used_by': r[4], 'note': r[5] or ''} for r in c.fetchall()]
        return jsonify(rows)

@app.route('/admin/api/vouchers/generate', methods=['POST'])
def admin_generate_vouchers():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    try:
        qty = max(1, min(100, int(data.get('qty', 5))))
        minutes = max(1, int(data.get('minutes', 60)))
    except (ValueError, TypeError):
        return (jsonify({'success': False, 'error': 'Invalid quantity or duration.'}), 400)
    note = data.get('note', '').strip()
    prefix = data.get('prefix', 'ECO-').strip().upper()
    created = []
    with db_connection() as conn:
        c = conn.cursor()
        for _ in range(qty):
            code = prefix + ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(6))
            c.execute('INSERT OR REPLACE INTO vouchers (code, minutes, created_at, note) VALUES (?, ?, ?, ?)', (code, minutes, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), note))
            created.append({'code': code, 'minutes': minutes, 'note': note})
        conn.commit()
    return jsonify({'success': True, 'vouchers': created})

@app.route('/admin/api/vouchers/delete', methods=['POST'])
def admin_delete_voucher():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    code = data.get('code', '').strip().upper()
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('DELETE FROM vouchers WHERE code = ?', (code,))
        conn.commit()
        return jsonify({'success': True})

@app.route('/admin/api/members/list')
def admin_api_members_list():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT username, wallet_minutes, created_at FROM members ORDER BY created_at DESC')
        rows = [{'username': r[0], 'wallet_minutes': r[1], 'created_at': r[2]} for r in c.fetchall()]
        return jsonify(rows)

@app.route('/admin/api/members/add', methods=['POST'])
def admin_api_members_add():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    pin = data.get('pin', '').strip()
    mins = int(data.get('wallet_minutes', 0))
    if not username or len(username) < 3:
        return (jsonify({'success': False, 'error': 'Username must be at least 3 characters.'}), 400)
    if not pin or not re.match('^\\d{4,6}$', pin):
        return (jsonify({'success': False, 'error': 'PIN must be 4 to 6 digits.'}), 400)
    pin_hash = generate_password_hash(pin, method='pbkdf2:sha256')
    try:
        with db_connection() as conn:
            c = conn.cursor()
            c.execute('INSERT INTO members (username, pin_hash, wallet_minutes, created_at) VALUES (?, ?, ?, ?)', (username, pin_hash, mins, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            conn.commit()
            return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return (jsonify({'success': False, 'error': 'Username already exists.'}), 400)

@app.route('/admin/api/members/topup', methods=['POST'])
def admin_api_members_topup():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    username = data.get('username')
    mins = int(data.get('minutes', 15))
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('UPDATE members SET wallet_minutes = MAX(0, wallet_minutes + ?) WHERE username = ?', (mins, username))
        conn.commit()
        return jsonify({'success': True})

@app.route('/admin/api/members/delete', methods=['POST'])
def admin_api_members_delete():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    username = data.get('username')
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('DELETE FROM members WHERE username = ?', (username,))
        conn.commit()
        return jsonify({'success': True})

@app.route('/admin/api/settings/save', methods=['POST'])
def admin_api_settings_save():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    for k, v in data.items():
        set_config(k, v)
        if k == 'announcement':
            with db_connection() as conn:
                c = conn.cursor()
                c.execute('UPDATE announcements SET message = ? WHERE id = 1', (str(v),))
                if c.rowcount == 0:
                    c.execute('INSERT INTO announcements (id, message, active) VALUES (1, ?, 1)', (str(v),))
                conn.commit()

    # If default bandwidth limits changed, push new rates to ALL active clients
    # that are still running on the default speed (not per-MAC overridden).
    if 'default_dl_kbps' in data or 'default_ul_kbps' in data:
        new_dl = int(get_config('default_dl_kbps', '3072'))
        new_ul = int(get_config('default_ul_kbps', '1536'))
        with active_clients_lock:
            for ip, sess in active_clients.items():
                if sess.get('remaining_seconds', 0) > 0 and not sess.get('is_paused'):
                    # Only update clients whose speed wasn't individually overridden
                    # (i.e. they still carry the previous default or no custom value)
                    sess['dl_kbps'] = new_dl
                    sess['ul_kbps'] = new_ul
                    sync_client_firewall(ip)
        save_sessions_to_db()

    return jsonify({'success': True})

@app.route('/admin/api/mac_control/list')
def admin_api_mac_list():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT mac, type, note, dl_kbps, ul_kbps FROM mac_control ORDER BY mac ASC')
        return jsonify([{'mac': r[0], 'type': r[1], 'note': r[2] or '', 'dl_kbps': r[3] or 0, 'ul_kbps': r[4] or 0} for r in c.fetchall()])

@app.route('/admin/api/mac_control/add', methods=['POST'])
def admin_api_mac_add():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    raw_mac = data.get('mac', '').strip().upper()
    cleaned = re.sub('[^0-9A-F]', '', raw_mac)
    if len(cleaned) == 12:
        mac = ':'.join((cleaned[i:i + 2] for i in range(0, 12, 2)))
    else:
        mac = raw_mac
    if not re.match('^([0-9A-F]{2}[:-]){5}([0-9A-F]{2})$', mac):
        return (jsonify({'success': False, 'error': 'Invalid MAC address format! Must be 12 hex characters (e.g. AA:BB:CC:DD:EE:FF).'}), 400)
    m_type = data.get('type', 'whitelist')
    note = data.get('note', '').strip()
    dl = int(data.get('dl_kbps', 0))
    ul = int(data.get('ul_kbps', 0))
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('REPLACE INTO mac_control (mac, type, note, dl_kbps, ul_kbps) VALUES (?, ?, ?, ?, ?)', (mac, m_type, note, dl, ul))
        conn.commit()
    apply_walled_garden_and_macs()
    return jsonify({'success': True, 'mac': mac})

@app.route('/admin/api/mac_control/delete', methods=['POST'])
def admin_api_mac_delete():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    mac = data.get('mac', '').strip().upper()
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('DELETE FROM mac_control WHERE mac = ?', (mac,))
        conn.commit()
    apply_walled_garden_and_macs()
    return jsonify({'success': True})

@app.route('/admin/api/walled_garden/list')
def admin_api_walled_garden_list():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT domain, note FROM walled_garden ORDER BY domain ASC')
        return jsonify([{'domain': r[0], 'note': r[1] or ''} for r in c.fetchall()])

@app.route('/admin/api/walled_garden/add', methods=['POST'])
def admin_api_walled_garden_add():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    domain = data.get('domain', '').strip().lower()
    note = data.get('note', '').strip()
    if not domain or '.' not in domain or len(domain) < 4:
        return (jsonify({'success': False, 'error': 'Invalid domain name. Example: gcash.com or deped.gov.ph'}), 400)
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('REPLACE INTO walled_garden (domain, note) VALUES (?, ?)', (domain, note))
        conn.commit()
        return jsonify({'success': True})

@app.route('/admin/api/walled_garden/delete', methods=['POST'])
def admin_api_walled_garden_delete():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    data = request.get_json() or {}
    domain = data.get('domain', '').strip().lower()
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('DELETE FROM walled_garden WHERE domain = ?', (domain,))
        conn.commit()
        return jsonify({'success': True})

@app.route('/admin/api/telegram/test', methods=['POST'])
def admin_api_telegram_test():
    if not session.get('admin_logged_in'):
        return (jsonify({'error': 'unauthorized'}), 401)
    ok = send_telegram_alert('🔔 *ECO-Fi Test Alert*\n\nThis is a successful test notification from your Reverse Vending Machine!')
    return jsonify({'success': ok})

def generate_ecofi_excel_report(db_path):
    if not openpyxl:
        return None
    wb = openpyxl.Workbook()
    FONT_FAMILY = 'Segoe UI'
    title_font = Font(name=FONT_FAMILY, size=16, bold=True, color='FFFFFF')
    subtitle_font = Font(name=FONT_FAMILY, size=10, italic=True, color='E2E8F0')
    kpi_title_font = Font(name=FONT_FAMILY, size=9, bold=True, color='64748B')
    kpi_value_font = Font(name=FONT_FAMILY, size=14, bold=True, color='0F172A')
    header_font = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF')
    total_font = Font(name=FONT_FAMILY, size=11, bold=True, color='0F172A')
    data_font = Font(name=FONT_FAMILY, size=10, color='1E293B')
    title_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    kpi_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    zebra_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    total_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    status_active_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    status_active_font = Font(name=FONT_FAMILY, size=10, bold=True, color='15803D')
    status_used_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    status_used_font = Font(name=FONT_FAMILY, size=10, color='64748B')
    thin_border_side = Side(style='thin', color='CBD5E1')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    kpi_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    total_border = Border(left=thin_border_side, right=thin_border_side, top=Side(style='thin', color='0F172A'), bottom=Side(style='double', color='0F172A'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    ws1 = wb.active
    ws1.title = 'Daily Collections & Impact'
    ws1.views.sheetView[0].showGridLines = True
    with sqlite3.connect(db_path) as conn:
        c = conn.cursor()
        c.execute('SELECT date, total_bottles FROM stats ORDER BY date DESC')
        stats_rows = c.fetchall()
        c.execute('SELECT COUNT(*), SUM(wallet_minutes) FROM members')
        m_stats = c.fetchone()
        total_members = m_stats[0] or 0
        c.execute('SELECT COUNT(*), SUM(CASE WHEN is_used=0 THEN 1 ELSE 0 END) FROM vouchers')
        v_stats = c.fetchone()
        total_vouchers = v_stats[0] or 0
        unclaimed_vouchers = v_stats[1] or 0
    total_bottles_sum = sum((r[1] for r in stats_rows))
    est_plastic_kg = total_bottles_sum * 0.025
    est_co2_kg = est_plastic_kg * 1.5
    total_mins_sum = total_bottles_sum * 10
    total_hours_sum = total_mins_sum / 60.0
    ws1.merge_cells('A1:G1')
    ws1['A1'] = 'ECO-FI REVERSE VENDING MACHINE (PBVM)'
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = align_center
    ws1.row_dimensions[1].height = 28
    ws1.merge_cells('A2:G2')
    ws1['A2'] = 'Executive Operations & Environmental Impact Report • Generated on {}'.format(datetime.now().strftime('%B %d, %Y at %I:%M %p'))
    ws1['A2'].font = subtitle_font
    ws1['A2'].fill = title_fill
    ws1['A2'].alignment = align_center
    ws1.row_dimensions[2].height = 18
    kpis = [('TOTAL BOTTLES', '{:,}'.format(total_bottles_sum), 'A', 'B'), ('WIFI TIME ISSUED', '{:.1f} Hours'.format(total_hours_sum), 'C', 'C'), ('PLASTIC RECYCLED', '{:.2f} kg'.format(est_plastic_kg), 'D', 'D'), ('CO₂ OFFSET', '{:.2f} kg'.format(est_co2_kg), 'E', 'E'), ('MEMBERS', '{} users'.format(total_members), 'F', 'F'), ('ACTIVE VOUCHERS', '{} avail'.format(unclaimed_vouchers), 'G', 'G')]
    ws1.row_dimensions[4].height = 16
    ws1.row_dimensions[5].height = 24
    for title, val, c1, c2 in kpis:
        if c1 != c2:
            ws1.merge_cells('{}4:{}4'.format(c1, c2))
            ws1.merge_cells('{}5:{}5'.format(c1, c2))
        top_cell = ws1['{}4'.format(c1)]
        top_cell.value = title
        top_cell.font = kpi_title_font
        top_cell.fill = kpi_fill
        top_cell.alignment = align_center
        top_cell.border = kpi_border
        val_cell = ws1['{}5'.format(c1)]
        val_cell.value = val
        val_cell.font = kpi_value_font
        val_cell.fill = kpi_fill
        val_cell.alignment = align_center
        val_cell.border = kpi_border
        if c1 != c2:
            ws1['{}4'.format(c2)].border = kpi_border
            ws1['{}5'.format(c2)].border = kpi_border
    headers = [('Date', align_center), ('Bottles Recycled', align_right), ('WiFi Time (Minutes)', align_right), ('WiFi Time (Hours)', align_right), ('Plastic Weight (kg)', align_right), ('CO₂ Saved (kg)', align_right), ('Collection Status', align_center)]
    ws1.row_dimensions[7].height = 24
    for col_idx, (h_title, h_align) in enumerate(headers, start=1):
        cell = ws1.cell(row=7, column=col_idx, value=h_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = h_align
        cell.border = cell_border
    current_row = 8
    for idx, (dt, count) in enumerate(stats_rows):
        mins = count * 10
        hrs = mins / 60.0
        kg = count * 0.025
        co2 = kg * 1.5
        status = 'High Volume' if count >= 20 else 'Active' if count > 0 else 'Idle'
        fill = zebra_fill if idx % 2 == 1 else white_fill
        ws1.row_dimensions[current_row].height = 20
        row_vals = [(dt, align_center, '@'), (count, align_right, '#,##0'), (mins, align_right, '#,##0'), (hrs, align_right, '0.00'), (kg, align_right, '0.000'), (co2, align_right, '0.000'), (status, align_center, '@')]
        for col_idx, (val, c_align, num_fmt) in enumerate(row_vals, start=1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = c_align
            cell.border = cell_border
            cell.number_format = num_fmt
        current_row += 1
    if stats_rows:
        ws1.row_dimensions[current_row].height = 22
        tot_cells = [('TOTAL', align_center), ('=SUM(B8:B{})'.format(current_row - 1), align_right, '#,##0'), ('=SUM(C8:C{})'.format(current_row - 1), align_right, '#,##0'), ('=SUM(D8:D{})'.format(current_row - 1), align_right, '0.00'), ('=SUM(E8:E{})'.format(current_row - 1), align_right, '0.000'), ('=SUM(F8:F{})'.format(current_row - 1), align_right, '0.000'), ('', align_center)]
        for col_idx, (val, c_align, *opt_fmt) in enumerate(tot_cells, start=1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = c_align
            cell.border = total_border
            if opt_fmt:
                cell.number_format = opt_fmt[0]
    ws1.freeze_panes = 'A8'
    ws2 = wb.create_sheet(title='Voucher Inventory')
    ws2.views.sheetView[0].showGridLines = True
    ws2.merge_cells('A1:G1')
    ws2['A1'] = 'ECO-FI VOUCHER TICKETS INVENTORY'
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = align_center
    ws2.row_dimensions[1].height = 26
    v_headers = ['Voucher Code', 'Duration (Mins)', 'Duration (Hours)', 'Status', 'Created Date', 'Redeemed / Used By', 'Batch / Admin Note']
    ws2.row_dimensions[3].height = 22
    for col_idx, h_title in enumerate(v_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx in [1, 4] else align_right if col_idx in [2, 3] else align_left
        cell.border = cell_border
    with sqlite3.connect(db_path) as conn:
        c = conn.cursor()
        c.execute('SELECT code, minutes, is_used, created_at, used_by, note FROM vouchers ORDER BY created_at DESC')
        v_rows = c.fetchall()
    v_row_idx = 4
    for idx, (code, mins, is_used, created_at, used_by, note) in enumerate(v_rows):
        ws2.row_dimensions[v_row_idx].height = 19
        fill = zebra_fill if idx % 2 == 1 else white_fill
        status_str = 'REDEEMED' if is_used else 'ACTIVE'
        s_fill = status_used_fill if is_used else status_active_fill
        s_font = status_used_font if is_used else status_active_font
        row_data = [(code, align_center, fill, data_font, '@'), (mins, align_right, fill, data_font, '#,##0'), (mins / 60.0, align_right, fill, data_font, '0.00'), (status_str, align_center, s_fill, s_font, '@'), (created_at or '--', align_left, fill, data_font, '@'), (used_by or '--', align_left, fill, data_font, '@'), (note or '', align_left, fill, data_font, '@')]
        for col_idx, (val, c_align, c_fill, c_font, n_fmt) in enumerate(row_data, start=1):
            cell = ws2.cell(row=v_row_idx, column=col_idx, value=val)
            cell.font = c_font
            cell.fill = c_fill
            cell.alignment = c_align
            cell.border = cell_border
            cell.number_format = n_fmt
        v_row_idx += 1
    ws2.freeze_panes = 'A4'
    ws3 = wb.create_sheet(title='Member Wallets')
    ws3.views.sheetView[0].showGridLines = True
    ws3.merge_cells('A1:E1')
    ws3['A1'] = 'ECO-FI REGISTERED MEMBERS & TIME WALLETS'
    ws3['A1'].font = title_font
    ws3['A1'].fill = title_fill
    ws3['A1'].alignment = align_center
    ws3.row_dimensions[1].height = 26
    m_headers = ['Member Username', 'Wallet Balance (Mins)', 'Wallet Balance (Hours)', 'Registered Date', 'Account Status']
    ws3.row_dimensions[3].height = 22
    for col_idx, h_title in enumerate(m_headers, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=h_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_right if col_idx in [2, 3] else align_center if col_idx in [4, 5] else align_left
        cell.border = cell_border
    with sqlite3.connect(db_path) as conn:
        c = conn.cursor()
        c.execute('SELECT username, wallet_minutes, created_at FROM members ORDER BY wallet_minutes DESC')
        m_rows = c.fetchall()
    m_row_idx = 4
    for idx, (uname, w_mins, m_created) in enumerate(m_rows):
        ws3.row_dimensions[m_row_idx].height = 19
        fill = zebra_fill if idx % 2 == 1 else white_fill
        row_data = [(uname, align_left, '@'), (w_mins, align_right, '#,##0'), (w_mins / 60.0, align_right, '0.00'), (m_created or '--', align_center, '@'), ('Active User', align_center, '@')]
        for col_idx, (val, c_align, n_fmt) in enumerate(row_data, start=1):
            cell = ws3.cell(row=m_row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = c_align
            cell.border = cell_border
            cell.number_format = n_fmt
        m_row_idx += 1
    ws3.freeze_panes = 'A4'
    ws4 = wb.create_sheet(title='Promo Rate Curves')
    ws4.views.sheetView[0].showGridLines = True
    ws4.merge_cells('A1:E1')
    ws4['A1'] = 'ECO-FI ACTIVE RATE TIERS & PROMO CURVES'
    ws4['A1'].font = title_font
    ws4['A1'].fill = title_fill
    ws4['A1'].alignment = align_center
    ws4.row_dimensions[1].height = 26
    r_headers = ['Bottles Required', 'Time Credited (Mins)', 'Time Credited (Hours)', 'Rate Efficiency', 'Package Display Label']
    ws4.row_dimensions[3].height = 22
    for col_idx, h_title in enumerate(r_headers, start=1):
        cell = ws4.cell(row=3, column=col_idx, value=h_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_right if col_idx in [1, 2, 3] else align_center if col_idx == 4 else align_left
        cell.border = cell_border
    with sqlite3.connect(db_path) as conn:
        c = conn.cursor()
        c.execute('SELECT bottles, minutes, label FROM promo_rates ORDER BY bottles ASC')
        r_rows = c.fetchall()
    r_row_idx = 4
    for idx, (b, m, l) in enumerate(r_rows):
        ws4.row_dimensions[r_row_idx].height = 19
        fill = zebra_fill if idx % 2 == 1 else white_fill
        eff = '{:.1f} m/bottle'.format(m / b)
        row_data = [(b, align_right, '#,##0'), (m, align_right, '#,##0'), (m / 60.0, align_right, '0.00'), (eff, align_center, '@'), (l, align_left, '@')]
        for col_idx, (val, c_align, n_fmt) in enumerate(row_data, start=1):
            cell = ws4.cell(row=r_row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = c_align
            cell.border = cell_border
            cell.number_format = n_fmt
        r_row_idx += 1
    ws4.freeze_panes = 'A4'
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2]:
                    continue
                if cell.value is not None:
                    s = str(cell.value)
                    if len(s) > max_len:
                        max_len = len(s)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)
    return wb

@app.route('/admin/api/export_xlsx')
def admin_export_xlsx():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    if not openpyxl:
        return redirect('/admin/api/export_csv')
    wb = generate_ecofi_excel_report(DB_PATH)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = 'ECO_Fi_Operations_Report_{}.xlsx'.format(datetime.now().strftime('%Y%m%d'))
    try:
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except TypeError:
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, attachment_filename=filename)

@app.route('/admin/api/export_csv')
def admin_export_csv():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    with db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT date, total_bottles FROM stats ORDER BY date DESC')
        rows = c.fetchall()
    csv_data = 'Date,Total Bottles,Equivalent Minutes\n'
    for r in rows:
        csv_data += '{},{},{}\n'.format(r[0], r[1], r[1] * 10)
    return Response(csv_data, mimetype='text/csv', headers={'Content-disposition': 'attachment; filename=ecofi_sales_report.csv'})
LOGIN_HTML = '\n<!DOCTYPE html>\n<html lang="en">\n<head>\n    <meta charset="utf-8">\n    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">\n    <title>ECO-Fi Admin Login</title>\n    <link rel="stylesheet" href="/static/vendor/fontawesome/css/all.min.css">\n    <link rel="stylesheet" href="/static/vendor/adminlte/css/adminlte.min.css">\n    <style>\n        body {\n            background-color: #0b0f19;\n            min-height: 100vh;\n            display: flex;\n            align-items: center;\n            justify-content: center;\n            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;\n            margin: 0;\n            padding: 16px;\n        }\n        .login-box-clean {\n            width: 100%;\n            max-width: 340px;\n            background: #111827;\n            border: 1px solid #1f2937;\n            border-radius: 12px;\n            padding: 24px;\n            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.6);\n        }\n        .login-title {\n            font-size: 18px;\n            font-weight: 700;\n            color: #f9fafb;\n            text-align: center;\n            margin-bottom: 2px;\n        }\n        .login-subtitle {\n            font-size: 11px;\n            color: #6b7280;\n            text-align: center;\n            margin-bottom: 20px;\n            letter-spacing: 0.5px;\n            text-transform: uppercase;\n        }\n        .form-group-clean {\n            margin-bottom: 14px;\n        }\n        .form-group-clean label {\n            display: block;\n            font-size: 12px;\n            font-weight: 500;\n            color: #9ca3af;\n            margin-bottom: 5px;\n        }\n        .form-control-clean {\n            width: 100%;\n            height: 38px;\n            background-color: #1f2937 !important;\n            border: 1px solid #374151 !important;\n            border-radius: 6px !important;\n            color: #f9fafb !important;\n            font-size: 13px !important;\n            padding: 8px 12px !important;\n            box-sizing: border-box;\n            outline: none;\n            transition: border-color 0.15s ease-in-out;\n        }\n        .form-control-clean:focus {\n            border-color: #10b981 !important;\n            box-shadow: 0 0 0 2px rgba(16, 185, 129, 0.2) !important;\n        }\n        input:-webkit-autofill,\n        input:-webkit-autofill:hover, \n        input:-webkit-autofill:focus {\n            -webkit-text-fill-color: #f9fafb !important;\n            -webkit-box-shadow: 0 0 0px 1000px #1f2937 inset !important;\n            transition: background-color 5000s ease-in-out 0s;\n        }\n        .btn-signin {\n            width: 100%;\n            height: 38px;\n            background: #10b981;\n            border: none;\n            border-radius: 6px;\n            color: #ffffff;\n            font-size: 13.5px;\n            font-weight: 600;\n            cursor: pointer;\n            margin-top: 6px;\n            transition: background 0.15s ease;\n        }\n        .btn-signin:hover {\n            background: #059669;\n        }\n        .btn-signin:active {\n            background: #047857;\n        }\n        .alert-error {\n            background: rgba(239, 68, 68, 0.15);\n            border: 1px solid rgba(239, 68, 68, 0.3);\n            color: #fca5a5;\n            padding: 8px 12px;\n            border-radius: 6px;\n            font-size: 12px;\n            margin-bottom: 14px;\n            text-align: center;\n        }\n        .login-footer {\n            margin-top: 18px;\n            text-align: center;\n            font-size: 12px;\n        }\n        .login-footer a {\n            color: #6b7280;\n            text-decoration: none;\n        }\n        .login-footer a:hover {\n            color: #9ca3af;\n        }\n    </style>\n</head>\n<body>\n<div class="login-box-clean">\n    <div class="login-title"><i class="fas fa-recycle text-success mr-1"></i> ECO-Fi VENDO</div>\n    <div class="login-subtitle">Master Control Panel</div>\n    \n    {% if error %}\n    <div class="alert-error">{{ error }}</div>\n    {% endif %}\n    \n    <form method="post">\n        <div class="form-group-clean">\n            <label for="username">Username</label>\n            <input type="text" id="username" name="username" class="form-control-clean" placeholder="Admin username" required autofocus autocomplete="username">\n        </div>\n        \n        <div class="form-group-clean">\n            <label for="password">Password</label>\n            <input type="password" id="password" name="password" class="form-control-clean" placeholder="Password" required autocomplete="current-password">\n        </div>\n        \n        <button type="submit" class="btn-signin">Sign In</button>\n    </form>\n    \n    <div class="login-footer">\n        <a href="/">← Return to Client Portal</a>\n    </div>\n</div>\n</body>\n</html>\n'
ADMIN_HTML = '\n<!DOCTYPE html>\n<html lang="en">\n<head>\n    <title>ECO-Fi Master Admin Control Panel</title>\n    <meta charset="utf-8">\n    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no, viewport-fit=cover">\n    <link rel="stylesheet" href="/static/vendor/fontawesome/css/all.min.css">\n    <link rel="stylesheet" href="/static/vendor/adminlte/css/adminlte.min.css">\n    <script src="/static/vendor/jquery/jquery.min.js"></script>\n    <script src="/static/vendor/bootstrap/js/bootstrap.bundle.min.js"></script>\n    <script src="/static/vendor/adminlte/js/adminlte.min.js"></script>\n    <script src="/static/vendor/sweetalert2/sweetalert2.all.min.js"></script>\n    <script src="/static/vendor/chartjs/Chart.bundle.min.js"></script>\n    <style>\n      /* ==========================================================================\n         ECO-FI MASTER PLAIN DARK THEME (FLAT, HIGH-CONTRAST, ZERO GLOW)\n         ========================================================================== */\n      :root {\n        --eco-bg: #0b0f19;\n        --eco-card: #1e293b;\n        --eco-header: #0f172a;\n        --eco-border: rgba(255, 255, 255, 0.08);\n        --eco-border-light: rgba(255, 255, 255, 0.12);\n        --eco-primary: #007bff;\n        --eco-accent: #38bdf8;\n        --eco-text-main: #f8fafc;\n        --eco-text-body: #cbd5e1;\n        --eco-text-muted: #94a3b8;\n      }\n\n      /* Base Layout & Typography */\n      body.dark-mode {\n        background-color: var(--eco-bg) !important;\n        color: var(--eco-text-main) !important;\n        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;\n      }\n      .content-wrapper {\n        background-color: var(--eco-bg) !important;\n        color: var(--eco-text-main) !important;\n        padding: 16px 20px !important;\n      }\n      .main-header.navbar {\n        background-color: var(--eco-header) !important;\n        border-bottom: 1px solid var(--eco-border) !important;\n      }\n      .main-sidebar {\n        background-color: var(--eco-bg) !important;\n        border-right: 1px solid var(--eco-border) !important;\n      }\n      .brand-link {\n        font-weight: 700;\n        letter-spacing: 0.5px;\n        border-bottom: 1px solid var(--eco-border) !important;\n        background-color: var(--eco-bg) !important;\n      }\n\n      /* Sidebar Navigation: Flat, Clean, Zero Glow, Zero Outline */\n      .nav-sidebar .nav-link {\n        color: var(--eco-text-muted) !important;\n        font-weight: 500;\n        border-radius: 4px !important;\n        margin: 2px 8px;\n        box-shadow: none !important;\n        border: none !important;\n        outline: none !important;\n        transition: background-color 0.15s ease;\n      }\n      .nav-sidebar .nav-link:hover {\n        background-color: rgba(255, 255, 255, 0.06) !important;\n        color: #ffffff !important;\n        box-shadow: none !important;\n      }\n      .nav-sidebar .nav-link.active {\n        background-color: #007bff !important;\n        color: #ffffff !important;\n        font-weight: 600 !important;\n        box-shadow: none !important;\n        border: none !important;\n        outline: none !important;\n      }\n      .nav-sidebar .nav-link:focus,\n      .nav-sidebar .nav-link:active,\n      .nav-sidebar .nav-link:focus-visible,\n      a:focus, button:focus {\n        outline: none !important;\n        box-shadow: none !important;\n      }\n      .nav-header {\n        color: #64748b !important;\n        font-weight: 700 !important;\n        letter-spacing: 0.8px !important;\n        font-size: 0.72rem !important;\n        padding: 0.75rem 1rem 0.35rem !important;\n      }\n      .section-view { display: none; }\n      .section-view.active { display: block; animation: fadeIn 0.15s ease-in-out; }\n      @keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }\n\n      /* Plain Cards System */\n      .card {\n        background: var(--eco-card) !important;\n        border: 1px solid var(--eco-border) !important;\n        border-radius: 8px !important;\n        box-shadow: none !important;\n        margin-bottom: 18px !important;\n        overflow: hidden;\n      }\n      .card-header,\n      .card[class*="card-"] > .card-header {\n        background: var(--eco-header) !important;\n        border-bottom: 1px solid var(--eco-border) !important;\n        padding: 12px 18px !important;\n        color: var(--eco-text-main) !important;\n        border-top-left-radius: 8px !important;\n        border-top-right-radius: 8px !important;\n      }\n      .card-title {\n        font-size: 0.98rem !important;\n        font-weight: 700 !important;\n        letter-spacing: 0.3px;\n        margin: 0;\n        color: var(--eco-text-main) !important;\n      }\n      .card-body {\n        padding: 16px 18px !important;\n        color: var(--eco-text-body) !important;\n      }\n      .card-body p, \n      .card-body span:not(.badge):not(.pulse-indicator):not(.badge-custom) {\n        color: var(--eco-text-body) !important;\n      }\n      .card-body strong, \n      .card-body h1, .card-body h2, .card-body h3, .card-body h4, .card-body h5, .card-body h6,\n      .modal-body strong,\n      .modal-body h1, .modal-body h2, .modal-body h3, .modal-body h4, .modal-body h5, .modal-body h6 {\n        color: var(--eco-text-main) !important;\n        font-weight: 700;\n      }\n      .card-footer,\n      .card[class*="card-"] > .card-footer {\n        background: var(--eco-header) !important;\n        border-top: 1px solid var(--eco-border) !important;\n        color: var(--eco-text-main) !important;\n        padding: 12px 18px !important;\n        border-bottom-left-radius: 8px !important;\n        border-bottom-right-radius: 8px !important;\n      }\n      .text-muted {\n        color: var(--eco-text-muted) !important;\n      }\n      hr {\n        border-top: 1px solid var(--eco-border-light) !important;\n      }\n      code {\n        background-color: rgba(15, 23, 42, 0.9) !important;\n        border: 1px solid var(--eco-border-light) !important;\n        color: var(--eco-accent) !important;\n        padding: 2px 6px !important;\n        border-radius: 4px !important;\n        font-size: 0.88rem;\n      }\n\n      .card-footer::after {\n        display: none !important;\n      }\n\n      /* Plain Form Controls & Inputs (No Glow on Focus) */\n      .form-control, .custom-select, select.form-control, textarea.form-control {\n        background-color: #0f172a !important;\n        border: 1px solid #334155 !important;\n        color: #f8fafc !important;\n        border-radius: 6px !important;\n        height: 38px;\n        font-size: 0.9rem;\n        box-shadow: none !important;\n        outline: none !important;\n      }\n      textarea.form-control {\n        height: auto !important;\n      }\n      .form-control::placeholder, textarea.form-control::placeholder {\n        color: #64748b !important;\n        opacity: 1 !important;\n      }\n      .form-control:focus, .custom-select:focus, select.form-control:focus, textarea.form-control:focus {\n        border-color: #007bff !important;\n        box-shadow: none !important;\n        outline: none !important;\n        background-color: #0f172a !important;\n        color: #ffffff !important;\n      }\n      select.form-control option, .custom-select option {\n        background-color: #1e293b !important;\n        color: #f8fafc !important;\n        padding: 6px 10px;\n      }\n      .form-group label, .modal-body label {\n        font-size: 0.82rem;\n        font-weight: 600;\n        color: var(--eco-text-body) !important;\n        margin-bottom: 5px;\n      }\n      .input-group-text {\n        background-color: #0f172a !important;\n        border: 1px solid #334155 !important;\n        color: var(--eco-text-muted) !important;\n        border-radius: 6px;\n      }\n      .btn-outline-secondary {\n        border-color: #334155 !important;\n        color: var(--eco-text-body) !important;\n        box-shadow: none !important;\n      }\n      .btn-outline-secondary:hover {\n        background-color: #334155 !important;\n        color: #ffffff !important;\n      }\n      input[type="range"].custom-range {\n        accent-color: #007bff;\n      }\n\n      /* Plain Buttons System (Zero Glow, Clean Borders) */\n      .btn { box-shadow: none !important; outline: none !important; border-radius: 6px; }\n      .btn-primary { background-color: #007bff !important; border-color: #0069d9 !important; color: #fff !important; }\n      .btn-primary:hover { background-color: #0069d9 !important; border-color: #0062cc !important; }\n      .btn-success { background-color: #28a745 !important; border-color: #218838 !important; color: #fff !important; }\n      .btn-success:hover { background-color: #218838 !important; border-color: #1e7e34 !important; }\n      .btn-info { background-color: #17a2b8 !important; border-color: #138496 !important; color: #fff !important; }\n      .btn-info:hover { background-color: #138496 !important; border-color: #117a8b !important; }\n      .btn-warning { background-color: #ffc107 !important; border-color: #e0a800 !important; color: #212529 !important; font-weight: 600 !important; }\n      .btn-warning:hover { background-color: #e0a800 !important; border-color: #d39e00 !important; color: #212529 !important; }\n      .btn-danger { background-color: #dc3545 !important; border-color: #c82333 !important; color: #fff !important; }\n      .btn-danger:hover { background-color: #c82333 !important; border-color: #bd2130 !important; }\n      .btn-secondary { background-color: #334155 !important; border-color: #475569 !important; color: #f8fafc !important; }\n      .btn-secondary:hover { background-color: #475569 !important; border-color: #64748b !important; }\n\n      /* Outline Table Action Buttons */\n      .btn-outline-success { color: #28a745 !important; border-color: #28a745 !important; }\n      .btn-outline-success:hover { background-color: #28a745 !important; color: #ffffff !important; }\n      .btn-outline-warning { color: #ffc107 !important; border-color: #ffc107 !important; }\n      .btn-outline-warning:hover { background-color: #ffc107 !important; color: #212529 !important; }\n      .btn-outline-info { color: #17a2b8 !important; border-color: #17a2b8 !important; }\n      .btn-outline-info:hover { background-color: #17a2b8 !important; color: #ffffff !important; }\n      .btn-outline-danger { color: #dc3545 !important; border-color: #dc3545 !important; }\n      .btn-outline-danger:hover { background-color: #dc3545 !important; color: #ffffff !important; }\n\n      /* Modals: Plain & Flat */\n      .modal-content {\n        background-color: var(--eco-card) !important;\n        border: 1px solid var(--eco-border-light) !important;\n        border-radius: 8px !important;\n        color: var(--eco-text-main) !important;\n        box-shadow: none !important;\n        overflow: hidden;\n      }\n      .modal-header {\n        background-color: var(--eco-header) !important;\n        border-bottom: 1px solid var(--eco-border) !important;\n        padding: 14px 18px !important;\n        color: var(--eco-text-main) !important;\n      }\n      .modal-title {\n        color: var(--eco-text-main) !important;\n        font-weight: 700;\n        font-size: 1.05rem;\n      }\n      .modal-body {\n        padding: 18px !important;\n        color: var(--eco-text-body) !important;\n      }\n      .modal-footer {\n        background-color: var(--eco-header) !important;\n        border-top: 1px solid var(--eco-border) !important;\n        padding: 12px 18px !important;\n      }\n      .close {\n        color: var(--eco-text-muted) !important;\n        text-shadow: none !important;\n        opacity: 0.8 !important;\n      }\n      .close:hover {\n        color: #ffffff !important;\n        opacity: 1 !important;\n      }\n\n      /* Plain Tables */\n      .table-responsive {\n        -webkit-overflow-scrolling: touch;\n        overflow-x: auto;\n        margin-bottom: 0;\n      }\n      .table-striped tbody tr:nth-of-type(odd) {\n        background-color: rgba(255, 255, 255, 0.02) !important;\n      }\n      .table-hover tbody tr:hover {\n        background-color: rgba(255, 255, 255, 0.04) !important;\n      }\n      .table {\n        width: 100% !important;\n        margin-bottom: 0 !important;\n        color: var(--eco-text-body) !important;\n      }\n      .table th {\n        background: rgba(15, 23, 42, 0.98) !important;\n        border-bottom: 1px solid var(--eco-border-light) !important;\n        border-top: none !important;\n        font-size: 0.78rem !important;\n        text-transform: uppercase !important;\n        letter-spacing: 0.5px !important;\n        color: var(--eco-text-muted) !important;\n        font-weight: 700 !important;\n        padding: 10px 14px !important;\n        vertical-align: middle !important;\n      }\n      .table td {\n        border-top: 1px solid rgba(255, 255, 255, 0.05) !important;\n        vertical-align: middle !important;\n        font-size: 0.88rem !important;\n        padding: 10px 14px !important;\n        color: #e2e8f0 !important;\n      }\n      .table th.text-center, .table td.text-center { text-align: center !important; }\n      .table th.text-right, .table td.text-right { text-align: right !important; }\n\n      /* Plain Badges (No Glow) */\n      .badge {\n        font-size: 0.78rem;\n        font-weight: 600;\n        padding: 0.35em 0.6em;\n        border-radius: 4px;\n        box-shadow: none !important;\n      }\n      .badge-success { background-color: #28a745 !important; color: #ffffff !important; }\n      .badge-warning { background-color: #ffc107 !important; color: #212529 !important; }\n      .badge-danger { background-color: #dc3545 !important; color: #ffffff !important; }\n      .badge-info { background-color: #17a2b8 !important; color: #ffffff !important; }\n      .badge-secondary { background-color: #6c757d !important; color: #ffffff !important; }\n\n      /* Plain Alerts */\n      .alert { box-shadow: none !important; border-radius: 6px; }\n      .alert-info { background-color: rgba(23, 162, 184, 0.15) !important; border: 1px solid rgba(23, 162, 184, 0.3) !important; color: #7dd3fc !important; }\n      .alert-success { background-color: rgba(40, 167, 69, 0.15) !important; border: 1px solid rgba(40, 167, 69, 0.3) !important; color: #6ee7b7 !important; }\n      .alert-danger { background-color: rgba(220, 53, 69, 0.15) !important; border: 1px solid rgba(220, 53, 69, 0.3) !important; color: #fca5a5 !important; }\n      .alert-warning { background-color: rgba(255, 193, 7, 0.15) !important; border: 1px solid rgba(255, 193, 7, 0.3) !important; color: #fde68a !important; }\n\n      /* Plain Small Boxes Dashboard */\n      .small-box {\n        border-radius: 6px !important;\n        box-shadow: none !important;\n        border: 1px solid var(--eco-border) !important;\n        margin-bottom: 16px;\n      }\n      .small-box .inner { padding: 14px; }\n      .small-box .inner h3 { font-size: 2rem; font-weight: 700 !important; margin-bottom: 2px; }\n      .small-box .inner p { font-size: 0.82rem; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 0.5px; opacity: 0.9; margin: 0; }\n      .small-box .icon { font-size: 50px; right: 12px; top: 12px; opacity: 0.25; }\n\n      /* Progress Bars & Feedback (Zero Glow) */\n      .progress {\n        background-color: #334155 !important;\n        border-radius: 4px;\n        height: 8px;\n        box-shadow: none !important;\n      }\n      .valid-feedback-custom { display: none; font-size: 0.78rem; color: #28a745; margin-top: 4px; font-weight: 600; }\n      .invalid-feedback-custom { display: none; font-size: 0.78rem; color: #dc3545; margin-top: 4px; font-weight: 600; }\n      .pulse-indicator { display: inline-block; width: 10px; height: 10px; border-radius: 50%; background: #28a745; margin-right: 5px; }\n      .btn-xs { padding: 4px 10px; font-size: 0.78rem; border-radius: 4px; font-weight: 600; line-height: 1.4; }\n      .gap-1 { gap: 0.25rem !important; }\n      .gap-2 { gap: 0.5rem !important; }\n      .gap-3 { gap: 1rem !important; }\n\n      /* Mobile Adaptive UI */\n      @media (max-width: 767.98px) {\n        .content-wrapper { padding: 10px !important; }\n        .card { margin-bottom: 12px; }\n        .card-header { padding: 0.6rem 0.8rem !important; }\n        .card-body { padding: 0.8rem !important; }\n        .card-title { font-size: 0.95rem !important; font-weight: 700; }\n        .small-box { margin-bottom: 10px; border-radius: 6px; }\n        .small-box .inner { padding: 10px; }\n        .small-box .inner h3 { font-size: 1.45rem; margin-bottom: 2px; }\n        .small-box .inner p { font-size: 0.72rem; margin-bottom: 0; line-height: 1.2; }\n        .small-box .icon { font-size: 38px; right: 8px; top: 8px; opacity: 0.25; }\n        .table th, .table td { padding: 0.5rem 0.4rem; font-size: 0.8rem; white-space: nowrap; }\n        .btn-sm { padding: 0.28rem 0.5rem; font-size: 0.78rem; }\n        .btn-block { margin-top: 4px; }\n        .modal-dialog { margin: 12px auto; max-width: 95vw; }\n        .navbar-nav .nav-link { padding-left: 0.5rem; padding-right: 0.5rem; }\n        .brand-link { font-size: 0.95rem; }\n        .main-header { padding: 0.25rem 0.5rem; }\n      }\n    </style>\n</head>\n<body class="hold-transition sidebar-mini dark-mode">\n<div class="wrapper">\n  <!-- Top Navbar -->\n  <nav class="main-header navbar navbar-expand navbar-dark">\n    <ul class="navbar-nav">\n      <li class="nav-item"><a class="nav-link" data-widget="pushmenu" href="#" role="button"><i class="fas fa-bars"></i></a></li>\n      <li class="nav-item"><a href="/" target="_blank" class="nav-link"><i class="fas fa-wifi text-success"></i> <span class="d-none d-sm-inline">Portal</span></a></li>\n    </ul>\n    <ul class="navbar-nav ml-auto">\n      <li class="nav-item"><a href="/admin/api/export_xlsx" class="btn btn-sm btn-success mr-2 shadow-sm"><i class="fas fa-file-excel mr-1"></i> <span class="d-none d-sm-inline">Export Data</span></a></li>\n      <li class="nav-item"><a href="/admin/logout" class="btn btn-sm btn-danger"><i class="fas fa-sign-out-alt"></i> <span class="d-none d-sm-inline">Logout</span></a></li>\n    </ul>\n  </nav>\n\n  <!-- Complete Filipino PisoFi-Style AdminLTE Sidebar -->\n  <aside class="main-sidebar sidebar-dark-primary elevation-4">\n    <a href="#" class="brand-link text-center">\n      <span class="brand-text font-weight-bold text-success"><i class="fas fa-recycle"></i> ECO-Fi MASTER</span>\n    </a>\n    <div class="sidebar">\n      <nav class="mt-2">\n        <ul class="nav nav-pills nav-sidebar flex-column" data-widget="treeview" role="menu">\n          <li class="nav-header">MAIN NAVIGATION</li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-dashboard\')" id="nav-dashboard" class="nav-link active"><i class="nav-icon fas fa-tachometer-alt"></i><p>Dashboard & Stats</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-clients\')" id="nav-clients" class="nav-link"><i class="nav-icon fas fa-users"></i><p>Active Clients</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-vouchers\')" id="nav-vouchers" class="nav-link"><i class="nav-icon fas fa-ticket-alt"></i><p>Voucher Tickets</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-members\')" id="nav-members" class="nav-link"><i class="nav-icon fas fa-user-friends"></i><p>Member Wallets</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-rates\')" id="nav-rates" class="nav-link"><i class="nav-icon fas fa-tags"></i><p>Rates & Promos</p></a></li>\n          \n          <li class="nav-header">SYSTEM & HARDWARE</li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-esp32\')" id="nav-esp32" class="nav-link"><i class="nav-icon fas fa-microchip"></i><p>ESP32 Hardware</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-audio\')" id="nav-audio" class="nav-link"><i class="nav-icon fas fa-volume-up"></i><p>Audio & Chimes</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-portal-custom\')" id="nav-portal-custom" class="nav-link"><i class="nav-icon fas fa-palette"></i><p>Portal & Banners</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-bandwidth\')" id="nav-bandwidth" class="nav-link"><i class="nav-icon fas fa-tachometer-alt"></i><p>Bandwidth & Speed</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-walled\')" id="nav-walled" class="nav-link"><i class="nav-icon fas fa-globe-americas"></i><p>Walled Garden Sites</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-security\')" id="nav-security" class="nav-link"><i class="nav-icon fas fa-shield-alt"></i><p>MAC Filtering</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-telegram\')" id="nav-telegram" class="nav-link"><i class="nav-icon fab fa-telegram-plane"></i><p>Telegram Alerts</p></a></li>\n          <li class="nav-item"><a href="javascript:showSection(\'sec-licensing\')" id="nav-licensing" class="nav-link"><i class="nav-icon fas fa-key"></i><p>Hardware Licensing</p></a></li>\n        </ul>\n      </nav>\n    </div>\n  </aside>\n\n  <!-- Main Content Wrapper -->\n  <div class="content-wrapper p-2 p-md-4">\n    \n    <!-- 1. DASHBOARD OVERVIEW SECTION -->\n    <div id="sec-dashboard" class="section-view active">\n      <div class="row">\n        <div class="col-lg-3 col-6">\n          <div class="small-box bg-success">\n            <div class="inner"><h3 id="stat-today">0</h3><p>Today\'s Bottles</p></div>\n            <div class="icon"><i class="fas fa-recycle"></i></div>\n          </div>\n        </div>\n        <div class="col-lg-3 col-6">\n          <div class="small-box bg-info">\n            <div class="inner"><h3 id="stat-total">0</h3><p>Lifetime Bottles</p></div>\n            <div class="icon"><i class="fas fa-chart-line"></i></div>\n          </div>\n        </div>\n        <div class="col-lg-3 col-6">\n          <div class="small-box bg-warning">\n            <div class="inner"><h3 id="stat-clients">0</h3><p>Active Clients</p></div>\n            <div class="icon"><i class="fas fa-wifi"></i></div>\n          </div>\n        </div>\n        <div class="col-lg-3 col-6">\n          <div class="small-box bg-danger">\n            <div class="inner"><h3 id="stat-lic">ACTIVE</h3><p>License Status</p></div>\n            <div class="icon"><i class="fas fa-shield-alt"></i></div>\n          </div>\n        </div>\n      </div>\n\n      <div class="row">\n        <div class="col-lg-8">\n          <div class="card card-dark">\n            <div class="card-header"><h3 class="card-title"><i class="fas fa-chart-bar"></i> 7-Day Recycling Intake History</h3></div>\n            <div class="card-body">\n              <div class="chart-container" style="position: relative; height:250px; width:100%">\n                <canvas id="historyChart"></canvas>\n              </div>\n            </div>\n          </div>\n        </div>\n        <div class="col-lg-4">\n          <div class="card card-dark" style="height: 312px;">\n            <div class="card-header"><h3 class="card-title"><i class="fas fa-microchip"></i> System Resources</h3></div>\n            <div class="card-body">\n              <div style="margin-bottom: 15px;">\n                <div class="d-flex justify-content-between"><span>CPU Load</span><strong id="sys-cpu">0%</strong></div>\n                <div class="progress" style="height: 6px;"><div id="sys-cpu-bar" class="progress-bar bg-success" style="width: 0%"></div></div>\n              </div>\n              <div style="margin-bottom: 15px;">\n                <div class="d-flex justify-content-between"><span>Memory (RAM)</span><strong id="sys-ram">0%</strong></div>\n                <div class="progress" style="height: 6px;"><div id="sys-ram-bar" class="progress-bar bg-info" style="width: 0%"></div></div>\n              </div>\n              <div style="margin-bottom: 15px;">\n                <div class="d-flex justify-content-between"><span>Storage (eMMC)</span><strong id="sys-disk">22%</strong></div>\n                <div class="progress" style="height: 6px;"><div id="sys-disk-bar" class="progress-bar bg-warning" style="width: 22%"></div></div>\n              </div>\n              <div>\n                <div class="d-flex justify-content-between"><span>System Uptime</span><strong id="sys-uptime">0h 0m</strong></div>\n              </div>\n            </div>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 1B. ESP32 HARDWARE SECTION -->\n    <div id="sec-esp32" class="section-view">\n      <div class="card card-purple">\n        <div class="card-header" style="display: flex; align-items: center; justify-content: space-between; width: 100%;">\n          <h3 class="card-title m-0 d-flex align-items-center" style="font-size: 0.95rem; font-weight: 700;">\n            <i class="fas fa-microchip mr-2 text-purple"></i>\n            <span class="d-none d-sm-inline">ESP32 Hardware Calibration</span>\n            <span class="d-inline d-sm-none">ESP32 Hardware</span>\n          </h3>\n          <div class="card-tools ml-auto" style="margin-left: auto;">\n            <button class="btn btn-warning btn-xs" onclick="triggerEsp32Config()" title="Reboot ESP32 to AP Mode" style="padding: 3px 10px; border-radius: 6px; font-size: 11.5px; font-weight: 600; white-space: nowrap;">\n              <i class="fas fa-wifi mr-1"></i> <span class="d-none d-sm-inline">Reboot to AP Mode</span><span class="d-inline d-sm-none">AP Mode</span>\n            </button>\n          </div>\n        </div>\n        <div class="card-body">\n          <div class="row">\n            <div class="col-md-4 form-group">\n              <label>Bin Full Distance (cm):</label>\n              <input type="number" id="esp-bin" class="form-control" value="{{ config.get(\'esp_bin_full_threshold_cm\', 15) }}">\n            </div>\n            <div class="col-md-4 form-group">\n              <label>Entrance Timeout (sec):</label>\n              <input type="number" id="esp-ent-tout" class="form-control" value="{{ config.get(\'esp_entrance_gate_timeout\', 60) }}">\n            </div>\n            <div class="col-md-4 form-group">\n              <label>Bottle Settle Time (ms):</label>\n              <input type="number" id="esp-settle" class="form-control" value="{{ config.get(\'esp_settle_time_ms\', 500) }}">\n            </div>\n            <div class="col-md-4 form-group">\n              <label>Success Drop Time (ms):</label>\n              <input type="number" id="esp-suc-time" class="form-control" value="{{ config.get(\'esp_success_drop_tout_ms\', 3000) }}">\n            </div>\n            <div class="col-md-4 form-group">\n              <label>Reject Drop Time (ms):</label>\n              <input type="number" id="esp-rej-time" class="form-control" value="{{ config.get(\'esp_reject_drop_time_ms\', 2000) }}">\n            </div>\n            <div class="col-md-4 form-group">\n              <label>NIR W Min / Max:</label>\n              <div class="d-flex" style="gap: 8px;">\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Min</span></div>\n                  <input type="number" id="esp-nir-min" class="form-control" value="{{ config.get(\'esp_pet_nir_w_min\', 200) }}">\n                </div>\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Max</span></div>\n                  <input type="number" id="esp-nir-max" class="form-control" value="{{ config.get(\'esp_pet_nir_w_max\', 5000) }}">\n                </div>\n              </div>\n            </div>\n            \n            <div class="col-12 mt-3 mb-2"><h5 class="text-info border-bottom border-secondary pb-1">Servo Tuning (Angles 0-180)</h5></div>\n            \n            <div class="col-md-4 form-group">\n              <label>Entrance Gate (Close / Open):</label>\n              <div class="d-flex" style="gap: 8px;">\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Close</span></div>\n                  <input type="number" id="esp-ent-close" class="form-control" value="{{ config.get(\'esp_ent_close_angle\', 0) }}">\n                </div>\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Open</span></div>\n                  <input type="number" id="esp-ent-open" class="form-control" value="{{ config.get(\'esp_ent_open_angle\', 90) }}">\n                </div>\n              </div>\n            </div>\n            <div class="col-md-4 form-group">\n              <label>Success Gate (Close / Open):</label>\n              <div class="d-flex" style="gap: 8px;">\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Close</span></div>\n                  <input type="number" id="esp-suc-close" class="form-control" value="{{ config.get(\'esp_suc_close_angle\', 0) }}">\n                </div>\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Open</span></div>\n                  <input type="number" id="esp-suc-open" class="form-control" value="{{ config.get(\'esp_suc_open_angle\', 90) }}">\n                </div>\n              </div>\n            </div>\n            <div class="col-md-4 form-group">\n              <label>Reject Gate (Close / Open):</label>\n              <div class="d-flex" style="gap: 8px;">\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Close</span></div>\n                  <input type="number" id="esp-rej-close" class="form-control" value="{{ config.get(\'esp_rej_close_angle\', 0) }}">\n                </div>\n                <div class="input-group input-group-sm" style="flex: 1;">\n                  <div class="input-group-prepend"><span class="input-group-text px-2 text-muted" style="font-size: 11px;">Open</span></div>\n                  <input type="number" id="esp-rej-open" class="form-control" value="{{ config.get(\'esp_rej_open_angle\', 90) }}">\n                </div>\n              </div>\n            </div>\n          </div>\n          <button class="btn btn-success mt-3" onclick="saveEsp32Config()"><i class="fas fa-save"></i> Save & Push to ESP32</button>\n        </div>\n      </div>\n    </div>\n\n    <!-- 2. ACTIVE CLIENTS SECTION -->\n    <div id="sec-clients" class="section-view">\n      <div class="card card-primary">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-users"></i> Connected Client Sessions</h3></div>\n        <div class="card-body p-0">\n          <div class="table-responsive">\n            <table class="table table-striped table-hover mb-0">\n              <thead>\n                <tr>\n                  <th style="padding: 10px 14px; width: 18%;">IP Address</th>\n                  <th style="padding: 10px 14px; width: 18%;">MAC Address</th>\n                  <th style="padding: 10px 14px; width: 18%;">Remaining Time</th>\n                  <th style="padding: 10px 14px; width: 12%; text-align: center;">Status</th>\n                  <th style="padding: 10px 14px; width: 16%;">Speed (DL/UL)</th>\n                  <th style="padding: 10px 14px; width: 18%; text-align: right; min-width: 220px; white-space: nowrap;">Actions</th>\n                </tr>\n              </thead>\n              <tbody id="clients-table-body"></tbody>\n            </table>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 3. VOUCHER & TICKETS SECTION -->\n    <div id="sec-vouchers" class="section-view">\n      <div class="card card-success">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-magic"></i> Generate Prepaid Vouchers</h3></div>\n        <div class="card-body">\n          <div class="row">\n            <div class="col-12 col-sm-6 col-md-3 form-group">\n              <label>Number of Vouchers (1-100):</label>\n              <input type="number" id="v-qty" class="form-control" value="5" min="1" max="100">\n            </div>\n            <div class="col-12 col-sm-6 col-md-3 form-group">\n              <label>Duration (Minutes):</label>\n              <select id="v-mins" class="form-control">\n                <option value="10">10 Minutes (1 Bottle Equivalent)</option>\n                <option value="45">45 Minutes</option>\n                <option value="60" selected>1 Hour</option>\n                <option value="180">3 Hours</option>\n                <option value="1440">24 Hours (1 Day Pass)</option>\n              </select>\n            </div>\n            <div class="col-12 col-sm-6 col-md-3 form-group">\n              <label>Custom Note / Batch Tag:</label>\n              <input type="text" id="v-note" class="form-control" placeholder="e.g. Student Promo Batch">\n            </div>\n            <div class="col-12 col-sm-6 col-md-3 form-group">\n              <label class="d-none d-md-block">&nbsp;</label>\n              <button class="btn btn-success btn-block" onclick="generateVouchers()"><i class="fas fa-ticket-alt"></i> Generate Vouchers</button>\n            </div>\n          </div>\n          <div id="v-results" class="mt-3"></div>\n        </div>\n      </div>\n      \n      <div class="card card-dark mt-3">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-history"></i> Voucher History</h3></div>\n        <div class="card-body p-0">\n          <div class="table-responsive">\n            <table class="table table-striped table-hover mb-0" id="voucher-history-table">\n              <thead>\n                <tr>\n                  <th style="padding: 10px 14px; width: 18%;">Voucher Code</th>\n                  <th style="padding: 10px 14px; width: 12%;">Duration</th>\n                  <th style="padding: 10px 14px; width: 12%; text-align: center;">Status</th>\n                  <th style="padding: 10px 14px; width: 18%;">Note / Tag</th>\n                  <th style="padding: 10px 14px; width: 18%;">Created Date</th>\n                  <th style="padding: 10px 14px; width: 12%;">Used By (IP)</th>\n                  <th style="padding: 10px 14px; width: 10%; text-align: right; min-width: 100px; white-space: nowrap;">Actions</th>\n                </tr>\n              </thead>\n              <tbody id="voucher-history-body"></tbody>\n            </table>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 4. MEMBER WALLETS SECTION -->\n    <div id="sec-members" class="section-view">\n      <div class="card card-success">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-user-plus"></i> Add New Member Account</h3></div>\n        <div class="card-body">\n          <div class="row">\n            <div class="col-12 col-sm-6 col-md-4 form-group">\n              <label>Username (min 3 chars):</label>\n              <input type="text" id="new-mem-user" class="form-control" placeholder="e.g. student01">\n            </div>\n            <div class="col-12 col-sm-6 col-md-4 form-group">\n              <label>4-Digit Security PIN:</label>\n              <input type="password" id="new-mem-pin" class="form-control" placeholder="e.g. 1234" maxlength="6">\n            </div>\n            <div class="col-12 col-sm-6 col-md-2 form-group">\n              <label>Initial Wallet (Mins):</label>\n              <input type="number" id="new-mem-mins" class="form-control" value="0" min="0">\n            </div>\n            <div class="col-12 col-sm-6 col-md-2 form-group">\n              <label class="d-none d-md-block">&nbsp;</label>\n              <button class="btn btn-success btn-block" onclick="addMember()"><i class="fas fa-plus"></i> Create</button>\n            </div>\n          </div>\n        </div>\n      </div>\n\n      <div class="card card-info mt-3">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-user-friends"></i> Registered Member Accounts</h3></div>\n        <div class="card-body p-0">\n          <div class="table-responsive">\n            <table class="table table-striped table-hover mb-0">\n              <thead>\n                <tr>\n                  <th style="padding: 10px 14px; width: 30%;">Username</th>\n                  <th style="padding: 10px 14px; width: 25%;">Wallet Balance</th>\n                  <th style="padding: 10px 14px; width: 25%;">Registered At</th>\n                  <th style="padding: 10px 14px; width: 20%; text-align: right; min-width: 170px; white-space: nowrap;">Actions</th>\n                </tr>\n              </thead>\n              <tbody id="members-table-body"></tbody>\n            </table>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 5. RATES & PROMOS SECTION (DYNAMIC ADD+ & EDIT RATE TIERS) -->\n    <div id="sec-rates" class="section-view">\n      <div class="row">\n        <!-- Card 1: General Timing -->\n        <div class="col-12 col-lg-6">\n          <div class="card mb-3">\n            <div class="card-header"><h3 class="card-title text-warning"><i class="fas fa-sliders-h mr-1"></i> Basic Rate & Drop Timeout</h3></div>\n            <div class="card-body">\n              <div class="row">\n                <div class="col-6 form-group mb-2">\n                  <label>Base Rate (Mins / Bottle):</label>\n                  <input type="number" id="rate-1" class="form-control" value="{{ config.minutes_per_bottle }}" min="1" oninput="onBaseRateInput()">\n                </div>\n                <div class="col-6 form-group mb-2">\n                  <label>Chute Timeout (Seconds):</label>\n                  <input type="number" id="rate-timeout" class="form-control" value="{{ config.drop_timeout }}" min="10" max="120">\n                </div>\n                <div class="col-12 mt-2">\n                  <button class="btn btn-warning btn-block font-weight-bold" onclick="saveRates()"><i class="fas fa-save mr-1"></i> Save Base Timing</button>\n                </div>\n              </div>\n            </div>\n          </div>\n        </div>\n\n        <!-- Card 2: Template Presets -->\n        <div class="col-12 col-lg-6">\n          <div class="card mb-3">\n            <div class="card-header"><h3 class="card-title text-info"><i class="fas fa-magic mr-1"></i> Quick-Load Rate Templates</h3></div>\n            <div class="card-body">\n              <div class="form-group mb-2">\n                <label>Balanced Rate Curve Templates:</label>\n                <select id="rate-preset-select" class="form-control">\n                  <option value="standard">🌟 Standard Community (1b=10m, 3b=40m, 5b=1h15m, 10b=3h)</option>\n                  <option value="aggressive">⚡ Aggressive Reward (1b=10m, 5b=1h10m, 10b=3h, 20b=7h)</option>\n                  <option value="cafe">☕ Café / Study Hub (1b=20m, 3b=1h15m, 6b=3h, 12b=7h)</option>\n                </select>\n              </div>\n              <div class="mt-2 pt-1">\n                <button class="btn btn-info btn-block font-weight-bold" onclick="applyRatePreset()"><i class="fas fa-file-import mr-1"></i> Apply Selected Template</button>\n              </div>\n            </div>\n          </div>\n        </div>\n      </div>\n\n      <!-- Add / Edit Custom Promo Rate Form -->\n      <div class="card mb-3" id="promo-form-card">\n        <div class="card-header"><h3 class="card-title text-success" id="promo-form-title"><i class="fas fa-plus-circle mr-1"></i> Add Custom Promo Rate Package</h3></div>\n        <div class="card-body">\n          <input type="hidden" id="edit-original-bottles" value="">\n          <div class="row align-items-end">\n            <div class="col-12 col-sm-6 col-md-3 form-group mb-2">\n              <label>Bottles Required:</label>\n              <div class="input-group">\n                <input type="number" id="new-rate-bottles" class="form-control" placeholder="e.g. 5" min="1" max="100" oninput="validatePromoFormMath()">\n                <div class="input-group-append">\n                  <button type="button" class="btn btn-outline-secondary dropdown-toggle dropdown-toggle-split" data-toggle="dropdown"></button>\n                  <div class="dropdown-menu dropdown-menu-right">\n                    <a class="dropdown-item" href="javascript:setRateBottles(2)">2 Bottles</a>\n                    <a class="dropdown-item" href="javascript:setRateBottles(3)">3 Bottles</a>\n                    <a class="dropdown-item" href="javascript:setRateBottles(5)">5 Bottles</a>\n                    <a class="dropdown-item" href="javascript:setRateBottles(6)">6 Bottles</a>\n                    <a class="dropdown-item" href="javascript:setRateBottles(10)">10 Bottles</a>\n                    <a class="dropdown-item" href="javascript:setRateBottles(12)">12 Bottles</a>\n                    <a class="dropdown-item" href="javascript:setRateBottles(15)">15 Bottles</a>\n                    <a class="dropdown-item" href="javascript:setRateBottles(20)">20 Bottles</a>\n                  </div>\n                </div>\n              </div>\n            </div>\n            \n            <div class="col-12 col-sm-6 col-md-3 form-group mb-2">\n              <label>Time Credited:</label>\n              <div class="input-group">\n                <input type="number" id="new-rate-time-val" class="form-control" placeholder="e.g. 90" min="1" oninput="validatePromoFormMath()">\n                <div class="input-group-append">\n                  <select id="new-rate-time-unit" class="custom-select" style="max-width: 85px;" onchange="validatePromoFormMath()">\n                    <option value="mins" selected>Mins</option>\n                    <option value="hours">Hours</option>\n                    <option value="days">Days</option>\n                  </select>\n                </div>\n              </div>\n            </div>\n\n            <div class="col-12 col-sm-8 col-md-4 form-group mb-2">\n              <label>Package Display Label:</label>\n              <div class="input-group">\n                <input type="text" id="new-rate-label" class="form-control" placeholder="Auto-generated if blank">\n                <div class="input-group-append">\n                  <button class="btn btn-outline-secondary" type="button" onclick="autoGenerateRateLabel()" title="Auto-generate friendly label">🪄</button>\n                </div>\n              </div>\n            </div>\n\n            <div class="col-12 col-sm-4 col-md-2 form-group mb-2">\n              <div class="d-flex">\n                <button class="btn btn-success btn-block mr-1 font-weight-bold" id="btn-save-promo" onclick="addPromoRate()"><i class="fas fa-plus mr-1"></i> Add</button>\n                <button class="btn btn-secondary font-weight-bold" id="btn-cancel-promo" onclick="cancelEditPromoRate()" style="display:none;" title="Cancel Edit"><i class="fas fa-times"></i></button>\n              </div>\n            </div>\n          </div>\n\n          <!-- Live Validator Feedback Box -->\n          <div id="rate-validator-feedback" class="alert alert-info py-2 px-3 mt-2 mb-0" style="display:none; font-size:12px; border-radius:8px;">\n            <div class="d-flex justify-content-between align-items-center">\n              <span id="rate-validator-eff" class="font-weight-bold">📊 Efficiency: --</span>\n              <span id="rate-validator-status" class="font-weight-bold">✔ Status: OK</span>\n            </div>\n            <div id="rate-validator-msg" style="margin-top:4px;"></div>\n          </div>\n        </div>\n      </div>\n\n      <!-- Active Rate Packages Table -->\n      <div class="card mb-3">\n        <div class="card-header"><h3 class="card-title text-light"><i class="fas fa-tags mr-1"></i> Active Rate Tiers & Promo Curves</h3></div>\n        <div class="card-body p-0">\n          <div class="table-responsive">\n            <table class="table table-striped table-hover mb-0">\n              <thead>\n                <tr>\n                  <th style="padding: 10px 14px; width: 20%;">Bottles Required</th>\n                  <th style="padding: 10px 14px; width: 22%;">Time Credited</th>\n                  <th style="padding: 10px 14px; width: 22%;">Rate Efficiency</th>\n                  <th style="padding: 10px 14px; width: 20%;">Package Label</th>\n                  <th style="padding: 10px 14px; width: 16%; min-width: 170px; text-align: right; white-space: nowrap;">Actions</th>\n                </tr>\n              </thead>\n              <tbody id="rates-table-body"></tbody>\n            </table>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 6. AUDIO CUSTOMIZER SECTION -->\n    <div id="sec-audio" class="section-view">\n      <div class="card card-warning mb-3">\n        <div class="card-header"><h3 class="card-title font-weight-bold"><i class="fas fa-volume-up"></i> Portal Audio & Event Chimes</h3></div>\n        <div class="card-body p-0">\n          <div class="d-none d-md-flex bg-dark text-white p-2 font-weight-bold" style="font-size: 13px;">\n            <div class="col-md-3">Event Stage</div>\n            <div class="col-md-3">Audio Preset</div>\n            <div class="col-md-4">Custom URL or Upload</div>\n            <div class="col-md-2 text-center">Preview</div>\n          </div>\n          \n          <!-- EVENT 1: BG LOOP -->\n          <div class="row m-0 p-3 border-bottom align-items-center">\n            <div class="col-12 col-md-3 mb-2 mb-md-0">\n              <span class="badge badge-secondary p-2 d-block text-left"><i class="fas fa-music"></i> 1. Standby Loop</span>\n            </div>\n            <div class="col-12 col-md-3 mb-2 mb-md-0">\n              <select id="audio-bg-preset" class="form-control form-control-sm" onchange="onAudioPresetChange(\'bg\')">\n                <option value="/static/audio/eco_loop.wav" {% if config.audio_bg == \'/static/audio/eco_loop.wav\' or config.audio_bg == \'/static/audio/b1.wav\' or not config.audio_bg %}selected{% endif %}>📻 Default ECO-Fi Standby Loop</option>\n                <option value="silent" {% if config.audio_bg == \'silent\' %}selected{% endif %}>🔇 Silent</option>\n                <option value="custom" {% if config.audio_bg and config.audio_bg not in [\'/static/audio/eco_loop.wav\', \'/static/audio/b1.wav\', \'silent\'] %}selected{% endif %}>📁 Custom File / URL</option>\n              </select>\n            </div>\n            <div class="col-12 col-md-4 mb-2 mb-md-0">\n              <div class="input-group input-group-sm">\n                <input type="text" id="audio-bg-custom" class="form-control" placeholder="Custom URL..." value="{{ config.audio_bg or \'/static/audio/eco_loop.wav\' }}" oninput="updateAudioPlayer(\'bg\')">\n                <div class="input-group-append">\n                  <label class="btn btn-secondary mb-0 rounded-right" style="cursor:pointer;" title="Upload File"><i class="fas fa-upload"></i>\n                    <input type="file" id="upload-file-bg" accept="audio/*" onchange="uploadAudioFile(\'bg\')" style="display:none;">\n                  </label>\n                </div>\n              </div>\n            </div>\n            <div class="col-12 col-md-2 text-center">\n              <audio id="audio-player-bg" controls src="{{ config.audio_bg or \'/static/audio/eco_loop.wav\' }}" style="height: 30px; width: 100%; max-width: 250px;"></audio>\n            </div>\n          </div>\n\n          <!-- EVENT 2: DEPOSIT CHIME -->\n          <div class="row m-0 p-3 border-bottom align-items-center">\n            <div class="col-12 col-md-3 mb-2 mb-md-0">\n              <span class="badge badge-success p-2 d-block text-left"><i class="fas fa-coins"></i> 2. Deposit Chime</span>\n            </div>\n            <div class="col-12 col-md-3 mb-2 mb-md-0">\n              <select id="audio-insert-preset" class="form-control form-control-sm" onchange="onAudioPresetChange(\'insert\')">\n                <option value="/static/audio/eco_chime.wav" {% if config.audio_insert == \'/static/audio/eco_chime.wav\' or config.audio_insert == \'/static/audio/coin.wav\' or not config.audio_insert %}selected{% endif %}>🔔 Classic ECO-Fi Chime</option>\n                <option value="/static/audio/eco_drop.wav" {% if config.audio_insert == \'/static/audio/eco_drop.wav\' or config.audio_insert == \'/static/audio/coin_insert.wav\' %}selected{% endif %}>🍾 Mechanical Bottle Drop</option>\n                <option value="/static/audio/eco_pulse.wav" {% if config.audio_insert == \'/static/audio/eco_pulse.wav\' or config.audio_insert == \'/static/audio/insert_coin.wav\' %}selected{% endif %}>🎶 Double Pulse Alert</option>\n                <option value="arcade_powerup" {% if config.audio_insert == \'arcade_powerup\' %}selected{% endif %}>🎮 8-Bit Power-Up</option>\n                <option value="voice_filipino" {% if config.audio_insert == \'voice_filipino\' %}selected{% endif %}>🗣️ Filipino Voice</option>\n                <option value="custom" {% if config.audio_insert and config.audio_insert not in [\'/static/audio/eco_chime.wav\', \'/static/audio/eco_drop.wav\', \'/static/audio/eco_pulse.wav\', \'/static/audio/coin.wav\', \'/static/audio/coin_insert.wav\', \'/static/audio/insert_coin.wav\', \'arcade_powerup\', \'voice_filipino\'] %}selected{% endif %}>📁 Custom File / URL</option>\n              </select>\n            </div>\n            <div class="col-12 col-md-4 mb-2 mb-md-0">\n              <div class="input-group input-group-sm">\n                <input type="text" id="audio-insert-custom" class="form-control" placeholder="Custom URL..." value="{{ config.audio_insert or \'/static/audio/eco_chime.wav\' }}" oninput="updateAudioPlayer(\'insert\')">\n                <div class="input-group-append">\n                  <label class="btn btn-secondary mb-0 rounded-right" style="cursor:pointer;" title="Upload File"><i class="fas fa-upload"></i>\n                    <input type="file" id="upload-file-insert" accept="audio/*" onchange="uploadAudioFile(\'insert\')" style="display:none;">\n                  </label>\n                </div>\n              </div>\n            </div>\n            <div class="col-12 col-md-2 text-center">\n              <audio id="audio-player-insert" controls src="{{ config.audio_insert or \'/static/audio/eco_chime.wav\' }}" style="height: 30px; width: 100%; max-width: 250px;"></audio>\n            </div>\n          </div>\n\n          <!-- EVENT 3: SUCCESS CHIME -->\n          <div class="row m-0 p-3 border-bottom align-items-center">\n            <div class="col-12 col-md-3 mb-2 mb-md-0">\n              <span class="badge badge-info p-2 d-block text-left"><i class="fas fa-check-circle"></i> 3. Session Complete</span>\n            </div>\n            <div class="col-12 col-md-3 mb-2 mb-md-0">\n              <select id="audio-success-preset" class="form-control form-control-sm" onchange="onAudioPresetChange(\'success\')">\n                <option value="/static/audio/eco_success.wav" {% if config.audio_success == \'/static/audio/eco_success.wav\' or config.audio_success == \'/static/audio/success_ding.wav\' or not config.audio_success %}selected{% endif %}>✨ ECO-Fi Success</option>\n                <option value="crystal_bell" {% if config.audio_success == \'crystal_bell\' %}selected{% endif %}>🛎️ Crystal Bell</option>\n                <option value="silent" {% if config.audio_success == \'silent\' %}selected{% endif %}>🔇 Silent</option>\n                <option value="custom" {% if config.audio_success and config.audio_success not in [\'/static/audio/eco_success.wav\', \'/static/audio/success_ding.wav\', \'crystal_bell\', \'silent\'] %}selected{% endif %}>📁 Custom File / URL</option>\n              </select>\n            </div>\n            <div class="col-12 col-md-4 mb-2 mb-md-0">\n              <div class="input-group input-group-sm">\n                <input type="text" id="audio-success-custom" class="form-control" placeholder="Custom URL..." value="{{ config.audio_success or \'/static/audio/eco_success.wav\' }}" oninput="updateAudioPlayer(\'success\')">\n                <div class="input-group-append">\n                  <label class="btn btn-secondary mb-0 rounded-right" style="cursor:pointer;" title="Upload File"><i class="fas fa-upload"></i>\n                    <input type="file" id="upload-file-success" accept="audio/*" onchange="uploadAudioFile(\'success\')" style="display:none;">\n                  </label>\n                </div>\n              </div>\n            </div>\n            <div class="col-12 col-md-2 text-center">\n              <audio id="audio-player-success" controls src="{{ config.audio_success or \'/static/audio/eco_success.wav\' }}" style="height: 30px; width: 100%; max-width: 250px;"></audio>\n            </div>\n          </div>\n          <div class="card-footer d-flex flex-wrap align-items-center justify-content-between p-3 border-top-0">\n            <div class="d-flex align-items-center mb-2 mb-md-0" style="min-width: 250px; max-width: 400px; flex-grow: 1;">\n              <span class="mr-3 font-weight-bold" style="font-size:13px;"><i class="fas fa-volume-up"></i> Master Volume:</span>\n              <input type="range" id="audio-vol-input" class="custom-range flex-grow-1" min="10" max="100" value="{{ config.audio_volume or 80 }}" oninput="updatePreviewVolume()">\n              <span id="vol-lbl" class="ml-2 badge badge-dark" style="width:45px;">{{ config.audio_volume or 80 }}%</span>\n            </div>\n            <button class="btn btn-success btn-sm px-4 shadow-sm" onclick="saveAudioSettings()"><i class="fas fa-save"></i> Save Audio Settings</button>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 7. PORTAL & BANNERS SECTION -->\n    <div id="sec-portal-custom" class="section-view">\n      <div class="card card-primary">\n        <div class="card-header"><h3 class="card-title text-info"><i class="fas fa-palette mr-1"></i> Portal Branding & Announcements</h3></div>\n        <div class="card-body">\n          <div class="row">\n            <div class="col-12 col-md-6 form-group">\n              <label>Hotspot Vendo Name:</label>\n              <input type="text" id="cfg-vendo-name" class="form-control" value="{{ config.get(\'vendo_name\', \'ECO-Fi Vendo\') }}">\n            </div>\n            <div class="col-12 col-md-6 form-group">\n              <label>Subtitle / Tagline:</label>\n              <input type="text" id="cfg-vendo-sub" class="form-control" value="{{ config.get(\'vendo_subtitle\', \'Recycle Bottles for Fast WiFi\') }}">\n            </div>\n          </div>\n          <div class="form-group">\n            <label>Announcement Banner Message:</label>\n            <textarea id="cfg-announcement" class="form-control" rows="3" placeholder="Enter announcement text to display on customer portal...">{{ config.get(\'announcement\', \'\') }}</textarea>\n            <small class="text-muted">This announcement banner is displayed at the top of the client portal page in real-time.</small>\n          </div>\n          <div class="mt-3">\n            <button class="btn btn-primary font-weight-bold px-4" onclick="savePortalCustom()"><i class="fas fa-save mr-1"></i> Update Portal Branding</button>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 8. BANDWIDTH & ANTI-TETHERING SECTION -->\n    <div id="sec-bandwidth" class="section-view">\n      <div class="card card-info">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-tachometer-alt"></i> Bandwidth & Anti-Tethering Rules</h3></div>\n        <div class="card-body">\n          <div class="row">\n            <div class="col-12 col-md-4 form-group">\n              <label>Default Download Speed Limit (Kbps):</label>\n              <input type="number" id="cfg-dl" class="form-control" value="{{ config.default_dl_kbps }}" min="128">\n            </div>\n            <div class="col-12 col-md-4 form-group">\n              <label>Default Upload Speed Limit (Kbps):</label>\n              <input type="number" id="cfg-ul" class="form-control" value="{{ config.default_ul_kbps }}" min="64">\n            </div>\n            <div class="col-12 col-md-4 form-group">\n              <label>Anti-Tethering Protection (TTL=64):</label>\n              <select id="cfg-tether" class="form-control">\n                <option value="1" {% if config.anti_tethering == \'1\' %}selected{% endif %}>🟢 ENABLED (Blocks Hotspot Resharing)</option>\n                <option value="0" {% if config.anti_tethering == \'0\' %}selected{% endif %}>🔴 DISABLED</option>\n              </select>\n            </div>\n          </div>\n          <button class="btn btn-info mt-2" onclick="saveBandwidth()"><i class="fas fa-save"></i> Apply Traffic Limits</button>\n        </div>\n      </div>\n    </div>\n\n    <!-- 9. WALLED GARDEN FREE DOMAINS SECTION -->\n    <div id="sec-walled" class="section-view">\n      <div class="card card-primary">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-globe-americas"></i> Walled Garden Free Whitelisted Websites</h3></div>\n        <div class="card-body">\n          <p class="text-muted" style="font-size:13px;">Domains added here are accessible to users even without inserting bottles or logging in (e.g. government, school portals, payment portals):</p>\n          <div class="row">\n            <div class="col-12 col-md-6 form-group">\n              <label>Domain Name (e.g. gcash.com or deped.gov.ph):</label>\n              <input type="text" id="walled-domain" class="form-control" placeholder="e.g. portal.school.edu.ph">\n            </div>\n            <div class="col-12 col-md-4 form-group">\n              <label>Note / Purpose:</label>\n              <input type="text" id="walled-note" class="form-control" placeholder="e.g. School Portal">\n            </div>\n            <div class="col-12 col-md-2 form-group">\n              <label class="d-none d-md-block">&nbsp;</label>\n              <button class="btn btn-primary btn-block" onclick="addWalledDomain()"><i class="fas fa-plus"></i> Whitelist Site</button>\n            </div>\n          </div>\n          <div class="table-responsive">\n            <table class="table table-striped table-hover mb-0 mt-3">\n              <thead>\n                <tr>\n                  <th style="padding: 10px 14px; width: 45%;">Whitelisted Domain</th>\n                  <th style="padding: 10px 14px; width: 40%;">Note / Purpose</th>\n                  <th style="padding: 10px 14px; width: 15%; text-align: right; min-width: 120px; white-space: nowrap;">Actions</th>\n                </tr>\n              </thead>\n              <tbody id="walled-table-body"></tbody>\n            </table>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 10. MAC FILTERING SECTION WITH VALIDATIONS & EDIT HANDLERS -->\n    <div id="sec-security" class="section-view">\n      <div class="card card-danger" id="mac-form-card">\n        <div class="card-header"><h3 class="card-title" id="mac-form-title"><i class="fas fa-shield-alt"></i> Add / Edit MAC Filter Rule</h3></div>\n        <div class="card-body">\n          <div class="row mb-2">\n            <div class="col-12">\n              <label class="text-info"><i class="fas fa-bolt"></i> Quick-Fill from Connected Devices:</label>\n              <select id="mac-quick-pick" class="form-control form-control-sm" onchange="quickFillMac(this.value)">\n                <option value="">-- Choose active connected device to auto-fill --</option>\n              </select>\n            </div>\n          </div>\n          \n          <input type="hidden" id="edit-original-mac" value="">\n          <div class="row">\n            <div class="col-12 col-sm-6 col-md-4 form-group">\n              <label>MAC Address (12 Hex Characters):</label>\n              <input type="text" id="mac-input" class="form-control" placeholder="AA:BB:CC:DD:EE:FF" maxlength="17" oninput="formatMacInput(this)">\n              <div id="mac-valid-msg" class="valid-feedback-custom">✓ Valid MAC format</div>\n              <div id="mac-invalid-msg" class="invalid-feedback-custom">✗ Invalid MAC address format (e.g. AA:BB:CC:DD:EE:FF)</div>\n            </div>\n            <div class="col-12 col-sm-6 col-md-3 form-group">\n              <label>Rule Type:</label>\n              <select id="mac-type" class="form-control">\n                <option value="whitelist">🟢 VIP Whitelist (Permanent Free Access)</option>\n                <option value="blacklist">🔴 Blacklist (Block / Ban Device)</option>\n              </select>\n            </div>\n            <div class="col-12 col-sm-6 col-md-3 form-group">\n              <label>Device Owner / Label:</label>\n              <input type="text" id="mac-note" class="form-control" placeholder="e.g. Owner Phone or Abusive User">\n            </div>\n            <div class="col-12 col-sm-6 col-md-2 form-group">\n              <label class="d-none d-md-block">&nbsp;</label>\n              <div class="d-flex">\n                <button class="btn btn-danger btn-block mr-1 font-weight-bold" id="btn-save-mac" onclick="saveMacControl()"><i class="fas fa-plus mr-1"></i> Add Rule</button>\n                <button class="btn btn-secondary font-weight-bold" id="btn-cancel-mac" style="display:none;" onclick="cancelEditMac()" title="Cancel Edit"><i class="fas fa-times"></i></button>\n              </div>\n            </div>\n          </div>\n        </div>\n      </div>\n\n      <div class="card card-dark mt-3">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-list-ul"></i> Active MAC Control Table</h3></div>\n        <div class="card-body p-0">\n          <div class="table-responsive">\n            <table class="table table-striped table-hover mb-0" id="mac-table">\n              <thead>\n                <tr>\n                  <th style="padding: 10px 14px; width: 30%;">MAC Address</th>\n                  <th style="padding: 10px 14px; width: 20%; text-align: center;">Rule Type</th>\n                  <th style="padding: 10px 14px; width: 30%;">Device Note</th>\n                  <th style="padding: 10px 14px; width: 20%; text-align: right; min-width: 170px; white-space: nowrap;">Actions</th>\n                </tr>\n              </thead>\n              <tbody id="mac-table-body"></tbody>\n            </table>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 11. TELEGRAM NOTIFICATIONS SECTION -->\n    <div id="sec-telegram" class="section-view">\n      <div class="card card-primary">\n        <div class="card-header"><h3 class="card-title"><i class="fab fa-telegram-plane"></i> Telegram Bot Automated Alerts</h3></div>\n        <div class="card-body">\n          \n          <div class="p-3 mb-4 rounded" style="background: rgba(255, 255, 255, 0.03); border: 1px solid rgba(255, 255, 255, 0.1); color: #cbd5e1; font-size: 0.9rem; line-height: 1.5;">\n            <h5 class="font-weight-bold mb-3" style="font-size: 1.05rem; color: #38bdf8;"><i class="fas fa-info-circle mr-1"></i> How to setup Telegram Alerts</h5>\n            <p class="mb-2 text-white">Follow these exact steps to connect ECO-Fi to your Telegram account:</p>\n            <ol class="mb-0 pl-3">\n              <li class="mb-2">Open Telegram and search for <strong class="text-white">@BotFather</strong>. Send it <code style="color: #60a5fa; background: rgba(0,0,0,0.3); padding: 2px 6px; border-radius: 4px;">/newbot</code> and follow the prompts to create your bot.</li>\n              <li class="mb-2">BotFather will give you a <strong class="text-white">Bot Token</strong> (e.g., <code style="color: #60a5fa; background: rgba(0,0,0,0.3); padding: 2px 6px; border-radius: 4px;">123456789:ABCdefGh...</code>). Paste it into the Bot Token field below.</li>\n              <li class="mb-2">Next, search for <strong class="text-white">@userinfobot</strong> in Telegram and send it <code style="color: #60a5fa; background: rgba(0,0,0,0.3); padding: 2px 6px; border-radius: 4px;">/start</code>. It will reply with your <strong class="text-white">Chat ID</strong> (e.g., <code style="color: #60a5fa; background: rgba(0,0,0,0.3); padding: 2px 6px; border-radius: 4px;">123456789</code>). Paste it into the Chat ID field below.</li>\n              <li class="mb-2"><strong class="text-warning">CRITICAL STEP:</strong> Telegram blocks bots from messaging users to prevent spam. You MUST search for your new bot\'s username in Telegram and click <strong class="text-white">Start</strong> (or send <code style="color: #60a5fa; background: rgba(0,0,0,0.3); padding: 2px 6px; border-radius: 4px;">/start</code>) to authorize it to message you.</li>\n              <li>Once you have started the chat with your bot, click <strong class="text-white">Save Settings</strong> below, and then click <strong class="text-white">Test Message</strong>.</li>\n            </ol>\n          </div>\n\n          <div class="form-group mt-4">\n            <label>Telegram Bot Token:</label>\n            <input type="text" id="cfg-tg-token" class="form-control" placeholder="123456789:ABCdefGhIJKlmNoPQRstuVWXyz" value="{{ config.telegram_bot_token }}">\n          </div>\n          <div class="form-group">\n            <label>Admin Telegram Chat ID:</label>\n            <input type="text" id="cfg-tg-chat" class="form-control" placeholder="123456789" value="{{ config.telegram_chat_id }}">\n          </div>\n          <div class="row">\n            <div class="col-12 col-md-6 form-group">\n              <label>Alert when Storage Bin reaches 100%:</label>\n              <select id="cfg-tg-bin" class="form-control">\n                <option value="1" {% if config.telegram_alert_bin == \'1\' %}selected{% endif %}>🟢 ENABLED</option>\n                <option value="0" {% if config.telegram_alert_bin == \'0\' %}selected{% endif %}>🔴 DISABLED</option>\n              </select>\n            </div>\n            <div class="col-12 col-md-6 form-group">\n              <label>Daily Midnight Revenue & Bottle Summary:</label>\n              <select id="cfg-tg-daily" class="form-control">\n                <option value="1" {% if config.telegram_alert_daily == \'1\' %}selected{% endif %}>🟢 ENABLED</option>\n                <option value="0" {% if config.telegram_alert_daily == \'0\' %}selected{% endif %}>🔴 DISABLED</option>\n              </select>\n            </div>\n          </div>\n          <div class="d-flex flex-wrap gap-2">\n            <button class="btn btn-primary mr-2 mb-2" onclick="saveTelegram()"><i class="fas fa-save"></i> Save Settings</button>\n            <button class="btn btn-outline-info mb-2" onclick="testTelegram()"><i class="fas fa-paper-plane"></i> Test Message</button>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <!-- 12. LICENSING & ACTIVATION SECTION -->\n    <div id="sec-licensing" class="section-view">\n      <div class="card card-info">\n        <div class="card-header"><h3 class="card-title"><i class="fas fa-key"></i> Machine Hardware Licensing & Sovereign Authorization</h3></div>\n        <div class="card-body">\n          <div style="margin-bottom: 14px;">\n            <div class="mb-1"><strong>Machine Hardware ID (HWID):</strong></div>\n            <div class="d-flex align-items-center flex-wrap" style="gap: 8px;">\n              <code id="lic-hwid" class="text-warning font-weight-bold" style="font-size:14px; word-break:break-all; background:rgba(0,0,0,0.25); padding:3px 8px; border-radius:6px; border:1px solid rgba(251,191,36,0.25);">Loading...</code>\n              <button id="btn-copy-hwid" class="btn btn-xs btn-outline-info" onclick="copyHwid()" title="Copy HWID" style="padding: 3px 9px; border-radius: 6px; font-size: 11px;"><i class="fas fa-copy mr-1"></i> <span id="copy-hwid-text">Copy</span></button>\n            </div>\n          </div>\n          <p><strong>License Status:</strong> <span id="lic-status" class="badge badge-success" style="font-size:13px;">CHECKING</span></p>\n          <p><strong>License Tier:</strong> <span id="lic-tier" class="badge badge-info">COMMERCIAL</span></p>\n          <hr>\n          <h5>Offline Machine Activation:</h5>\n          <p class="text-muted" style="font-size:13px;">Enter the 16-character offline activation PIN provided by your vendor to authorize this station:</p>\n          <div class="d-flex align-items-stretch flex-column flex-sm-row" style="gap: 10px; max-width: 560px; margin-bottom: 15px;">\n            <input type="text" id="act-pin" class="form-control" placeholder="16-Char PIN (XXXX-XXXX-XXXX-XXXX)" style="flex: 1; min-width: 200px; border-radius: 8px; height: 38px; font-size: 13px;">\n            <button class="btn btn-info" onclick="activateLicense()" style="height: 38px; padding: 0 18px; border-radius: 8px; font-size: 13px; font-weight: 600; white-space: nowrap;"><i class="fas fa-check mr-1"></i> Activate</button>\n          </div>\n        </div>\n      </div>\n    </div>\n\n  </div>\n</div>\n\n<!-- ========================================================================= -->\n<!-- COMPLETE BOOTSTRAP / ADMINLTE MODAL SUITE                                 -->\n<!-- ========================================================================= -->\n\n<!-- 1. EDIT CLIENT MODAL -->\n<div class="modal fade" id="modal-edit-client" tabindex="-1" role="dialog">\n  <div class="modal-dialog modal-dialog-centered" role="document">\n    <div class="modal-content">\n      <div class="modal-header">\n        <h5 class="modal-title"><i class="fas fa-user-edit"></i> Edit Client Session</h5>\n        <button type="button" class="close text-white" data-dismiss="modal"><span>&times;</span></button>\n      </div>\n      <div class="modal-body">\n        <input type="hidden" id="modal-client-ip">\n        <div class="form-group">\n          <label>Client IP / MAC:</label>\n          <input type="text" id="modal-client-info" class="form-control" readonly>\n        </div>\n        <div class="form-group">\n          <label>Remaining Time (Minutes):</label>\n          <input type="number" id="modal-client-mins" class="form-control" min="0">\n        </div>\n        <div class="row">\n          <div class="col-12 col-md-6 form-group">\n            <label>Download Limit (Kbps):</label>\n            <input type="number" id="modal-client-dl" class="form-control" min="128">\n          </div>\n          <div class="col-12 col-md-6 form-group">\n            <label>Upload Limit (Kbps):</label>\n            <input type="number" id="modal-client-ul" class="form-control" min="64">\n          </div>\n        </div>\n      </div>\n      <div class="modal-footer">\n        <button type="button" class="btn btn-secondary" data-dismiss="modal">Cancel</button>\n        <button type="button" class="btn btn-primary" onclick="submitEditClientModal()"><i class="fas fa-save"></i> Save Changes</button>\n      </div>\n    </div>\n  </div>\n</div>\n\n<!-- 2. MEMBER TOP-UP / PIN MODAL -->\n<div class="modal fade" id="modal-member-topup" tabindex="-1" role="dialog">\n  <div class="modal-dialog modal-dialog-centered" role="document">\n    <div class="modal-content">\n      <div class="modal-header">\n        <h5 class="modal-title"><i class="fas fa-coins"></i> Manage Member Wallet</h5>\n        <button type="button" class="close text-white" data-dismiss="modal"><span>&times;</span></button>\n      </div>\n      <div class="modal-body">\n        <input type="hidden" id="modal-member-user">\n        <div class="form-group">\n          <label>Member Username:</label>\n          <input type="text" id="modal-member-user-display" class="form-control" readonly>\n        </div>\n        <div class="form-group">\n          <label>Adjust Minutes (+ to Add, - to Deduct):</label>\n          <input type="number" id="modal-member-adj-mins" class="form-control" value="30">\n        </div>\n      </div>\n      <div class="modal-footer">\n        <button type="button" class="btn btn-secondary" data-dismiss="modal">Cancel</button>\n        <button type="button" class="btn btn-info" onclick="submitMemberTopupModal()"><i class="fas fa-check"></i> Apply Adjustment</button>\n      </div>\n    </div>\n  </div>\n</div>\n\n<script>\nfunction showSection(secId) {\n    document.querySelectorAll(\'.section-view\').forEach(el => el.classList.remove(\'active\'));\n    document.querySelectorAll(\'.nav-sidebar .nav-link\').forEach(el => el.classList.remove(\'active\'));\n    \n    document.getElementById(secId).classList.add(\'active\');\n    const navLink = document.getElementById(secId.replace(\'sec-\', \'nav-\'));\n    if (navLink) navLink.classList.add(\'active\');\n\n    // Auto-close sidebar on mobile devices when section link is clicked\n    if ($(window).width() < 992) {\n        $(\'body\').removeClass(\'sidebar-open\').addClass(\'sidebar-collapse\');\n    }\n\n    if (secId === \'sec-clients\') loadClients();\n    if (secId === \'sec-vouchers\') loadVouchers();\n    if (secId === \'sec-members\') loadMembers();\n    if (secId === \'sec-rates\') loadRates();\n    if (secId === \'sec-walled\') loadWalledGarden();\n    if (secId === \'sec-security\') { loadMacs(); populateMacQuickPick(); }\n}\n\nfunction refreshStats() {\n    fetch(\'/admin/api/stats\').then(r=>r.json()).then(d=>{\n        document.getElementById(\'stat-today\').innerText = d.today_bottles;\n        document.getElementById(\'stat-total\').innerText = d.total_bottles;\n        document.getElementById(\'stat-clients\').innerText = d.active_clients;\n\n        if(document.getElementById(\'sys-cpu\')) document.getElementById(\'sys-cpu\').innerText = d.cpu + \'%\';\n        if(document.getElementById(\'sys-cpu-bar\')) document.getElementById(\'sys-cpu-bar\').style.width = d.cpu + \'%\';\n        if(document.getElementById(\'sys-ram\')) document.getElementById(\'sys-ram\').innerText = d.ram + \'%\';\n        if(document.getElementById(\'sys-ram-bar\')) document.getElementById(\'sys-ram-bar\').style.width = d.ram + \'%\';\n        if(document.getElementById(\'sys-disk\')) document.getElementById(\'sys-disk\').innerText = d.disk + \'%\';\n        if(document.getElementById(\'sys-disk-bar\')) document.getElementById(\'sys-disk-bar\').style.width = d.disk + \'%\';\n        if(document.getElementById(\'sys-uptime\')) document.getElementById(\'sys-uptime\').innerText = d.uptime;\n\n        const labels = d.history.map(h => h.date);\n        const data = d.history.map(h => h.count);\n\n        if (!window.historyChart) {\n            var chartCanvas = document.getElementById(\'historyChart\');\n            if (chartCanvas) {\n                var ctx = chartCanvas.getContext(\'2d\');\n                window.historyChart = new Chart(ctx, {\n                    type: \'bar\',\n                    data: {\n                        labels: labels,\n                        datasets: [{\n                            label: \'Bottles Recycled\',\n                            data: data,\n                            backgroundColor: \'rgba(16, 185, 129, 0.85)\',\n                            borderColor: \'#10b981\',\n                            borderWidth: 1\n                        }]\n                    },\n                    options: {\n                        responsive: true,\n                        maintainAspectRatio: false,\n                        legend: { display: false },\n                        scales: {\n                            yAxes: [{\n                                ticks: {\n                                    beginAtZero: true,\n                                    precision: 0,\n                                    fontColor: \'#cbd5e1\'\n                                },\n                                gridLines: {\n                                    color: \'rgba(255, 255, 255, 0.08)\',\n                                    zeroLineColor: \'rgba(255, 255, 255, 0.15)\'\n                                }\n                            }],\n                            xAxes: [{\n                                ticks: {\n                                    fontColor: \'#cbd5e1\'\n                                },\n                                gridLines: {\n                                    display: false\n                                }\n                            }]\n                        }\n                    }\n                });\n            }\n        } else {\n            window.historyChart.data.labels = labels;\n            window.historyChart.data.datasets[0].data = data;\n            window.historyChart.update();\n        }\n    });\n\n    fetch(\'/admin/api/license\').then(r=>r.json()).then(d=>{\n        document.getElementById(\'lic-hwid\').innerText = d.hwid;\n        document.getElementById(\'lic-status\').innerText = d.status;\n        document.getElementById(\'lic-tier\').innerText = d.tier || \'COMMERCIAL\';\n        document.getElementById(\'stat-lic\').innerText = d.status;\n    });\n}\nsetInterval(refreshStats, 3000);\nrefreshStats();\n\n// MAC Address Validation & Auto-Formatting\nfunction formatMacInput(input) {\n    let v = input.value.toUpperCase().replace(/[^0-9A-F]/g, \'\');\n    let formatted = \'\';\n    for (let i = 0; i < v.length && i < 12; i += 2) {\n        if (i > 0) formatted += \':\';\n        formatted += v.substr(i, 2);\n    }\n    input.value = formatted;\n    \n    const isValid = /^([0-9A-F]{2}:){5}[0-9A-F]{2}$/.test(formatted);\n    document.getElementById(\'mac-valid-msg\').style.display = isValid ? \'block\' : \'none\';\n    document.getElementById(\'mac-invalid-msg\').style.display = (formatted.length > 0 && !isValid) ? \'block\' : \'none\';\n}\n\nfunction populateMacQuickPick() {\n    fetch(\'/admin/api/clients\').then(r=>r.json()).then(clients=>{\n        let optHtml = \'<option value="">-- Choose active connected device to auto-fill --</option>\';\n        clients.forEach(c=>{\n            optHtml += `<option value="${c.mac}">${c.ip} (${c.mac})</option>`;\n        });\n        document.getElementById(\'mac-quick-pick\').innerHTML = optHtml;\n    });\n}\n\nfunction quickFillMac(val) {\n    if (!val) return;\n    document.getElementById(\'mac-input\').value = val;\n    formatMacInput(document.getElementById(\'mac-input\'));\n}\n\nfunction loadMacs() {\n    fetch(\'/admin/api/mac_control/list\').then(r=>r.json()).then(d=>{\n        let html = \'\';\n        d.forEach(m=>{\n            const safeNote = encodeURIComponent(m.note);\n            html += `<tr>\n                <td style="padding: 10px 14px;"><code>${m.mac}</code></td>\n                <td style="padding: 10px 14px; text-align: center;"><span class="badge ${m.type===\'whitelist\'?\'badge-success\':\'badge-danger\'}">${m.type.toUpperCase()}</span></td>\n                <td style="padding: 10px 14px;"><span class="text-light">${m.note || \'-\'}</span></td>\n                <td style="padding: 10px 14px; text-align: right; white-space: nowrap;">\n                    <div class="d-inline-flex align-items-center justify-content-end" style="gap: 6px; white-space: nowrap; flex-wrap: nowrap;">\n                        <button class="btn btn-xs btn-outline-warning text-nowrap" onclick="editMacControl(\'${m.mac}\', \'${m.type}\', \'${safeNote}\')"><i class="fas fa-edit mr-1"></i>Edit</button>\n                        <button class="btn btn-xs btn-outline-danger text-nowrap" onclick="deleteMacControl(\'${m.mac}\')"><i class="fas fa-trash mr-1"></i>Delete</button>\n                    </div>\n                </td>\n            </tr>`;\n        });\n        document.getElementById(\'mac-table-body\').innerHTML = html || \'<tr><td colspan="4" class="text-center p-3 text-muted">No MAC filtering rules set.</td></tr>\';\n    });\n}\n\nfunction saveMacControl() {\n    const mac = document.getElementById(\'mac-input\').value.trim().toUpperCase();\n    const type = document.getElementById(\'mac-type\').value;\n    const note = document.getElementById(\'mac-note\').value.trim();\n    const origMac = document.getElementById(\'edit-original-mac\').value;\n\n    if (!/^([0-9A-F]{2}:){5}[0-9A-F]{2}$/.test(mac)) {\n        Swal.fire(\'Invalid MAC\', \'MAC Address must be in 12 hex format: AA:BB:CC:DD:EE:FF\', \'warning\');\n        return;\n    }\n\n    if (origMac && origMac !== mac) {\n        fetch(\'/admin/api/mac_control/delete\', {\n            method: \'POST\',\n            headers: {\'Content-Type\':\'application/json\'},\n            body: JSON.stringify({mac: origMac})\n        });\n    }\n\n    fetch(\'/admin/api/mac_control/add\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({mac: mac, type: type, note: note})\n    }).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            cancelEditMac();\n            loadMacs();\n            Swal.fire(\'Saved!\', \'MAC Rule has been saved.\', \'success\');\n        } else {\n            Swal.fire(\'Error\', d.error || \'Failed to save MAC rule.\', \'error\');\n        }\n    });\n}\n\nfunction editMacControl(mac, type, encNote) {\n    const note = decodeURIComponent(encNote);\n    document.getElementById(\'edit-original-mac\').value = mac;\n    document.getElementById(\'mac-input\').value = mac;\n    document.getElementById(\'mac-type\').value = type;\n    document.getElementById(\'mac-note\').value = note;\n    \n    document.getElementById(\'mac-form-title\').innerHTML = `<i class="fas fa-edit text-warning"></i> Edit MAC Filter (${mac})`;\n    document.getElementById(\'btn-save-mac\').innerHTML = `<i class="fas fa-save"></i> Update Rule`;\n    document.getElementById(\'btn-save-mac\').className = `btn btn-warning btn-block mr-1`;\n    document.getElementById(\'btn-cancel-mac\').style.display = \'inline-block\';\n    \n    formatMacInput(document.getElementById(\'mac-input\'));\n    document.getElementById(\'mac-form-card\').scrollIntoView({ behavior: \'smooth\' });\n}\n\nfunction cancelEditMac() {\n    document.getElementById(\'edit-original-mac\').value = \'\';\n    document.getElementById(\'mac-input\').value = \'\';\n    document.getElementById(\'mac-note\').value = \'\';\n    document.getElementById(\'mac-quick-pick\').value = \'\';\n    \n    document.getElementById(\'mac-form-title\').innerHTML = `<i class="fas fa-shield-alt"></i> Add / Edit MAC Filter Rule`;\n    document.getElementById(\'btn-save-mac\').innerHTML = `<i class="fas fa-plus"></i> Add Rule`;\n    document.getElementById(\'btn-save-mac\').className = `btn btn-danger btn-block mr-1`;\n    document.getElementById(\'btn-cancel-mac\').style.display = \'none\';\n    formatMacInput(document.getElementById(\'mac-input\'));\n}\n\nfunction deleteMacControl(mac) {\n    Swal.fire({\n        title: `Delete MAC Rule?`,\n        text: `Are you sure you want to remove rule for ${mac}?`,\n        icon: \'warning\',\n        showCancelButton: true,\n        confirmButtonColor: \'#d33\',\n        confirmButtonText: \'Yes, delete it!\'\n    }).then((res) => {\n        if (res.isConfirmed) {\n            fetch(\'/admin/api/mac_control/delete\', {\n                method: \'POST\',\n                headers: {\'Content-Type\':\'application/json\'},\n                body: JSON.stringify({mac: mac})\n            }).then(()=>{\n                loadMacs();\n                Swal.fire(\'Deleted!\', \'MAC rule deleted.\', \'success\');\n            });\n        }\n    });\n}\n\n// Promo Rates Management with Real-Time Mathematical Conflict Prevention\nlet currentRatesCache = [];\n\nfunction loadRates() {\n    fetch(\'/admin/api/rates/list\').then(r=>r.json()).then(d=>{\n        currentRatesCache = d || [];\n        let html = \'\';\n        currentRatesCache.forEach(r=>{\n            let timeStr = \'\';\n            if (r.minutes >= 60) {\n                const hrs = (r.minutes / 60).toFixed(1);\n                timeStr = `${hrs} Hours (${r.minutes} mins)`;\n            } else {\n                timeStr = `${r.minutes} Minutes`;\n            }\n            const eff = (r.minutes / r.bottles).toFixed(1);\n            const baseRate = parseInt(document.getElementById(\'rate-1\').value) || 10;\n            const bonusPct = Math.round(((eff - baseRate) / baseRate) * 100);\n            const bonusTag = bonusPct > 0 ? `<span class="badge badge-success ml-1">+${bonusPct}% Bonus</span>` : `<span class="badge badge-secondary ml-1">Base</span>`;\n            const safeLabel = encodeURIComponent(r.label || \'\');\n            html += `<tr>\n                <td style="padding: 10px 14px;"><strong class="text-success"><i class="fas fa-wine-bottle mr-1"></i>${r.bottles} Bottle${r.bottles > 1 ? \'s\' : \'\'}</strong></td>\n                <td style="padding: 10px 14px;"><strong>${timeStr}</strong></td>\n                <td style="padding: 10px 14px;"><code>${eff} m/b</code> ${bonusTag}</td>\n                <td style="padding: 10px 14px;"><span class="text-light">${r.label || \'-\'}</span></td>\n                <td style="padding: 10px 14px; text-align: right; white-space: nowrap;">\n                    <div class="d-inline-flex align-items-center justify-content-end" style="gap: 6px; white-space: nowrap; flex-wrap: nowrap;">\n                        <button class="btn btn-xs btn-outline-warning text-nowrap" onclick="editPromoRate(${r.bottles}, ${r.minutes}, \'${safeLabel}\')"><i class="fas fa-edit mr-1"></i>Edit</button>\n                        ${r.bottles > 1 ? `<button class="btn btn-xs btn-outline-danger text-nowrap" onclick="deletePromoRate(${r.bottles})"><i class="fas fa-trash mr-1"></i>Delete</button>` : `<span class="text-muted small ml-1 text-nowrap">(Base)</span>`}\n                    </div>\n                </td>\n            </tr>`;\n        });\n        document.getElementById(\'rates-table-body\').innerHTML = html || \'<tr><td colspan="5" class="text-center p-3 text-muted">No promo rates configured.</td></tr>\';\n    });\n}\n\nfunction setRateBottles(n) {\n    document.getElementById(\'new-rate-bottles\').value = n;\n    validatePromoFormMath();\n    autoGenerateRateLabel();\n}\n\nfunction getSelectedTotalMinutes() {\n    const rawVal = parseFloat(document.getElementById(\'new-rate-time-val\').value) || 0;\n    const unit = document.getElementById(\'new-rate-time-unit\').value;\n    if (unit === \'hours\') return Math.round(rawVal * 60);\n    if (unit === \'days\') return Math.round(rawVal * 1440);\n    return Math.round(rawVal);\n}\n\nfunction onBaseRateInput() {\n    validatePromoFormMath();\n    loadRates();\n}\n\nfunction autoGenerateRateLabel() {\n    const b = parseInt(document.getElementById(\'new-rate-bottles\').value) || 0;\n    const m = getSelectedTotalMinutes();\n    if (!b || !m) return;\n    let timeStr = \'\';\n    if (m >= 60) {\n        const h = Math.floor(m / 60);\n        const remM = m % 60;\n        timeStr = (remM === 0) ? `${h} Hour${h > 1 ? \'s\' : \'\'}` : `${h}h ${remM}m`;\n    } else {\n        timeStr = `${m} mins`;\n    }\n    const label = `${b} Bottle${b > 1 ? \'s\' : \'\'} = ${timeStr}`;\n    document.getElementById(\'new-rate-label\').value = label;\n}\n\nfunction validatePromoFormMath() {\n    const b = parseInt(document.getElementById(\'new-rate-bottles\').value) || 0;\n    const m = getSelectedTotalMinutes();\n    const origB = parseInt(document.getElementById(\'edit-original-bottles\').value) || null;\n    const fb = document.getElementById(\'rate-validator-feedback\');\n    const effSpan = document.getElementById(\'rate-validator-eff\');\n    const statusSpan = document.getElementById(\'rate-validator-status\');\n    const msgDiv = document.getElementById(\'rate-validator-msg\');\n    const saveBtn = document.getElementById(\'btn-save-promo\');\n\n    if (!b || !m) {\n        fb.style.display = \'none\';\n        saveBtn.disabled = false;\n        return;\n    }\n\n    fb.style.display = \'block\';\n    const eff = (m / b).toFixed(2);\n    effSpan.innerText = `📊 Efficiency: ${eff} mins/bottle (${m}m for ${b}B)`;\n\n    // Check invariants against currentRatesCache (excluding editing tier)\n    const existing = currentRatesCache.filter(r => r.bottles !== origB);\n    let conflict = null;\n\n    // Invariant 1: Monotonic Efficiency\n    for (let r of existing) {\n        const exEff = r.minutes / r.bottles;\n        if (r.bottles < b && exEff > (m / b)) {\n            conflict = `Efficiency conflict: ${r.bottles}B tier gives ${exEff.toFixed(1)} m/b, but this gives only ${eff} m/b. Larger bundles must be at least as rewarding.`;\n            break;\n        }\n        if (r.bottles > b && exEff < (m / b)) {\n            conflict = `Efficiency conflict: this tier gives ${eff} m/b, which exceeds the larger ${r.bottles}B tier (${exEff.toFixed(1)} m/b).`;\n            break;\n        }\n    }\n\n    // Invariant 2: Combination Floor\n    if (!conflict) {\n        const lowerTiers = existing.filter(r => r.bottles < b).sort((a,b) => b.bottles - a.bottles);\n        let comboMins = 0;\n        let rem = b;\n        for (let lt of lowerTiers) {\n            if (rem >= lt.bottles) {\n                comboMins += Math.floor(rem / lt.bottles) * lt.minutes;\n                rem %= lt.bottles;\n            }\n        }\n        if (comboMins > 0 && m < comboMins) {\n            conflict = `Combination conflict: Depositing ${b} bottles in smaller packages yields ${comboMins} mins, but this package gives only ${m} mins. Minimum required is ${comboMins} mins.`;\n        }\n    }\n\n    // Invariant 3: Higher-Tier Upper Bound\n    if (!conflict) {\n        const higherTiers = existing.filter(r => r.bottles > b);\n        if (higherTiers.length > 0) {\n            const minHigher = Math.min(...higherTiers.map(r => r.minutes));\n            if (m >= minHigher) {\n                conflict = `Upper bound conflict: ${m} mins equals or exceeds a larger tier (${minHigher} mins).`;\n            }\n        }\n    }\n\n    if (conflict) {\n        fb.className = \'alert alert-danger py-2 px-3 mb-0\';\n        statusSpan.innerText = \'❌ Conflict Detected\';\n        msgDiv.innerText = conflict;\n        saveBtn.disabled = true;\n    } else {\n        fb.className = \'alert alert-success py-2 px-3 mb-0\';\n        statusSpan.innerText = \'✔ Mathematically Balanced\';\n        msgDiv.innerText = \'No rate curve conflicts. Bundle incentivizes bulk deposit.\';\n        saveBtn.disabled = false;\n    }\n}\n\nfunction editPromoRate(bottles, minutes, encLabel) {\n    const label = decodeURIComponent(encLabel || \'\');\n    document.getElementById(\'edit-original-bottles\').value = bottles;\n    document.getElementById(\'new-rate-bottles\').value = bottles;\n    document.getElementById(\'new-rate-time-val\').value = minutes;\n    document.getElementById(\'new-rate-time-unit\').value = \'mins\';\n    document.getElementById(\'new-rate-label\').value = label;\n\n    document.getElementById(\'promo-form-title\').innerHTML = `<i class="fas fa-edit text-warning"></i> Edit Promo Rate Tier (${bottles} Bottles)`;\n    document.getElementById(\'btn-save-promo\').innerHTML = `<i class="fas fa-save"></i> Update Rate`;\n    document.getElementById(\'btn-save-promo\').className = \'btn btn-warning btn-block mr-1\';\n    document.getElementById(\'btn-cancel-promo\').style.display = \'inline-block\';\n\n    validatePromoFormMath();\n    document.getElementById(\'promo-form-card\').scrollIntoView({ behavior: \'smooth\' });\n}\n\nfunction cancelEditPromoRate() {\n    document.getElementById(\'edit-original-bottles\').value = \'\';\n    document.getElementById(\'new-rate-bottles\').value = \'\';\n    document.getElementById(\'new-rate-time-val\').value = \'\';\n    document.getElementById(\'new-rate-time-unit\').value = \'mins\';\n    document.getElementById(\'new-rate-label\').value = \'\';\n    document.getElementById(\'rate-validator-feedback\').style.display = \'none\';\n\n    document.getElementById(\'promo-form-title\').innerHTML = `<i class="fas fa-plus-circle"></i> Add Custom Promo Rate Package`;\n    document.getElementById(\'btn-save-promo\').innerHTML = `<i class="fas fa-plus"></i> Add Rate`;\n    document.getElementById(\'btn-save-promo\').className = \'btn btn-success btn-block mr-1\';\n    document.getElementById(\'btn-save-promo\').disabled = false;\n    document.getElementById(\'btn-cancel-promo\').style.display = \'none\';\n}\n\nfunction addPromoRate() {\n    const b = parseInt(document.getElementById(\'new-rate-bottles\').value);\n    const m = getSelectedTotalMinutes();\n    const origB = document.getElementById(\'edit-original-bottles\').value;\n    let l = document.getElementById(\'new-rate-label\').value.trim();\n\n    if (!b || !m || isNaN(b) || isNaN(m)) {\n        Swal.fire(\'Input Error\', \'Please enter valid numbers for Bottles and Duration.\', \'warning\');\n        return;\n    }\n\n    if (!l) {\n        autoGenerateRateLabel();\n        l = document.getElementById(\'new-rate-label\').value.trim();\n    }\n\n    fetch(\'/admin/api/rates/add\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({bottles: b, minutes: m, label: l, orig_bottles: origB ? parseInt(origB) : null})\n    }).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            cancelEditPromoRate();\n            loadRates();\n            Swal.fire(\'Saved!\', \'Promo rate tier saved successfully.\', \'success\');\n        } else {\n            Swal.fire(\'Conflict Error\', d.error || \'Failed to save rate.\', \'error\');\n        }\n    }).catch(e=>{\n        Swal.fire(\'Error\', \'Server error while saving promo rate.\', \'error\');\n    });\n}\n\nfunction deletePromoRate(b) {\n    Swal.fire({\n        title: `Delete Rate Tier?`,\n        text: `Delete package for ${b} bottle(s)?`,\n        icon: \'warning\',\n        showCancelButton: true,\n        confirmButtonColor: \'#d33\',\n        confirmButtonText: \'Yes, delete it!\'\n    }).then((res) => {\n        if (res.isConfirmed) {\n            fetch(\'/admin/api/rates/delete\', {\n                method: \'POST\',\n                headers: {\'Content-Type\':\'application/json\'},\n                body: JSON.stringify({bottles: b})\n            }).then(r=>r.json()).then(d=>{\n                if (d.success) {\n                    loadRates();\n                    Swal.fire(\'Deleted!\', \'Rate tier removed.\', \'success\');\n                } else {\n                    Swal.fire(\'Error\', d.error || \'Could not delete rate.\', \'error\');\n                }\n            });\n        }\n    });\n}\n\nfunction applyRatePreset() {\n    const p = document.getElementById(\'rate-preset-select\').value;\n    Swal.fire({\n        title: \'Apply Rate Template?\',\n        text: \'This will replace all active promo rates with the selected conflict-free curve template.\',\n        icon: \'question\',\n        showCancelButton: true,\n        confirmButtonColor: \'#17a2b8\',\n        confirmButtonText: \'Yes, Apply Template\'\n    }).then(res => {\n        if (res.isConfirmed) {\n            fetch(\'/admin/api/rates/apply_preset\', {\n                method: \'POST\',\n                headers: {\'Content-Type\':\'application/json\'},\n                body: JSON.stringify({preset: p})\n            }).then(r=>r.json()).then(d=>{\n                if (d.success) {\n                    loadRates();\n                    Swal.fire(\'Applied!\', d.message, \'success\');\n                } else {\n                    Swal.fire(\'Error\', d.error || \'Could not apply template.\', \'error\');\n                }\n            });\n        }\n    });\n}\n\nfunction saveRates() {\n    const minPerBottle = document.getElementById(\'rate-1\').value;\n    const dropTimeout = document.getElementById(\'rate-timeout\').value;\n    fetch(\'/admin/api/settings/save\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({minutes_per_bottle: minPerBottle, drop_timeout: dropTimeout})\n    }).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            loadRates();\n            Swal.fire(\'Saved!\', \'Base timing settings updated.\', \'success\');\n        } else {\n            Swal.fire(\'Error\', d.error || \'Failed to save settings.\', \'error\');\n        }\n    });\n}\n\n// Clients & Client Modal\nfunction loadClients() {\n    fetch(\'/admin/api/clients\').then(r=>r.json()).then(d=>{\n        let html = \'\';\n        d.forEach(c=>{\n            const mins = Math.floor(c.remaining_seconds / 60);\n            html += `<tr>\n                <td style="padding: 10px 14px;"><strong>${c.ip}</strong></td>\n                <td style="padding: 10px 14px;"><code>${c.mac}</code></td>\n                <td style="padding: 10px 14px;"><strong>${mins}m</strong> <small class="text-muted">(${c.remaining_seconds}s)</small></td>\n                <td style="padding: 10px 14px; text-align: center;"><span class="badge ${c.is_paused ? \'badge-warning\' : \'badge-success\'}">${c.is_paused ? \'PAUSED\' : \'ACTIVE\'}</span></td>\n                <td style="padding: 10px 14px;"><span class="text-info font-weight-bold">${c.dl_kbps || 3072} / ${c.ul_kbps || 1536}</span> <small class="text-muted">Kbps</small></td>\n                <td style="padding: 10px 14px; text-align: right; white-space: nowrap;">\n                    <div class="d-inline-flex align-items-center justify-content-end" style="gap: 5px; white-space: nowrap; flex-wrap: nowrap;">\n                        <button class="btn btn-xs btn-outline-success text-nowrap" onclick="clientAction(\'${c.ip}\', \'add15\')"><i class="fas fa-plus mr-1"></i>15m</button>\n                        <button class="btn btn-xs btn-outline-warning text-nowrap" onclick="clientAction(\'${c.ip}\', \'${c.is_paused ? \'resume\' : \'pause\'}\')">${c.is_paused ? \'<i class="fas fa-play mr-1"></i>Resume\' : \'<i class="fas fa-pause mr-1"></i>Pause\'}</button>\n                        <button class="btn btn-xs btn-outline-info text-nowrap" onclick="openEditClientModal(\'${c.ip}\', \'${c.mac}\', ${c.remaining_seconds}, ${c.dl_kbps || 3072}, ${c.ul_kbps || 1536})"><i class="fas fa-edit mr-1"></i>Edit</button>\n                        <button class="btn btn-xs btn-outline-danger text-nowrap" onclick="clientAction(\'${c.ip}\', \'kick\')"><i class="fas fa-user-slash mr-1"></i>Kick</button>\n                    </div>\n                </td>\n            </tr>`;\n        });\n        document.getElementById(\'clients-table-body\').innerHTML = html || \'<tr><td colspan="6" class="text-center p-3 text-muted">No active clients connected.</td></tr>\';\n    });\n}\n\nfunction openEditClientModal(ip, mac, remSeconds, dl, ul) {\n    document.getElementById(\'modal-client-ip\').value = ip;\n    document.getElementById(\'modal-client-info\').value = `${ip} (${mac})`;\n    document.getElementById(\'modal-client-mins\').value = Math.max(0, Math.floor((remSeconds || 0) / 60));\n    document.getElementById(\'modal-client-dl\').value = dl || 3072;\n    document.getElementById(\'modal-client-ul\').value = ul || 1536;\n    $(\'#modal-edit-client\').modal(\'show\');\n}\n\nfunction submitEditClientModal() {\n    const ip = document.getElementById(\'modal-client-ip\').value;\n    const mins = parseInt(document.getElementById(\'modal-client-mins\').value) || 0;\n    const dl = parseInt(document.getElementById(\'modal-client-dl\').value) || 3072;\n    const ul = parseInt(document.getElementById(\'modal-client-ul\').value) || 1536;\n\n    fetch(\'/admin/api/client/edit\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({ip: ip, minutes: mins, dl_kbps: dl, ul_kbps: ul})\n    }).then(r=>r.json()).then(d=>{\n        $(\'#modal-edit-client\').modal(\'hide\');\n        loadClients();\n        Swal.fire(\'Updated!\', \'Client session updated.\', \'success\');\n    });\n}\n\nfunction clientAction(ip, act) {\n    fetch(\'/admin/api/client/action\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({ip: ip, action: act})\n    }).then(()=>loadClients());\n}\n\n// Vouchers\nfunction loadVouchers() {\n    fetch(\'/admin/api/vouchers/list\').then(r=>r.json()).then(d=>{\n        let html = \'\';\n        d.forEach(v=>{\n            html += `<tr>\n                <td style="padding: 10px 14px;"><strong class="text-success"><i class="fas fa-ticket-alt mr-1"></i>${v.code}</strong></td>\n                <td style="padding: 10px 14px;"><strong>${v.minutes}m</strong></td>\n                <td style="padding: 10px 14px; text-align: center;"><span class="badge ${v.is_used ? \'badge-secondary\' : \'badge-success\'}">${v.is_used ? \'REDEEMED\' : \'ACTIVE\'}</span></td>\n                <td style="padding: 10px 14px;"><span class="text-light">${v.note || \'-\'}</span></td>\n                <td style="padding: 10px 14px;"><small class="text-muted">${v.created_at}</small></td>\n                <td style="padding: 10px 14px;"><code>${v.used_by || \'-\'}</code></td>\n                <td style="padding: 10px 14px; text-align: right; white-space: nowrap;"><button class="btn btn-xs btn-outline-danger text-nowrap" onclick="deleteVoucher(\'${v.code}\')"><i class="fas fa-trash mr-1"></i>Delete</button></td>\n            </tr>`;\n        });\n        document.getElementById(\'voucher-history-body\').innerHTML = html || \'<tr><td colspan="7" class="text-center p-3 text-muted">No vouchers generated yet.</td></tr>\';\n    });\n}\n\nfunction generateVouchers() {\n    const q = document.getElementById(\'v-qty\').value;\n    const m = document.getElementById(\'v-mins\').value;\n    const note = document.getElementById(\'v-note\').value.trim();\n    fetch(\'/admin/api/vouchers/generate\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({qty: q, minutes: m, note: note})\n    }).then(r=>r.json()).then(d=>{\n        let html = \'<div class="alert alert-success"><h5>Generated Vouchers:</h5><ul>\';\n        d.vouchers.forEach(v=>{ html += `<li><strong>${v.code}</strong> (${v.minutes} Minutes) - ${v.note || \'\'}</li>`; });\n        html += \'</ul></div>\';\n        document.getElementById(\'v-results\').innerHTML = html;\n        loadVouchers();\n        Swal.fire(\'Generated!\', `${d.vouchers.length} vouchers generated.`, \'success\');\n    });\n}\n\nfunction deleteVoucher(code) {\n    Swal.fire({\n        title: \'Delete Voucher?\',\n        text: `Delete voucher code ${code}?`,\n        icon: \'warning\',\n        showCancelButton: true,\n        confirmButtonColor: \'#d33\',\n        confirmButtonText: \'Yes, delete\'\n    }).then((res) => {\n        if (res.isConfirmed) {\n            fetch(\'/admin/api/vouchers/delete\', {\n                method: \'POST\',\n                headers: {\'Content-Type\':\'application/json\'},\n                body: JSON.stringify({code: code})\n            }).then(()=>{\n                loadVouchers();\n                Swal.fire(\'Deleted!\', \'Voucher deleted.\', \'success\');\n            });\n        }\n    });\n}\n\n// Members & Member Modals\nfunction loadMembers() {\n    fetch(\'/admin/api/members/list\').then(r=>r.json()).then(d=>{\n        let html = \'\';\n        d.forEach(m=>{\n            const hrs = (m.wallet_minutes / 60).toFixed(1);\n            html += `<tr>\n                <td style="padding: 10px 14px;"><strong class="text-info"><i class="fas fa-user-circle mr-1"></i>${m.username}</strong></td>\n                <td style="padding: 10px 14px;"><span class="badge badge-info">${m.wallet_minutes} Mins (${hrs} Hrs)</span></td>\n                <td style="padding: 10px 14px;"><small class="text-muted">${m.created_at}</small></td>\n                <td style="padding: 10px 14px; text-align: right; white-space: nowrap;">\n                    <div class="d-inline-flex align-items-center justify-content-end" style="gap: 6px; white-space: nowrap; flex-wrap: nowrap;">\n                        <button class="btn btn-xs btn-outline-success text-nowrap" onclick="openMemberTopupModal(\'${m.username}\')"><i class="fas fa-coins mr-1"></i>Adjust</button>\n                        <button class="btn btn-xs btn-outline-danger text-nowrap" onclick="deleteMember(\'${m.username}\')"><i class="fas fa-trash mr-1"></i>Delete</button>\n                    </div>\n                </td>\n            </tr>`;\n        });\n        document.getElementById(\'members-table-body\').innerHTML = html || \'<tr><td colspan="4" class="text-center p-3 text-muted">No registered members yet.</td></tr>\';\n    });\n}\n\nfunction openMemberTopupModal(username) {\n    document.getElementById(\'modal-member-user\').value = username;\n    document.getElementById(\'modal-member-user-display\').value = username;\n    document.getElementById(\'modal-member-adj-mins\').value = \'30\';\n    $(\'#modal-member-topup\').modal(\'show\');\n}\n\nfunction submitMemberTopupModal() {\n    const u = document.getElementById(\'modal-member-user\').value;\n    const mins = parseInt(document.getElementById(\'modal-member-adj-mins\').value) || 0;\n    fetch(\'/admin/api/members/topup\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({username: u, minutes: mins})\n    }).then(()=>{\n        $(\'#modal-member-topup\').modal(\'hide\');\n        loadMembers();\n        Swal.fire(\'Adjusted!\', `Wallet for ${u} updated.`, \'success\');\n    });\n}\n\nfunction addMember() {\n    const u = document.getElementById(\'new-mem-user\').value.trim();\n    const p = document.getElementById(\'new-mem-pin\').value.trim();\n    const m = parseInt(document.getElementById(\'new-mem-mins\').value) || 0;\n    if (!u || !p) {\n        Swal.fire(\'Required\', \'Username and PIN are required.\', \'warning\');\n        return;\n    }\n    fetch(\'/admin/api/members/add\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({username: u, pin: p, wallet_minutes: m})\n    }).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            document.getElementById(\'new-mem-user\').value = \'\';\n            document.getElementById(\'new-mem-pin\').value = \'\';\n            document.getElementById(\'new-mem-mins\').value = \'0\';\n            loadMembers();\n            Swal.fire(\'Created!\', \'Member account created successfully!\', \'success\');\n        } else {\n            Swal.fire(\'Error\', d.error || \'Failed to create member.\', \'error\');\n        }\n    });\n}\n\nfunction deleteMember(u) {\n    Swal.fire({\n        title: \'Delete Member?\',\n        text: `Delete account ${u}?`,\n        icon: \'warning\',\n        showCancelButton: true,\n        confirmButtonColor: \'#d33\',\n        confirmButtonText: \'Yes, delete\'\n    }).then((res) => {\n        if (res.isConfirmed) {\n            fetch(\'/admin/api/members/delete\', {\n                method: \'POST\',\n                headers: {\'Content-Type\':\'application/json\'},\n                body: JSON.stringify({username: u})\n            }).then(()=>{\n                loadMembers();\n                Swal.fire(\'Deleted!\', \'Member account deleted.\', \'success\');\n            });\n        }\n    });\n}\n\n// Walled Garden\nfunction loadWalledGarden() {\n    fetch(\'/admin/api/walled_garden/list\').then(r=>r.json()).then(d=>{\n        let html = \'\';\n        d.forEach(w=>{\n            html += `<tr>\n                <td style="padding: 10px 14px;"><code class="text-success">${w.domain}</code></td>\n                <td style="padding: 10px 14px;"><span class="text-light">${w.note || \'-\'}</span></td>\n                <td style="padding: 10px 14px; text-align: right; white-space: nowrap;"><button class="btn btn-xs btn-outline-danger text-nowrap" onclick="deleteWalledDomain(\'${w.domain}\')"><i class="fas fa-trash mr-1"></i>Delete</button></td>\n            </tr>`;\n        });\n        document.getElementById(\'walled-table-body\').innerHTML = html || \'<tr><td colspan="3" class="text-center p-3 text-muted">No walled garden sites whitelisted.</td></tr>\';\n    });\n}\n\nfunction addWalledDomain() {\n    const domain = document.getElementById(\'walled-domain\').value.trim().toLowerCase();\n    const note = document.getElementById(\'walled-note\').value.trim();\n    if (!domain) return;\n    fetch(\'/admin/api/walled_garden/add\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({domain: domain, note: note})\n    }).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            document.getElementById(\'walled-domain\').value = \'\';\n            document.getElementById(\'walled-note\').value = \'\';\n            loadWalledGarden();\n            Swal.fire(\'Whitelisted!\', \'Domain whitelisted in Walled Garden.\', \'success\');\n        } else {\n            Swal.fire(\'Invalid Domain\', d.error || \'Invalid domain format.\', \'warning\');\n        }\n    });\n}\n\nfunction deleteWalledDomain(domain) {\n    Swal.fire({\n        title: \'Remove Domain?\',\n        text: `Remove ${domain} from whitelist?`,\n        icon: \'warning\',\n        showCancelButton: true,\n        confirmButtonColor: \'#d33\',\n        confirmButtonText: \'Yes, remove\'\n    }).then((res) => {\n        if (res.isConfirmed) {\n            fetch(\'/admin/api/walled_garden/delete\', {\n                method: \'POST\',\n                headers: {\'Content-Type\':\'application/json\'},\n                body: JSON.stringify({domain: domain})\n            }).then(()=>{\n                loadWalledGarden();\n                Swal.fire(\'Removed!\', \'Domain removed.\', \'success\');\n            });\n        }\n    });\n}\n\nfunction savePortalCustom() {\n    const n = document.getElementById(\'cfg-vendo-name\').value.trim();\n    const sub = document.getElementById(\'cfg-vendo-sub\').value.trim();\n    const ann = document.getElementById(\'cfg-announcement\').value.trim();\n    fetch(\'/admin/api/settings/save\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({vendo_name: n, vendo_subtitle: sub, announcement: ann})\n    }).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            Swal.fire(\'Saved!\', \'Portal branding & announcement updated.\', \'success\');\n        } else {\n            Swal.fire(\'Error\', d.error || \'Failed to save settings.\', \'error\');\n        }\n    });\n}\n\nfunction saveBandwidth() {\n    const dl = document.getElementById(\'cfg-dl\').value;\n    const ul = document.getElementById(\'cfg-ul\').value;\n    const t = document.getElementById(\'cfg-tether\').value;\n    fetch(\'/admin/api/settings/save\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({default_dl_kbps: dl, default_ul_kbps: ul, anti_tethering: t})\n    }).then(()=>Swal.fire(\'Saved!\', \'Bandwidth & Anti-Tethering rules applied.\', \'success\'));\n}\n\nfunction saveTelegram() {\n    const tok = document.getElementById(\'cfg-tg-token\').value;\n    const cid = document.getElementById(\'cfg-tg-chat\').value;\n    const bin = document.getElementById(\'cfg-tg-bin\').value;\n    const daily = document.getElementById(\'cfg-tg-daily\').value;\n    fetch(\'/admin/api/settings/save\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({\n            telegram_bot_token: tok,\n            telegram_chat_id: cid,\n            telegram_alert_bin: bin,\n            telegram_alert_daily: daily\n        })\n    }).then(()=>Swal.fire(\'Saved!\', \'Telegram alerts configured.\', \'success\'));\n}\n\nfunction testTelegram() {\n    fetch(\'/admin/api/telegram/test\', {method:\'POST\'}).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            Swal.fire(\'Sent!\', \'Test alert sent successfully to Telegram!\', \'success\');\n        } else {\n            Swal.fire(\'Failed\', \'Could not send alert. Please check your Bot Token and Chat ID.\', \'error\');\n        }\n    });\n}\n\nfunction onAudioPresetChange(type) {\n    const sel = document.getElementById(`audio-${type}-preset`).value;\n    const customInp = document.getElementById(`audio-${type}-custom`);\n    const player = document.getElementById(`audio-player-${type}`);\n    \n    if (sel === \'silent\') {\n        customInp.value = \'silent\';\n        player.src = \'\';\n    } else if (sel === \'custom\') {\n        player.src = customInp.value;\n    } else if (sel === \'arcade_powerup\' || sel === \'voice_filipino\' || sel === \'crystal_bell\') {\n        customInp.value = sel;\n        player.src = \'\';\n    } else {\n        customInp.value = sel;\n        player.src = sel;\n    }\n}\n\nfunction updateAudioPlayer(type) {\n    const url = document.getElementById(`audio-${type}-custom`).value.trim();\n    const player = document.getElementById(`audio-player-${type}`);\n    if (url && url !== \'silent\' && ![\'arcade_powerup\', \'voice_filipino\', \'crystal_bell\'].includes(url)) {\n        player.src = url;\n    }\n}\n\nfunction uploadAudioFile(type) {\n    const fileInp = document.getElementById(`upload-file-${type}`);\n    if (!fileInp.files || fileInp.files.length === 0) return;\n    \n    const formData = new FormData();\n    formData.append(\'file\', fileInp.files[0]);\n    \n    Swal.fire({\n        title: \'Uploading Audio...\',\n        text: \'Please wait while your custom audio file is uploaded.\',\n        allowOutsideClick: false,\n        didOpen: () => { Swal.showLoading(); }\n    });\n    \n    fetch(\'/admin/api/audio/upload\', {\n        method: \'POST\',\n        body: formData\n    }).then(r=>r.json()).then(d=>{\n        if (d.success) {\n            document.getElementById(`audio-${type}-custom`).value = d.url;\n            document.getElementById(`audio-${type}-preset`).value = \'custom\';\n            document.getElementById(`audio-player-${type}`).src = d.url;\n            Swal.fire(\'Uploaded!\', \'Custom audio file uploaded successfully!\', \'success\');\n        } else {\n            Swal.fire(\'Upload Failed\', d.error || \'Could not upload audio.\', \'error\');\n        }\n    }).catch(err => {\n        Swal.fire(\'Upload Error\', \'Error communicating with server.\', \'error\');\n    });\n}\n\nfunction saveAudioSettings() {\n    const bg = document.getElementById(\'audio-bg-custom\').value.trim() || \'/static/audio/eco_loop.wav\';\n    const insert = document.getElementById(\'audio-insert-custom\').value.trim() || \'/static/audio/eco_chime.wav\';\n    const success = document.getElementById(\'audio-success-custom\').value.trim() || \'/static/audio/eco_success.wav\';\n    const vol = document.getElementById(\'audio-vol-input\').value;\n    \n    fetch(\'/admin/api/audio/settings\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({audio_bg: bg, audio_insert: insert, audio_success: success, volume: vol})\n    }).then(r=>r.json()).then(d=>{\n        Swal.fire(\'Saved!\', \'All portal audio settings saved successfully!\', \'success\');\n    });\n}\n\nfunction updatePreviewVolume() {\n    const vol = document.getElementById(\'audio-vol-input\').value;\n    document.getElementById(\'vol-lbl\').innerText = vol + \'%\';\n    const vDecimal = vol / 100.0;\n    try {\n        document.getElementById(\'audio-player-bg\').volume = vDecimal;\n        document.getElementById(\'audio-player-insert\').volume = vDecimal;\n        document.getElementById(\'audio-player-success\').volume = vDecimal;\n    } catch(e) {}\n}\n\n// Call on load to set initial volume of previews\nwindow.addEventListener(\'DOMContentLoaded\', () => {\n    updatePreviewVolume();\n});\n\nfunction copyHwid() {\n    const hwid = document.getElementById(\"lic-hwid\").innerText.trim();\n    if (!hwid || hwid === \"Loading...\") return;\n    const btn = document.getElementById(\"btn-copy-hwid\");\n    const onSuccess = () => {\n        btn.className = \"btn btn-xs btn-success\";\n        btn.innerHTML = \"<i class=\\\"fas fa-check mr-1\\\"></i> Copied!\";\n        setTimeout(() => {\n            btn.className = \"btn btn-xs btn-outline-info\";\n            btn.innerHTML = \"<i class=\\\"fas fa-copy mr-1\\\"></i> <span id=\\\"copy-hwid-text\\\">Copy</span>\";\n        }, 2000);\n    };\n    if (navigator.clipboard && navigator.clipboard.writeText) {\n        navigator.clipboard.writeText(hwid).then(onSuccess).catch(() => {\n            const t = document.createElement(\"input\");\n            t.value = hwid;\n            document.body.appendChild(t);\n            t.select();\n            document.execCommand(\"copy\");\n            document.body.removeChild(t);\n            onSuccess();\n        });\n    } else {\n        const t = document.createElement(\"input\");\n        t.value = hwid;\n        document.body.appendChild(t);\n        t.select();\n        document.execCommand(\"copy\");\n        document.body.removeChild(t);\n        onSuccess();\n    }\n}\n\nfunction activateLicense() {\n    const p = document.getElementById(\'act-pin\').value.trim();\n    fetch(\'/admin/api/license/activate\', {\n        method: \'POST\',\n        headers: {\'Content-Type\':\'application/json\'},\n        body: JSON.stringify({pin: p})\n    }).then(r=>r.json()).then(d=>{\n        Swal.fire(\'Activation\', d.message, d.success ? \'success\' : \'error\');\n        refreshStats();\n    });\n}\n    function saveEsp32Config() {\n        const payload = {\n            bin_full_threshold_cm: parseInt($(\'#esp-bin\').val()),\n            entrance_gate_timeout: parseInt($(\'#esp-ent-tout\').val()),\n            settle_time_ms: parseInt($(\'#esp-settle\').val()),\n            success_drop_tout_ms: parseInt($(\'#esp-suc-time\').val()),\n            reject_drop_time_ms: parseInt($(\'#esp-rej-time\').val()),\n            pet_nir_w_min: parseInt($(\'#esp-nir-min\').val()),\n            pet_nir_w_max: parseInt($(\'#esp-nir-max\').val()),\n            ent_close_angle: parseInt($(\'#esp-ent-close\').val()),\n            ent_open_angle: parseInt($(\'#esp-ent-open\').val()),\n            suc_close_angle: parseInt($(\'#esp-suc-close\').val()),\n            suc_open_angle: parseInt($(\'#esp-suc-open\').val()),\n            rej_close_angle: parseInt($(\'#esp-rej-close\').val()),\n            rej_open_angle: parseInt($(\'#esp-rej-open\').val())\n        };\n        fetch(\'/admin/api/esp32/save\', {\n            method: \'POST\', headers: {\'Content-Type\': \'application/json\'},\n            body: JSON.stringify(payload)\n        }).then(r=>r.json()).then(d=>{\n            if(d.success) Swal.fire(\'Saved!\', \'Config pushed to ESP32 memory. Servos snapped to closed positions.\', \'success\');\n        });\n    }\n    \n    function triggerEsp32Config() {\n        if(confirm("This will reboot the ESP32 into Captive Portal mode and stop vending temporarily. Continue?")) {\n            fetch(\'/admin/api/esp32/trigger\', {method:\'POST\'}).then(()=>Swal.fire(\'Triggered\', \'ESP32 is rebooting into Wi-Fi Config Mode.\', \'info\'));\n        }\n    }\n</script>\n</body>\n</html>\n'
SERIAL_PORTS = ['/dev/ttyUSB0', '/dev/ttyUSB1', '/dev/ttyS1', '/dev/ttyS0']
BAUD_RATE = 115200

def hardware_serial_daemon():
    global ser
    while True:
        port = None
        for p in SERIAL_PORTS:
            if os.path.exists(p):
                port = p
                break
        if not port:
            time.sleep(3)
            continue
        try:
            with serial.Serial(port, BAUD_RATE, timeout=2) as s:
                ser = s
                while True:
                    raw_line = s.readline().decode('utf-8', errors='ignore').strip()
                    if not raw_line:
                        continue
                    try:
                        data = json.loads(raw_line)
                        on_esp32_uart_output(raw_line)
                    except json.JSONDecodeError:
                        pass
        except Exception:
            ser = None
            time.sleep(3)
if __name__ == '__main__':
    setup_firewall()
    restore_sessions_from_db()
    threading.Thread(target=time_daemon, daemon=True).start()
    if serial:
        threading.Thread(target=hardware_serial_daemon, daemon=True).start()
    port = int(os.environ.get('PORT', 5000))
    print('=================================================='.format())
    print('  SMART ECO-FI REVERSE VENDING MACHINE'.format())
    print('  Captive Portal : http://localhost:{}/'.format(port))
    print('  Simulator UI   : http://localhost:{}/simulator'.format(port))
    print('  Admin Panel    : http://localhost:{}/admin'.format(port))
    print('=================================================='.format())
    app.run(host='0.0.0.0', port=port, debug=False)
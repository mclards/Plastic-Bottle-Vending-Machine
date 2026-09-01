import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_econnect_bom_workbook(filename):
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # -------------------------------------------------------------
    # STYLES & COLOR PALETTE (ECO-Fi BRANDED PALETTE)
    # -------------------------------------------------------------
    forest_green = "1B4D3E"
    emerald_mid = "2E7D32"
    emerald_light = "388E3C"
    slate_gray = "4A5568"
    zebra_gray = "F4FBF7"
    accent_amber = "D97706"
    header_fill_gray = "2D3748"
    client_scope_fill = "FEF3C7"  # Soft amber for client scope

    font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_subhdr = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True, color="555555")
    font_grand_total = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_client_note = Font(name="Segoe UI", size=9, bold=True, color="92400E")

    fill_title = PatternFill(start_color=forest_green, end_color=forest_green, fill_type="solid")
    fill_section = PatternFill(start_color=emerald_mid, end_color=emerald_mid, fill_type="solid")
    fill_teal_section = PatternFill(start_color="0D5C75", end_color="0D5C75", fill_type="solid")
    fill_amber_section = PatternFill(start_color=accent_amber, end_color=accent_amber, fill_type="solid")
    fill_subhdr = PatternFill(start_color=header_fill_gray, end_color=header_fill_gray, fill_type="solid")
    fill_zebra = PatternFill(start_color=zebra_gray, end_color=zebra_gray, fill_type="solid")
    fill_grand_total = PatternFill(start_color=forest_green, end_color=forest_green, fill_type="solid")
    fill_green_highlight = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_client_scope = PatternFill(start_color=client_scope_fill, end_color=client_scope_fill, fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E0")
    double_border_bottom = Side(border_style="double", color="1A202C")
    thick_border_bottom = Side(border_style="medium", color="1A202C")

    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=thin_border_side, bottom=thick_border_bottom, left=thin_border_side, right=thin_border_side)
    border_grand_total = Border(top=thin_border_side, bottom=double_border_bottom, left=thin_border_side, right=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    curr_format = "₱#,##0.00"
    num_format = "#,##0"

    # =============================================================
    # SHEET 1: PROJECT SUMMARY & COSTING (BUILDER SCOPE @ ₱500/HR)
    # =============================================================
    ws_summary = wb.create_sheet(title="Project Summary & Costing")
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Block
    ws_summary.row_dimensions[1].height = 32
    ws_summary.merge_cells("A1:G1")
    ws_summary["A1"] = "ECO-Fi PLASTIC BOTTLE-TO-WIFI VENDO SYSTEM"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].fill = fill_title
    ws_summary["A1"].alignment = align_center

    ws_summary.row_dimensions[2].height = 22
    ws_summary.merge_cells("A2:G2")
    ws_summary["A2"] = "Turnkey Project Costing, Bill of Materials (BoM) & Professional Engineering Quotation"
    ws_summary["A2"].font = font_subtitle
    ws_summary["A2"].fill = PatternFill(start_color=emerald_mid, end_color=emerald_mid, fill_type="solid")
    ws_summary["A2"].alignment = align_center

    # Metadata Block
    metadata = [
        ("Project Standard:", "ECO-Fi Bottle-to-WiFi Reverse Vending System", "Builder Scope:", "Firmware, Electronics Assembly, AS7263 NIR, PCA9685, Chute & Portal"),
        ("Developer / Builder:", "Electrical Engineer & IoT Project Builder", "Professional Rate:", "₱500.00 / hour (Senior Electrical & IoT Engineering Rate)"),
        ("Date Prepared:", "August 2026", "Cabinet Scope:", "Client-Supplied / Excluded (Builder provides dimensional cutout guide)"),
        ("Sensing & Actuation:", "4\" PVC Drop Chute + AS7263 NIR Spectrometer + LJ12A3 + 2x Servos", "Core Gateway:", "Orange Pi One (Allwinner H3 512MB) + Armbian Server")
    ]

    r = 4
    for label1, val1, label2, val2 in metadata:
        ws_summary.row_dimensions[r].height = 20
        ws_summary.cell(row=r, column=1, value=label1).font = font_bold
        ws_summary.cell(row=r, column=1).alignment = align_left
        ws_summary.cell(row=r, column=2, value=val1).font = font_regular
        ws_summary.cell(row=r, column=2).alignment = align_left
        ws_summary.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

        ws_summary.cell(row=r, column=5, value=label2).font = font_bold
        ws_summary.cell(row=r, column=5).alignment = align_left
        ws_summary.cell(row=r, column=6, value=val2).font = font_regular
        ws_summary.cell(row=r, column=6).alignment = align_left
        ws_summary.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        r += 1

    # Section 1: Executive Cost Breakdown Table
    r += 1
    ws_summary.row_dimensions[r].height = 26
    ws_summary.merge_cells(f"A{r}:G{r}")
    ws_summary[f"A{r}"] = "1. EXECUTIVE COST BREAKDOWN (BUILDER SCOPE OF WORK)"
    ws_summary[f"A{r}"].font = font_section
    ws_summary[f"A{r}"].fill = fill_section
    ws_summary[f"A{r}"].alignment = align_left

    r += 1
    ws_summary.row_dimensions[r].height = 24
    summary_headers = ["Category No.", "Cost Category Description", "Key Scope / Components Included", "Scope Owner", "Amount (PHP)", "% Share", "Notes / References"]
    for c_idx, h_text in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=r, column=c_idx, value=h_text)
        cell.font = font_subhdr
        cell.fill = fill_subhdr
        cell.alignment = align_center
        cell.border = border_cell

    summary_rows = [
        ("Group A", "Computing, Storage & Networking", "Orange Pi One (512MB), ESP32 DevKit, MicroSD 32GB, TP-Link Outdoor AP", "Builder Scope", "='Bill of Materials (BoM)'!K11", "='Bill of Materials (BoM)'!K11/$E$18", "Proven Piso-WiFi Linux gateway & real-time sensor controller"),
        ("Group B", "Sensing, Chute & Actuation", "4\" PVC Drop Chute, AS7263 NIR Spectrometer, LJ12A3 Metal Sensor, PCA9685, 2x Servos, LCD", "Builder Scope", "='Bill of Materials (BoM)'!K21", "='Bill of Materials (BoM)'!K21/$E$18", "Rigid PVC drop pathway, AS7263 NIR PET signature, PCA9685 2-servo airlock & LCD"),
        ("Group C", "Power Supply, Regulation & Safety", "12V 5A (60W) SMPS, XL4015 5A Buck (5.1V logic rail), C14 Fused Inlet, 80mm Fan", "Builder Scope", "='Bill of Materials (BoM)'!K28", "='Bill of Materials (BoM)'!K28/$E$18", "Clean regulated power, master fuse switch & cooling"),
        ("Group D", "Cabinet, Enclosure & Woodwork", "Marine Plywood, Formica, Laser-cut Acrylic Hopper, Cam Locks, Hinges, Decals", "CLIENT SCOPE (EXCLUDED)", 0.00, "=0", "Client will build; Builder provides dimensions & fitment guide"),
        ("Group E", "Fasteners, Shield & Wiring Loom", "ESP32 screw terminal shield, silicone wiring, brass standoffs, heatshrink, Cat6 UTP", "Builder Scope", "='Bill of Materials (BoM)'!K42", "='Bill of Materials (BoM)'!K42/$E$18", "Vibration-proof terminal connections & cabling"),
        ("Labor", "Professional Engineering & Assembly Labor", "46 Hours (Schematics, AS7263 I2C, PCA9685, 2-Servo Airlock, Linux Gateway, Eco-Fi Portal, QA)", "Builder Scope", "='Labor Breakdown (WBS)'!F10", "='Labor Breakdown (WBS)'!F10/$E$18", "Licensed Electrical Engineer rate @ ₱500.00/hour"),
        ("Contingency", "Electronics & Sourcing Contingency", "Extra solder, test calibration bottles, spare connectors, shipping buffer (approx 5%)", "Builder Scope", 450.00, "=E17/$E$18", "Safety buffer for electronic components & shipping variance")
    ]

    start_summary_r = r + 1
    for row_data in summary_rows:
        r += 1
        ws_summary.row_dimensions[r].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if col_idx in [1, 4]:
                cell.alignment = align_center
            elif col_idx == 5:
                cell.alignment = align_right
                cell.number_format = curr_format
                cell.font = font_bold
            elif col_idx == 6:
                cell.alignment = align_right
                cell.number_format = "0.0%"
            else:
                cell.alignment = align_left
        
        # Highlight client scope row
        if "CLIENT SCOPE" in str(row_data[3]):
            for c in range(1, 8):
                ws_summary.cell(row=r, column=c).fill = fill_client_scope
            ws_summary.cell(row=r, column=4).font = font_client_note

    # Total Row
    r += 1
    total_r = r
    ws_summary.row_dimensions[r].height = 26
    ws_summary.cell(row=r, column=1, value="TOTAL").alignment = align_center
    ws_summary.cell(row=r, column=2, value="TOTAL CONTRACT COST (BUILDER SCOPE)").alignment = align_left
    ws_summary.cell(row=r, column=3, value="Electronics + AS7263 NIR Spectrometer + 2x Servos + PVC Chute + Firmware + Gateway + Portal + Labor").alignment = align_left
    ws_summary.cell(row=r, column=4, value="Turnkey Scope").alignment = align_center
    ws_summary.cell(row=r, column=5, value=f"=SUM(E{start_summary_r}:E{r-1})").alignment = align_right
    ws_summary.cell(row=r, column=5).number_format = curr_format
    ws_summary.cell(row=r, column=6, value=f"=SUM(F{start_summary_r}:F{r-1})").alignment = align_right
    ws_summary.cell(row=r, column=6).number_format = "0.0%"
    ws_summary.cell(row=r, column=7, value="Ready for integration into client-built cabinet").alignment = align_left

    for c in range(1, 8):
        cell = ws_summary.cell(row=r, column=c)
        cell.font = font_grand_total
        cell.fill = fill_grand_total
        cell.border = border_grand_total

    # Section 2: Contract Proposal Packages for Client
    r += 2
    ws_summary.row_dimensions[r].height = 26
    ws_summary.merge_cells(f"A{r}:G{r}")
    ws_summary[f"A{r}"] = "2. CONTRACT QUOTATION PACKAGES (FOR CLIENT PROPOSAL)"
    ws_summary[f"A{r}"].font = font_section
    ws_summary[f"A{r}"].fill = fill_teal_section
    ws_summary[f"A{r}"].alignment = align_left

    r += 1
    ws_summary.row_dimensions[r].height = 24
    pkg_headers = ["Option", "Contract Model", "Quoted Price (PHP)", "Builder Earnings", "Builder Deliverables & Support", "Client Responsibility", "Recommendation"]
    for c_idx, h_text in enumerate(pkg_headers, 1):
        cell = ws_summary.cell(row=r, column=c_idx, value=h_text)
        cell.font = font_subhdr
        cell.fill = fill_subhdr
        cell.alignment = align_center
        cell.border = border_cell

    pkg_rows = [
        ("Tier 1", "Direct Build (At-Cost Electronics + Labor @ ₱500/hr)", f"=E{total_r}", "₱23,000.00 (Labor Only)", "Tested electronics chassis, AS7263 NIR sensor, 2-servo airlock chute, Eco-Fi portal, wiring loom", "Builds cabinet, supplies enclosure & mounts electronics", "Best for transparent / direct contract"),
        ("Tier 2", "Standard Turnkey Contract (+15% Sourcing Buffer)", f"=ROUND(E{total_r}+('Bill of Materials (BoM)'!K43*0.15), -2)", "₱23,000.00 Labor + ₱1,274.00 Parts Buffer", "Pre-tested harness, AS7263 NIR sensor, 2x servos, PCA9685, PVC chute, dimensional drawings, 1-mo support", "Builds cabinet according to provided cutout drawing", "(Recommended) Standard Commercial Contract"),
        ("Tier 3", "Turnkey Package with On-site Setup & 3-Mo SLA", f"=ROUND(E{total_r}+5500, -2)", "₱23,000.00 Labor + ₱5,500.00 SLA & Setup Profit", "Full electronics package, AS7263 NIR, 2x servos, on-site setup assistance, staff training, 3-month SLA warranty", "Builds cabinet & handles physical installation", "Best for SK Councils, LGUs & Institutions")
    ]

    for p_data in pkg_rows:
        r += 1
        ws_summary.row_dimensions[r].height = 32
        for col_idx, val in enumerate(p_data, 1):
            cell = ws_summary.cell(row=r, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if col_idx in [1, 7]:
                cell.alignment = align_center
            elif col_idx == 3:
                cell.alignment = align_right
                cell.number_format = curr_format
                cell.font = font_bold
            elif col_idx in [5, 6]:
                cell.alignment = align_wrap_left
            else:
                cell.alignment = align_left
        if "Tier 2" in p_data[0]:
            for c in range(1, 8):
                ws_summary.cell(row=r, column=c).fill = fill_green_highlight

    # Auto-adjust column widths with generous margins
    summary_widths = [16, 44, 52, 26, 22, 14, 46]
    for idx, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(idx)].width = width

    # =============================================================
    # SHEET 2: BILL OF MATERIALS (BOM - WITH AS7263 NIR & 2 SERVOS)
    # =============================================================
    ws_bom = wb.create_sheet(title="Bill of Materials (BoM)")
    ws_bom.views.sheetView[0].showGridLines = True

    # Title Block
    ws_bom.row_dimensions[1].height = 32
    ws_bom.merge_cells("A1:K1")
    ws_bom["A1"] = "ECO-Fi VENDO - ITEMIZED BILL OF MATERIALS (BOM)"
    ws_bom["A1"].font = font_title
    ws_bom["A1"].fill = fill_title
    ws_bom["A1"].alignment = align_center

    ws_bom.row_dimensions[2].height = 22
    ws_bom.merge_cells("A2:K2")
    ws_bom["A2"] = "Hardware Architecture with AS7263 6-Channel NIR Spectrometer, 2x Servos, PCA9685 & PH Market Pricing"
    ws_bom["A2"].font = font_subtitle
    ws_bom["A2"].fill = PatternFill(start_color=emerald_mid, end_color=emerald_mid, fill_type="solid")
    ws_bom["A2"].alignment = align_center

    bom_headers = ["Item Code", "Category", "Component / Part Name", "Model / Spec", "Qty", "Unit", "Engineering Purpose & Function", "Base Price (₱)", "Shipping / Buffer (₱)", "Final Unit Price (₱)", "Subtotal (₱)"]

    r = 4
    ws_bom.row_dimensions[r].height = 24
    for c_idx, h_text in enumerate(bom_headers, 1):
        cell = ws_bom.cell(row=r, column=c_idx, value=h_text)
        cell.font = font_subhdr
        cell.fill = fill_subhdr
        cell.alignment = align_center
        cell.border = border_cell

    # Grouped Items Data for ECO-Fi (Including AS7263 NIR Spectrometer, 4" PVC Chute & PCA9685 Driver)
    bom_groups = [
        ("GROUP A: COMPUTING, STORAGE & NETWORKING (BUILDER SCOPE)", [
            ("A1", "Computing", "Orange Pi One (512MB RAM)", "Allwinner H3 Quad-Core Cortex-A7", 1, "pc", "Linux gateway, Eco-Fi Flask captive portal, SQLite, iptables", 1250.00, 80.00),
            ("A2", "Microcontroller", "ESP32 DevKit V1", "Dual-Core Xtensa LX6 240MHz (30/38 Pin)", 1, "pc", "Real-time optical drop pulse counting, I2C master & LCD driver", 220.00, 40.00),
            ("A3", "Storage", "MicroSD Card 32GB Class 10", "SanDisk Ultra A1 98MB/s High Endurance", 1, "pc", "Armbian OS root filesystem, session database, log storage", 320.00, 40.00),
            ("A4", "Networking", "Outdoor High-Power AP", "TP-Link CPE220 / EAP110-Outdoor 2.4GHz", 1, "unit", "Long-range Wi-Fi broadcasting for Eco-Fi hotspot clients", 2100.00, 120.00),
            ("A5", "Connectivity", "Short USB Data Cable", "USB-A to Type-C / Micro-B 0.3m shielded", 1, "pc", "Serial UART link between Orange Pi & ESP32 (115200 baud)", 80.00, 20.00),
        ]),
        ("GROUP B: SENSING, CHUTE & ACTUATION (BUILDER SCOPE)", [
            ("B1", "Drop Chute", "4\" (100mm) PVC Pipe & 45° Elbow", "Series 1000 PVC pipe (~1m) + 45° elbow fitting", 1, "set", "Rigid, smooth drop pathway for PET bottles; pre-drilled for sensors", 240.00, 40.00),
            ("B2", "Optical IR", "E18-D80NK Adjustable IR Sensor", "NPN NO 3-80cm Diffuse Optical Beam", 2, "pcs", "Chute entrance presence & drop confirmation beam counter (2ms)", 135.00, 30.00),
            ("B3", "Inductive", "LJ12A3-4-Z/BX Metal Sensor", "Inductive Proximity 4mm NPN NO (6-36V)", 1, "pc", "Mounted on PVC airlock chamber to detect & lockout tin/metal cans", 140.00, 30.00),
            ("B4", "Spectrometry", "AS7263 6-Channel NIR Spectrometer", "610-860nm Near-Infrared Optical Spectral Sensor (I2C)", 1, "module", "Directly measures NIR reflectance signature to confirm PET plastic material", 950.00, 50.00),
            ("B5", "Servo Driver", "PCA9685 16-Channel PWM Driver", "12-bit I2C PWM driver with power terminal & capacitor", 1, "module", "Dedicated hardware I2C servo controller; isolates servo power from ESP32", 110.00, 25.00),
            ("B6", "Actuators", "MG996R Metal Gear Servo", "High-Torque 13kg.cm Digital Servo", 2, "pcs", "2x Servos: 1x Outer Entrance Security Door + 1x Internal Drop Trapdoor", 210.00, 40.00),
            ("B7", "Display", "20x4 Character LCD + I2C", "HD44780 with PCF8574 I2C adapter backpack", 1, "pc", "Primary on-screen guide (Insert Bottle / Rate / Bottles Count)", 295.00, 35.00),
            ("B8", "Audio/Visual", "5V Active Buzzer & LED Kit", "5V Beeper + 5mm Green/Red Status LEDs", 1, "set", "Beep feedback on valid drop & visual status lights", 75.00, 25.00),
        ]),
        ("GROUP C: POWER DISTRIBUTION, REGULATION & SAFETY (BUILDER SCOPE)", [
            ("C1", "Main Power", "12V 5A (60W) Industrial SMPS", "Enclosed metal switching power supply", 1, "unit", "Main AC-DC power converter for entire Eco-Fi vendo", 320.00, 50.00),
            ("C2", "Buck Converter", "XL4015 5A Step-Down DC-DC", "Adjustable buck converter with heatsink", 1, "module", "Steps 12V down to 5.1V logic rail for Orange Pi, ESP32 & PCA9685 V+", 110.00, 25.00),
            ("C3", "AC Safety", "AC C14 Socket with Fuse & Switch", "IEC 320 C14 with 10A fuse & toggle switch", 1, "pc", "Main 220V power inlet with surge fuse and master switch", 120.00, 25.00),
            ("C4", "AC Cable", "Heavy-Duty 3-Prong AC Power Cord", "1.8m 3x0.75mm² grounded plug cord", 1, "pc", "Connects machine safely to building wall outlet", 110.00, 20.00),
            ("C5", "Cooling", "12V 80mm DC Exhaust Fan + Grill", "Brushless cooling fan with finger grill", 1, "set", "Continuous exhaust ventilation for Orange Pi and power supply", 130.00, 30.00),
        ]),
        ("GROUP D: CABINET, ENCLOSURE & FABRICATION [CLIENT SCOPE - REFERENCE ONLY]", [
            ("D1", "Structure", "Marine Plywood Cabinet + Formica", "1/2\" or 3/4\" marine plywood with laminate", 1, "unit", "[CLIENT-BUILT] Weather-resistant structural wooden housing", 2400.00, 200.00),
            ("D2", "Faceplate", "Laser-Cut Acrylic Bezel & Hopper", "3mm-5mm clear & tinted acrylic panels", 1, "set", "[CLIENT-BUILT] Display viewing window and intake hopper", 950.00, 100.00),
            ("D3", "Locks", "Heavy-Duty Cam Locks with Keys", "Tubular cam locks with master keys", 2, "sets", "[CLIENT-BUILT] Secures maintenance door and bottle collection bin", 120.00, 30.00),
            ("D4", "Hardware", "Stainless Steel Piano Hinges", "Continuous corrosion-proof door hinges", 1, "set", "[CLIENT-BUILT] Heavy-duty door hinge mounting", 180.00, 30.00),
            ("D5", "Branding", "Outdoor Vinyl Sticker Decal Wrap", "Laminated waterproof UV-resistant print", 1, "set", "[CLIENT-BUILT] Eco-Fi WiFi Vendo instructions & graphic branding", 450.00, 50.00),
        ]),
        ("GROUP E: FASTENERS, SHIELD & WIRING LOOM (BUILDER SCOPE)", [
            ("E1", "PCB Shield", "ESP32 Terminal Shield Breakout", "Screw terminal breakout PCB for ESP32", 1, "pc", "Vibration-proof screw connections for all sensor wires", 180.00, 30.00),
            ("E2", "Hardware", "M3/M4 Brass Standoffs, Screws & Nuts", "Stainless steel & brass fastener kit", 1, "box", "Rigid mounting for PCBs, buck converters, and sensors", 150.00, 30.00),
            ("E3", "Wiring", "22AWG/18AWG Silicone Wire & Lugs", "Flexible heat-resistant multi-strand wire", 1, "set", "Internal high-reliability power and signal wiring loom", 180.00, 25.00),
            ("E4", "Cable Org", "Heatshrink, Cable Ties & Spiral Wrap", "Polyolefin heatshrink & nylon cable ties", 1, "set", "Strain relief, insulation and clean cable management", 100.00, 20.00),
            ("E5", "Network Cable", "Cat6 Outdoor UTP Patch Cable", "5-meter weather-shielded RJ45 cable", 1, "pc", "PoE link from Orange Pi to outdoor access point", 120.00, 20.00),
        ])
    ]

    builder_scope_subtotal_cells = []
    client_scope_subtotal_cells = []

    for group_title, items in bom_groups:
        r += 1
        group_start_r = r + 1
        is_client_group = "CLIENT SCOPE" in group_title

        ws_bom.row_dimensions[r].height = 26
        ws_bom.merge_cells(f"A{r}:K{r}")
        ws_bom[f"A{r}"] = group_title
        ws_bom[f"A{r}"].font = font_section
        ws_bom[f"A{r}"].fill = fill_amber_section if is_client_group else fill_section
        ws_bom[f"A{r}"].alignment = align_left

        for item in items:
            r += 1
            ws_bom.row_dimensions[r].height = 24
            code, cat, name, model, qty, unit, desc, base, buffer = item
            ws_bom.cell(row=r, column=1, value=code).alignment = align_center
            ws_bom.cell(row=r, column=2, value=cat).alignment = align_left
            ws_bom.cell(row=r, column=3, value=name).alignment = align_left
            ws_bom.cell(row=r, column=4, value=model).alignment = align_left
            ws_bom.cell(row=r, column=5, value=qty).alignment = align_center
            ws_bom.cell(row=r, column=5).number_format = num_format
            ws_bom.cell(row=r, column=6, value=unit).alignment = align_center
            ws_bom.cell(row=r, column=7, value=desc).alignment = align_wrap_left
            ws_bom.cell(row=r, column=8, value=base).alignment = align_right
            ws_bom.cell(row=r, column=8).number_format = curr_format
            ws_bom.cell(row=r, column=9, value=buffer).alignment = align_right
            ws_bom.cell(row=r, column=9).number_format = curr_format
            ws_bom.cell(row=r, column=10, value=f"=H{r}+I{r}").alignment = align_right
            ws_bom.cell(row=r, column=10).number_format = curr_format
            ws_bom.cell(row=r, column=11, value=f"=E{r}*J{r}").alignment = align_right
            ws_bom.cell(row=r, column=11).number_format = curr_format
            ws_bom.cell(row=r, column=11).font = font_bold

            for c in range(1, 12):
                cell = ws_bom.cell(row=r, column=c)
                cell.font = font_bold if c in [1, 11] else font_regular
                cell.border = border_cell
                if is_client_group:
                    cell.fill = fill_client_scope

        # Subtotal Row for Group
        r += 1
        group_subtotal_r = r
        ws_bom.row_dimensions[r].height = 24
        if is_client_group:
            client_scope_subtotal_cells.append(f"K{group_subtotal_r}")
        else:
            builder_scope_subtotal_cells.append(f"K{group_subtotal_r}")

        ws_bom.merge_cells(f"A{r}:J{r}")
        ws_bom[f"A{r}"] = f"SUBTOTAL {group_title.split(':')[0]} {'[CLIENT ESTIMATE ONLY]' if is_client_group else '[BUILDER SCOPE]'}"
        ws_bom[f"A{r}"].alignment = align_right
        ws_bom[f"A{r}"].font = font_bold
        ws_bom.cell(row=r, column=11, value=f"=SUM(K{group_start_r}:K{r-1})").alignment = align_right
        ws_bom.cell(row=r, column=11).number_format = curr_format
        ws_bom.cell(row=r, column=11).font = font_bold

        for c in range(1, 12):
            cell = ws_bom.cell(row=r, column=c)
            cell.fill = fill_client_scope if is_client_group else fill_zebra
            cell.border = border_total

    # Grand Total Builder Scope Row
    r += 1
    grand_builder_bom_r = r
    ws_bom.row_dimensions[r].height = 26
    ws_bom.merge_cells(f"A{r}:J{r}")
    ws_bom[f"A{r}"] = "TOTAL BUILDER ELECTRONICS & HARDWARE SCOPE (GROUPS A, B, C, E)"
    ws_bom[f"A{r}"].alignment = align_right
    ws_bom[f"A{r}"].font = font_grand_total
    ws_bom.cell(row=r, column=11, value=f"={'+'.join(builder_scope_subtotal_cells)}").alignment = align_right
    ws_bom.cell(row=r, column=11).number_format = curr_format
    ws_bom.cell(row=r, column=11).font = font_grand_total

    for c in range(1, 12):
        cell = ws_bom.cell(row=r, column=c)
        cell.fill = fill_grand_total
        cell.border = border_grand_total

    # Auto-adjust column widths with extra clearance to prevent clipping
    bom_widths = [14, 18, 36, 44, 8, 10, 56, 16, 18, 18, 20]
    for idx, width in enumerate(bom_widths, 1):
        ws_bom.column_dimensions[get_column_letter(idx)].width = width

    # =============================================================
    # SHEET 3: LABOR BREAKDOWN (WBS - ₱500/HR PROFESSIONAL RATE)
    # =============================================================
    ws_labor = wb.create_sheet(title="Labor Breakdown (WBS)")
    ws_labor.views.sheetView[0].showGridLines = True

    # Title Block
    ws_labor.row_dimensions[1].height = 32
    ws_labor.merge_cells("A1:G1")
    ws_labor["A1"] = "ENGINEERING LABOR BREAKDOWN & WORK BREAKDOWN STRUCTURE (WBS)"
    ws_labor["A1"].font = font_title
    ws_labor["A1"].fill = fill_title
    ws_labor["A1"].alignment = align_center

    ws_labor.row_dimensions[2].height = 22
    ws_labor.merge_cells("A2:G2")
    ws_labor["A2"] = "Electrical Engineering, Circuit Design, Firmware & Eco-Fi Portal Integration (Rate: ₱500.00 / hour)"
    ws_labor["A2"].font = font_subtitle
    ws_labor["A2"].fill = PatternFill(start_color=emerald_mid, end_color=emerald_mid, fill_type="solid")
    ws_labor["A2"].alignment = align_center

    labor_headers = ["Phase No.", "Engineering Phase / Milestone", "Detailed Tasks, Sub-deliverables & Scope", "Est. Hours", "Hourly Rate (PHP)", "Phase Cost (PHP)", "Primary Discipline"]

    r = 4
    ws_labor.row_dimensions[r].height = 24
    for c_idx, h_text in enumerate(labor_headers, 1):
        cell = ws_labor.cell(row=r, column=c_idx, value=h_text)
        cell.font = font_subhdr
        cell.fill = fill_subhdr
        cell.alignment = align_center
        cell.border = border_cell

    hourly_rate = 500.00

    labor_phases = [
        ("Phase 1", "Electrical Architecture & Power Rail Setup", 
         "• Circuit schematic design, pinout mapping & fuse protection calculations\n• Tuning XL4015 buck converter (5.1V logic rail for Orange Pi One, ESP32 & PCA9685)\n• Wiring 3.3V I2C bus for AS7263 NIR spectrometer & 10k/10k divider for LJ12A3 inductive\n• Assembling AC C14 inlet socket with 10A fuse & master rocker switch", 8.0, hourly_rate, "Electrical Engineering"),
        
        ("Phase 2", "PVC Drop Pathway, AS7263 NIR & 2-Servo Airlock Sub-Assembly", 
         "• Assembling & pre-drilling 4\" PVC drop pipe & 45° elbow for optical sensor mounts\n• Mounting AS7263 NIR spectrometer & LJ12A3 inductive sensor into sealed airlock chamber\n• Wiring PCA9685 16-channel I2C PWM driver board & 2x MG996R servos (Outer & Drop)\n• Assembling electronics chassis plate, terminal blocks & buck converters", 8.0, hourly_rate, "Electrical & Mechatronics Consulting"),
        
        ("Phase 3", "ESP32 FreeRTOS Embedded Firmware Development", 
         "• Implementing high-speed ISR optical beam-break drop detection\n• Developing I2C PCA9685 2-servo airlock state machine (Outer Gate & Internal Trapdoor)\n• AS7263 6-channel NIR spectral calibration (860nm PET peak) & tin lockout logic\n• Implementing JSON UART serialization protocol to stream telemetry to Orange Pi", 10.0, hourly_rate, "Firmware / Embedded C++"),
        
        ("Phase 4", "Orange Pi Armbian & Eco-Fi Portal Engine", 
         "• Flashing Armbian Linux on Orange Pi One & configuring systemd service daemons\n• Deploying branded Eco-Fi Flask captive portal with live bottle deposit modal\n• Setting up dynamic ipset & iptables firewall rules for time-based access\n• Implementing SQLite bottle statistics tracking and anti-tethering (TTL=64) rules", 12.0, hourly_rate, "Linux / Network Software"),
        
        ("Phase 5", "System Integration, Anti-Cheat Testing & QA", 
         "• Full bottle drop testing down the 4\" PVC chute and AS7263 NIR spectral validation\n• Anti-cheat testing (verifying rapid drops, metal can rejection & string pull defense)\n• 24-hour continuous burn-in load test and Wi-Fi throughput validation\n• Builder documentation, wiring diagram handoff, and client briefing", 8.0, hourly_rate, "QA & Systems Engineering")
    ]

    start_labor_r = r + 1
    for p in labor_phases:
        r += 1
        ws_labor.row_dimensions[r].height = 64
        p_no, name, desc, hours, rate, disc = p
        ws_labor.cell(row=r, column=1, value=p_no).alignment = align_center
        ws_labor.cell(row=r, column=2, value=name).alignment = align_left
        ws_labor.cell(row=r, column=3, value=desc).alignment = align_wrap_left
        ws_labor.cell(row=r, column=4, value=hours).alignment = align_center
        ws_labor.cell(row=r, column=4).number_format = "0.0"
        ws_labor.cell(row=r, column=5, value=rate).alignment = align_right
        ws_labor.cell(row=r, column=5).number_format = curr_format
        ws_labor.cell(row=r, column=6, value=f"=D{r}*E{r}").alignment = align_right
        ws_labor.cell(row=r, column=6).number_format = curr_format
        ws_labor.cell(row=r, column=6).font = font_bold
        ws_labor.cell(row=r, column=7, value=disc).alignment = align_left

        for c in range(1, 8):
            cell = ws_labor.cell(row=r, column=c)
            cell.font = font_bold if c in [1, 6] else font_regular
            cell.border = border_cell

    # Total Labor Row
    r += 1
    total_labor_r = r
    ws_labor.row_dimensions[r].height = 26
    ws_labor.merge_cells(f"A{r}:C{r}")
    ws_labor[f"A{r}"] = "TOTAL ESTIMATED ENGINEERING & ASSEMBLY LABOR (BUILDER SCOPE)"
    ws_labor[f"A{r}"].alignment = align_right
    ws_labor[f"A{r}"].font = font_grand_total
    ws_labor.cell(row=r, column=4, value=f"=SUM(D{start_labor_r}:D{r-1})").alignment = align_center
    ws_labor.cell(row=r, column=4).number_format = "0.0"
    ws_labor.cell(row=r, column=5, value="Rate @ ₱500/hr").alignment = align_center
    ws_labor.cell(row=r, column=6, value=f"=SUM(F{start_labor_r}:F{r-1})").alignment = align_right
    ws_labor.cell(row=r, column=6).number_format = curr_format
    ws_labor.cell(row=r, column=6).font = font_grand_total
    ws_labor.cell(row=r, column=7, value="EE & Firmware Scope").alignment = align_center

    for c in range(1, 8):
        cell = ws_labor.cell(row=r, column=c)
        cell.fill = fill_grand_total
        cell.font = font_grand_total
        cell.border = border_grand_total

    # Rate Comparison Benchmarking Section
    r += 2
    ws_labor.row_dimensions[r].height = 26
    ws_labor.merge_cells(f"A{r}:G{r}")
    ws_labor[f"A{r}"] = "MARKET RATE BENCHMARKING (PHILIPPINE ENGINEERING LABOR)"
    ws_labor[f"A{r}"].font = font_section
    ws_labor[f"A{r}"].fill = fill_teal_section
    ws_labor[f"A{r}"].alignment = align_left

    r += 1
    ws_labor.row_dimensions[r].height = 24
    bench_headers = ["Labor Tier", "Experience / Skill Profile", "Hourly Rate (PHP)", "Est. Project Labor (46h)", "Market Assessment", "Target Applicability"]
    for c_idx, h_text in enumerate(bench_headers, 1):
        cell = ws_labor.cell(row=r, column=c_idx if c_idx < 4 else (c_idx+1 if c_idx >= 5 else 4), value=h_text)
        cell.font = font_subhdr
        cell.fill = fill_subhdr
        cell.alignment = align_center
        cell.border = border_cell

    benchmarks = [
        ("Tier 1: General Assembler", "Basic wiring, screw tightening, physical assembly", 120.00, "=C#*46", "Entry-level / Apprentice rate", "Component assembly helper"),
        ("Tier 2: Electronics Technician", "PCB soldering, harness wiring, component testing", 180.00, "=C#*46", "Mid-level technician rate", "Electronics wiring technician"),
        ("Tier 3: Budget Engineering Rate", "Junior engineer / Discounted freelance rate", 250.00, "=C#*46", "Budget client-friendly rate", "Discounted partner projects"),
        ("Tier 4: Standard Freelance EE", "Licensed EE, standard embedded & circuit design", 350.00, "=C#*46", "Standard professional engineering rate", "Commercial client contracts"),
        ("Tier 5: Your Professional Rate", "Senior EE, FreeRTOS Firmware, Linux Gateway, Full-Stack IoT", 500.00, "=C#*46", "(Recommended) Senior Engineering & Turnkey IoT Specialist", "LGU, SK, Commercial & Turnkey Contracts")
    ]

    for b in benchmarks:
        r += 1
        ws_labor.row_dimensions[r].height = 24
        tier, exp, rate_val, formula_templ, assess, appl = b
        ws_labor.cell(row=r, column=1, value=tier).alignment = align_left
        ws_labor.cell(row=r, column=2, value=exp).alignment = align_left
        ws_labor.cell(row=r, column=3, value=rate_val).alignment = align_right
        ws_labor.cell(row=r, column=3).number_format = curr_format
        ws_labor.cell(row=r, column=4, value=f"=C{r}*46").alignment = align_right
        ws_labor.cell(row=r, column=4).number_format = curr_format
        ws_labor.cell(row=r, column=4).font = font_bold
        ws_labor.cell(row=r, column=5, value=assess).alignment = align_left
        ws_labor.merge_cells(f"E{r}:F{r}")
        ws_labor.cell(row=r, column=7, value=appl).alignment = align_left

        for c in range(1, 8):
            cell = ws_labor.cell(row=r, column=c)
            cell.font = font_bold if c in [1, 4] else font_regular
            cell.border = border_cell
        if "Your Professional Rate" in tier:
            for c in range(1, 8):
                ws_labor.cell(row=r, column=c).fill = fill_green_highlight

    # Auto-adjust column widths with extra padding
    labor_widths = [16, 38, 58, 16, 20, 20, 30]
    for idx, width in enumerate(labor_widths, 1):
        ws_labor.column_dimensions[get_column_letter(idx)].width = width

    # Save to target filenames
    filenames = ["Smart_EcoFi_Vendo_BoM_and_Costing.xlsx", "ECO_Fi_Vendo_BoM_and_Costing.xlsx"]
    for fn in filenames:
        try:
            wb.save(fn)
            print(f"Successfully saved: {fn}")
        except PermissionError:
            print(f"Notice: '{fn}' is currently open/locked in another application, skipped.")

if __name__ == "__main__":
    create_econnect_bom_workbook("Smart_EcoFi_Vendo_BoM_and_Costing.xlsx")
